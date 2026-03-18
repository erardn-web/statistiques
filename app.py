import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime, timedelta
import altair as alt

# --- CONFIGURATION PAGE WEB ---
st.set_page_config(page_title="36.9° Analytique", layout="wide", page_icon="📊")

# --- CONSTANTES ET LOGIQUE MÉTIER ---
MOTS_EXCLUSION = {"BERNOIS", "NEUCHATELOIS", "VALAISANS", "GENEVOIS", "VAUDOIS", "FRIBOURGEOIS"}
COULEURS_PROF = {"Physiothérapie": "#00CCFF", "Ergothérapie": "#FF9900", "Massage": "#00CC96", "Autre": "#AB63FA"}

# --- UTILITAIRE PDF ---
def chf(valeur):
    """Formate un nombre en CHF avec apostrophe suisse : 13'340.50 CHF"""
    try:
        entier, decimale = f"{abs(float(valeur)):.2f}".split(".")
        entier_fmt = ""
        for i, c in enumerate(reversed(entier)):
            if i > 0 and i % 3 == 0:
                entier_fmt = "'" + entier_fmt
            entier_fmt = c + entier_fmt
        signe = "-" if float(valeur) < 0 else ""
        return f"{signe}{entier_fmt}.{decimale}"
    except:
        return str(valeur)

def chf_int(valeur):
    """Formate un entier en CHF avec apostrophe suisse : 13'340"""
    try:
        entier = str(int(round(float(valeur))))
        result = ""
        for i, c in enumerate(reversed(entier)):
            if i > 0 and i % 3 == 0:
                result = "'" + result
            result = c + result
        return result
    except:
        return str(valeur)

def nettoyer_code_tarif(val):
    """Convertit un code tarifaire lu comme float en string propre.
    7311.0 → '7311' | 25.11 → '25.110' | 7301.0 → '7301' | déjà string → inchangé"""
    s = str(val).strip()
    # Si c'est un float style "7311.0" → supprimer le .0
    if s.endswith('.0') and s[:-2].isdigit():
        return s[:-2]
    # Si c'est "25.11" → pad à 3 décimales → "25.110"
    if '.' in s:
        entier, dec = s.split('.', 1)
        if entier.isdigit() and dec.isdigit() and len(entier) <= 3:
            return f"{entier}.{dec.ljust(3, '0')}"
    return s

def resoudre_colonnes(df):
    """Détecte les colonnes d'un export Factures Ephysio par leur nom.
    Compatible export mono-thérapeute (20 col) et multi-thérapeutes (23 col).
    Retourne un dict {nom_logique: nom_colonne_réel}."""
    cols = {str(c).strip(): c for c in df.columns}
    cols_lower = {str(c).strip().lower(): c for c in df.columns}

    def trouver(candidats):
        for c in candidats:
            if c.lower() in cols_lower:
                return cols_lower[c.lower()]
        return None

    return {
        # Export Factures
        "date_facture":   trouver(["date"]),
        "loi":            trouver(["loi"]),
        "tp_tg":          trouver(["tp/tg"]),
        "patient":        trouver(["patient"]),
        "medecin":        trouver(["médecin prescripteur", "medecin prescripteur"]),
        "assureur":       trouver(["assurance"]),
        "fournisseur":    trouver(["fournisseur de prestation", "fournisseur"]),
        "statut":         trouver(["statut"]),
        "montant":        trouver(["montant chf", "montant"]),
        "chiffre":        trouver(["chiffre chf", "chiffre"]),
        "date_paiement":  trouver(["date payment"]),
        "montant_paye":   trouver(["montant payment"]),
        "num_patient":    trouver(["#patient"]),
        # Export Prestations (onglet Prestation — stable dans tous les exports)
        "code_tarifaire": trouver(["code tarifaire"]),
        "description":    trouver(["description du tarif", "description"]),
        "quantite":       trouver(["quantité", "quantite"]),
        "nb_points":      trouver(["nombre de points"]),
        "valeur_point":   trouver(["valeur du point"]),
        "therapeute":     trouver(["thérapeute", "therapeute"]),
        "facturation":    trouver(["facturation"]),
    }

def generer_pdf_tableau(titre, df, sous_titre=""):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io as _io_pdf

    buf = _io_pdf.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle('titre', fontSize=14, fontName='Helvetica-Bold',
                                 spaceAfter=6, alignment=TA_CENTER)
    sous_style  = ParagraphStyle('sous',  fontSize=9,  fontName='Helvetica',
                                 spaceAfter=12, textColor=colors.grey, alignment=TA_CENTER)
    elems = [Paragraph(titre, titre_style)]
    if sous_titre:
        elems.append(Paragraph(sous_titre, sous_style))
    elems.append(Spacer(1, 0.3*cm))

    # Construire les données du tableau
    cols = list(df.columns)
    data = [cols] + [[str(v) if v is not None and str(v) != 'nan' else '—'
                      for v in row] for row in df.values]

    # Largeur colonnes auto
    page_w = landscape(A4)[0] - 3*cm
    col_w = page_w / len(cols)
    col_widths = [col_w] * len(cols)

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#1A6B9A')),
        ('TEXTCOLOR',    (0,0), (-1,0),  colors.white),
        ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0),  9),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,1), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4F9')]),
        ('GRID',         (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 0.5*cm))
    elems.append(Paragraph(f"Généré le {datetime.today().strftime('%d.%m.%Y')}", sous_style))
    doc.build(elems)
    buf.seek(0)
    return buf

def generer_pdf_graphique_matplotlib(titre, df, sous_titre="", kind="line", xlabel="", ylabel="CHF"):
    """Génère un PDF contenant un graphique matplotlib à partir d'un DataFrame."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Image as RLImage, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io as _io_pdf
    import tempfile, os

    # Tracer le graphique
    fig, ax = plt.subplots(figsize=(14, 6))
    if kind == "line":
        for col in df.columns:
            ax.plot(df.index, df[col], marker='o', linewidth=2, label=str(col))
    elif kind == "bar":
        df.plot(kind='bar', ax=ax, width=0.7)
    ax.set_title(titre, fontsize=14, fontweight='bold', pad=12)
    ax.set_xlabel(xlabel, fontsize=10)
    ax.set_ylabel(ylabel, fontsize=10)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{chf_int(x)}"))
    ax.legend(loc='upper left', fontsize=8, ncol=min(4, len(df.columns)))
    ax.grid(axis='y', linestyle='--', alpha=0.5)
    plt.xticks(rotation=30, ha='right', fontsize=8)
    plt.tight_layout()

    # Sauvegarder en PNG temporaire
    tmpf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    fig.savefig(tmpf.name, dpi=150, bbox_inches='tight')
    plt.close(fig)

    # Créer le PDF
    buf = _io_pdf.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    titre_style = ParagraphStyle('t', fontSize=13, fontName='Helvetica-Bold', spaceAfter=4, alignment=TA_CENTER)
    sous_style  = ParagraphStyle('s', fontSize=9,  fontName='Helvetica', spaceAfter=10, textColor=colors.grey, alignment=TA_CENTER)
    page_w = landscape(A4)[0] - 3*cm
    elems = [Paragraph(titre, titre_style)]
    if sous_titre:
        elems.append(Paragraph(sous_titre, sous_style))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(RLImage(tmpf.name, width=page_w, height=page_w * 6/14))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(Paragraph(f"Généré le {datetime.today().strftime('%d.%m.%Y')}", sous_style))
    doc.build(elems)
    os.unlink(tmpf.name)
    buf.seek(0)
    return buf

def generer_pdf_plotly(titre, fig, sous_titre=""):
    """Génère un PDF à partir d'un graphique Plotly (nécessite kaleido)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Image as RLImage, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io as _io_pdf
    import tempfile, os

    tmpf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    fig.write_image(tmpf.name, width=1400, height=600, scale=2)

    buf = _io_pdf.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    titre_style = ParagraphStyle('t', fontSize=13, fontName='Helvetica-Bold', spaceAfter=4, alignment=TA_CENTER)
    sous_style  = ParagraphStyle('s', fontSize=9,  fontName='Helvetica', spaceAfter=10, textColor=colors.grey, alignment=TA_CENTER)
    page_w = landscape(A4)[0] - 3*cm
    elems = [Paragraph(titre, titre_style)]
    if sous_titre:
        elems.append(Paragraph(sous_titre, sous_style))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(RLImage(tmpf.name, width=page_w, height=page_w * 6/14))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(Paragraph(f"Généré le {datetime.today().strftime('%d.%m.%Y')}", sous_style))
    doc.build(elems)
    os.unlink(tmpf.name)
    buf.seek(0)
    return buf


def jours_ouvres(date_debut, date_fin, jours_cabinet=None):
    """Nombre de jours où le cabinet était réellement ouvert entre deux dates.
    Si jours_cabinet (set de date) est fourni, on compte les jours avec prestations.
    Sinon, repli sur lun-ven (bdate_range) pour les modules sans ce contexte."""
    if jours_cabinet is not None:
        return max(sum(1 for d in pd.date_range(date_debut, date_fin) if d.date() in jours_cabinet), 1)
    return max(len(pd.bdate_range(date_debut, date_fin)), 1)

def calculer_tendance(ca_60j, ca_365j, jo_60, jo_365):
    """Compare le taux journalier (CHF/jour ouvré) des 60 derniers jours
    vs les 365 derniers jours. Neutre aux vacances, Noël, ponts, etc.
    Seuils : variation > +10% → Hausse, < -10% → Baisse.
    Si pas d'historique sur la période de référence → Nouveau."""
    if ca_60j > 0 and ca_365j == 0:
        return "🆕 Nouveau"
    if ca_60j == 0 and ca_365j == 0:
        return "—"
    if ca_365j > 0 and jo_365 > 0 and jo_60 > 0:
        taux_90  = ca_60j  / jo_60
        taux_365 = ca_365j / jo_365
        variation = (taux_90 - taux_365) / taux_365 * 100
        if variation <= -10: return f"↘️ Baisse ({variation:+.1f}%/j)"
        if variation >=  10: return f"↗️ Hausse ({variation:+.1f}%/j)"
        return f"➡️ Stable ({variation:+.1f}%/j)"
    return "—"

def valider_colonnes(df, nb_min, nom_module):
    """Valide que le DataFrame a assez de colonnes, lève une erreur claire sinon."""
    if len(df.columns) < nb_min:
        raise ValueError(f"[{nom_module}] Le fichier semble incorrect : {len(df.columns)} colonnes trouvées, {nb_min} attendues minimum.")

def assigner_profession(code):
    """Logique métier spécifique au module Tarifs"""
    c = str(code).strip().lower()
    if 'rem' in c: return "Autre"
    if any(x in c for x in ['abo', 'thais']): return "Autre"
    if any(x in c for x in ['privé']) or c.startswith(('73', '25', '15.30')): 
        return "Physiothérapie"
    if any(x in c for x in ['foyer']) or c.startswith(('76', '31', '32')): 
        return "Ergothérapie"
    if c.startswith('1062'): 
        return "Massage"
    return "Autre"

def convertir_date(val):
    """Conversion robuste des dates pour tous les modules"""
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    if isinstance(val, pd.Timestamp): return val
    try:
        return pd.to_datetime(str(val).strip(), format="%d.%m.%Y", errors="coerce")
    except:
        return pd.to_datetime(val, errors="coerce")

def calculer_jours_versement(p_hist):
    """Calcule le jour de versement effectif par assureur.
    Basé sur les 2 derniers mois uniquement pour refléter le comportement récent.
    - Jour dominant >= 50% : on utilise ce jour.
    - Jour dominant < 50% : on décale d'une semaine (conservateur).
    Retourne {assureur: (weekday 0-6, decaler_semaine bool)}
    """
    import pandas as pd
    resultat = {}
    if p_hist.empty:
        return resultat
    # Restreindre aux 2 derniers mois pour capter les changements récents
    date_max = p_hist["date_paiement"].dropna().max()
    limite   = date_max - pd.DateOffset(months=2)
    p_recent = p_hist[p_hist["date_paiement"] >= limite]
    if p_recent.empty:
        p_recent = p_hist  # fallback sur tout l'historique
    jours_fr = {0:"Lun",1:"Mar",2:"Mer",3:"Jeu",4:"Ven",5:"Sam",6:"Dim"}
    for ass, grp in p_recent.groupby("assureur"):
        jours = grp["date_paiement"].dropna().dt.dayofweek
        if jours.empty:
            continue
        counts = jours.value_counts()
        total  = len(jours)
        dominant_j   = counts.idxmax()
        dominant_pct = counts.max() / total
        # Résumé des 2 premiers jours par fréquence décroissante
        top2 = counts.head(2)
        detail = ", ".join(f"{round(n/total*100)}% {jours_fr.get(j,'?')}" for j, n in top2.items())
        resultat[str(ass)] = (int(dominant_j), dominant_pct < 0.50, detail)
    return resultat

def jours_avant_prochain_versement(date_ref, weekday_cible, decaler_semaine=False):
    """Jours de date_ref jusqu'au prochain versement de l'assureur.
    delta=0 = aujourd'hui → inclus (on ne sait pas si le virement a déjà eu lieu).
    decaler_semaine=True : +7 jours supplémentaires (pattern < 50%).
    """
    delta = (weekday_cible - date_ref.weekday()) % 7
    if decaler_semaine:
        delta += 7
    return delta


def calculer_liquidites_fournisseur(f_attente, p_hist, jours_horizons,
                                    jours_versement=None, date_ref=None):
    """Calcul de probabilité de paiement pour le module Facturation.
    3 niveaux de granularité : assureur×fournisseur → fournisseur → global.

    Pour chaque facture ouverte, on calcule son âge à la date cible
    (age_actuel + h), puis on applique P(delai ≤ age_à_horizon) sur
    l'historique de l'assureur×fournisseur.

    Si jours_versement est fourni, on n'inclut une facture dans l'horizon h
    que si le prochain versement de son assureur tombe dans les h jours.
    """
    liq = {h: 0.0 for h in jours_horizons}
    taux_glob = {h: 0.0 for h in jours_horizons}
    if p_hist.empty: return liq, taux_glob

    import numpy as np
    delais_croises = p_hist.groupby(["assureur", "fournisseur"])["delai"].apply(np.array).to_dict()
    delais_fourn   = p_hist.groupby("fournisseur")["delai"].apply(np.array).to_dict()
    delais_global  = p_hist["delai"].values

    for h in jours_horizons:
        taux_glob[h] = (delais_global <= h).mean()
        total_h = 0.0
        for _, row in f_attente.iterrows():
            ass = str(row["assureur"])

            # Correction jour de versement
            if jours_versement is not None and date_ref is not None:
                if ass in jours_versement:
                    weekday_cible, decaler, _ = jours_versement[ass]
                    if jours_avant_prochain_versement(date_ref, weekday_cible, decaler) > h:
                        continue  # assureur ne versera pas dans cet horizon

            age   = int(row.get("delai_actuel", 0))
            seuil = age + h
            key   = (row["assureur"], row["fournisseur"])
            if key in delais_croises:
                d = delais_croises[key]
            elif row["fournisseur"] in delais_fourn:
                d = delais_fourn[row["fournisseur"]]
            else:
                d = delais_global
            prob = (d <= seuil).mean()
            total_h += row["montant"] * prob
        liq[h] = total_h
    return liq, taux_glob


# 👥 MODULE : PILOTAGE FLUX
# ==========================================
def render_stats_patients():
    if st.sidebar.button("⬅️ Retour Accueil", key="btn_back_final"):
        st.session_state.page = "accueil"
        st.rerun()

    st.sidebar.markdown("---")
    st.sidebar.markdown("**📂 Fichier(s) de prestations**")
    uploaded_file  = st.sidebar.file_uploader("Export récent (obligatoire)", type="xlsx", key="uploader_flux_1")
    uploaded_file2 = st.sidebar.file_uploader("Export plus ancien (optionnel, pour étendre l'historique)", type="xlsx", key="uploader_flux_2")
    st.sidebar.markdown("---")
    st.sidebar.markdown("**⚙️ Paramètres**")
    delai_fin_traitement = st.sidebar.number_input(
        "Délai fin de traitement présumé (jours sans séance) :",
        min_value=14, max_value=180, value=60, step=7,
        help="Un patient dont la dernière séance date de plus de N jours est considéré comme terminé. Utilisé pour calculer la moyenne de séances/traitement."
    )
    seuil_jour_flux = st.sidebar.number_input(
        "Montant min. pour jour ouvert (CHF) :",
        min_value=0, max_value=500, value=50, step=10, key="seuil_flux",
        help="Somme minimale facturée sur la journée pour qu'elle soit comptée comme jour ouvré."
    )


    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.title("👥 Pilotage du Flux Patients")

    if not uploaded_file:
        st.info("👋 Chargez au moins le fichier récent pour activer l'analyse.")
        return

    try:
        @st.cache_data
        def lire_prestations(f):
            onglets = pd.ExcelFile(f).sheet_names
            ong = next((s for s in onglets if s.strip().lower() == 'prestation'), None) or                   next((s for s in onglets if 'prestation' in s.lower()), onglets[0])
            df = pd.read_excel(f, sheet_name=ong)
            df.columns = [str(c).strip() for c in df.columns]
            col_code = df.columns[2]
            df[col_code] = df[col_code].apply(nettoyer_code_tarif)
            return df

        @st.cache_data
        def get_full_analysis(file1, file2, delai_fin, seuil_jour):
            def parser(f):
                df = lire_prestations(f)
                _csp = resoudre_colonnes(df)
                c_date  = _csp["date_facture"]
                c_tarif = df.columns[2]   # Code tarifaire — position stable col 2 dans Prestations
                c_pat   = _csp["num_patient"] or df.columns[8]
                c_mont  = _csp["chiffre"] or df.columns[11]
                df[c_date] = pd.to_datetime(df[c_date], errors='coerce')
                df[c_tarif] = df[c_tarif].astype(str).str.strip()
                df[c_mont] = pd.to_numeric(df[c_mont], errors='coerce').fillna(0)
                # CA journalier sur TOUTES les prestations (pour jours ouvrés réels)
                ca_jour = df[df[c_mont] > 0].dropna(subset=[c_date]).copy()
                ca_jour = ca_jour.groupby(ca_jour[c_date].dt.date)[c_mont].sum()
                # Trois flux séparés selon la logique de détection :
                # - 7350 : bilan premier traitement (source principale)
                # - 7301/7311 : garde uniquement pour le rythme et la moyenne séances
                # - 25.110 : première apparition du patient (traitements courts)
                df_pos = df[df[c_mont] > 0].dropna(subset=[c_date, c_pat]).copy()
                df_pos[c_tarif] = df_pos[c_tarif].astype(str).str.strip()

                df_7350 = df_pos[df_pos[c_tarif] == "7350"][[c_date, c_pat]].rename(columns={c_date: "_date", c_pat: "_pat"})
                df_7350["_type"] = "7350"

                df_physio = df_pos[df_pos[c_tarif].isin(["7301", "7311"])][[c_date, c_pat]].rename(columns={c_date: "_date", c_pat: "_pat"})
                df_physio["_type"] = "physio"

                df_25 = df_pos[df_pos[c_tarif] == "25.110"][[c_date, c_pat]].rename(columns={c_date: "_date", c_pat: "_pat"})
                df_25["_type"] = "25.110"

                df_f = pd.concat([df_7350, df_physio, df_25]).drop_duplicates(subset=["_date", "_pat", "_type"])
                return df_f, ca_jour

            df_f, ca_jour = parser(file1)
            if file2 is not None:
                df_f2, ca_jour2 = parser(file2)
                df_f = pd.concat([df_f, df_f2]).drop_duplicates().reset_index(drop=True)
                ca_jour = pd.concat([ca_jour, ca_jour2]).groupby(level=0).sum()
                nb_fichiers = 2
            else:
                nb_fichiers = 1

            # Jours ouvrés réels : jours où le CA total >= seuil_jour
            jours_cabinet_flux = set(ca_jour[ca_jour >= seuil_jour].index)

            df_f = df_f.sort_values("_date")
            derniere_date = df_f["_date"].max()
            premiere_date = df_f["_date"].min()

            # Sous-ensembles par type
            df_physio = df_f[df_f["_type"] == "physio"]
            df_7350   = df_f[df_f["_type"] == "7350"]
            df_25     = df_f[df_f["_type"] == "25.110"]

            # --- 1. DÉCOUPAGE EN ÉPISODES DE TRAITEMENT ---
            # Un épisode = séquence continue de séances 7301/7311 sans pause > PAUSE_TRAITEMENT jours.
            # Le 7350 est refacturé tous les 36 séances ou 6 mois pour des raisons admin,
            # il ne marque donc PAS le début d'un nouveau traitement.
            PAUSE_TRAITEMENT = delai_fin  # paramètre utilisateur (défaut 60j)

            episodes = []
            for pat, grp in df_physio.groupby("_pat"):
                seances = sorted(grp["_date"].tolist())
                debut = seances[0]
                precedente = seances[0]
                count = 1
                for s in seances[1:]:
                    if (s - precedente).days > PAUSE_TRAITEMENT:
                        episodes.append({"_pat": pat, "debut": debut, "fin": precedente, "nb_seances": count})
                        debut = s
                        count = 1
                    else:
                        count += 1
                    precedente = s
                episodes.append({"_pat": pat, "debut": debut, "fin": precedente, "nb_seances": count})

            df_ep = pd.DataFrame(episodes)

            # --- 2. RYTHME HEBDOMADAIRE (semaines actives par épisode) ---
            # Pour chaque épisode, on compte les semaines distinctes avec au moins une séance.
            rythmes_ep = []
            for pat, grp in df_physio.groupby("_pat"):
                seances = sorted(grp["_date"].tolist())
                precedente = seances[0]
                ep_seances = [seances[0]]
                for s in seances[1:]:
                    if (s - precedente).days > PAUSE_TRAITEMENT:
                        if len(ep_seances) >= 2:
                            semaines = pd.Series(ep_seances).dt.isocalendar().apply(
                                lambda r: f"{r['year']}-{r['week']:02d}", axis=1
                            ).nunique()
                            if semaines >= 2:
                                rythmes_ep.append(len(ep_seances) / semaines)
                        ep_seances = [s]
                    else:
                        ep_seances.append(s)
                    precedente = s
                if len(ep_seances) >= 2:
                    semaines = pd.Series(ep_seances).dt.isocalendar().apply(
                        lambda r: f"{r['year']}-{r['week']:02d}", axis=1
                    ).nunique()
                    if semaines >= 2:
                        rythmes_ep.append(len(ep_seances) / semaines)

            rythme = pd.Series(rythmes_ep).mean() if rythmes_ep else 1.1

            # --- 3. CHRONIQUES & MOYENNE SÉANCES/TRAITEMENT ---
            seuil_termine = derniere_date - timedelta(days=delai_fin)
            # Épisodes encore actifs (fin récente = en cours)
            ep_en_cours = df_ep[df_ep["fin"] > seuil_termine]
            ep_termines  = df_ep[df_ep["fin"] <= seuil_termine]

            # Chroniques actifs = épisodes en cours satisfaisant au moins une condition :
            #   1. Présents sur les 365 derniers jours (début <= derniere_date - 365j)
            #   2. Episode avec >=45 séances
            seuil_365_chron = derniere_date - timedelta(days=365)
            chroniques_actifs = ep_en_cours[
                (ep_en_cours["debut"] <= seuil_365_chron) |
                (ep_en_cours["nb_seances"] >= 45)
            ]
            nb_chroniques = len(chroniques_actifs)

            # Cadence hebdomadaire des chroniques (sur les 60 derniers jours de l'export)
            # = nombre moyen de séances/semaine par chronique récemment
            seuil_60 = derniere_date - timedelta(days=60)
            jo_60_flux = jours_ouvres(seuil_60, derniere_date, jours_cabinet_flux)
            semaines_90 = jo_60_flux / 5  # semaines ouvrées
            seances_chroniques_90 = df_physio[
                df_physio["_pat"].isin(chroniques_actifs["_pat"]) &
                (df_physio["_date"] >= seuil_60)
            ]
            rdv_chron_sem = (len(seances_chroniques_90) / semaines_90) if semaines_90 > 0 else 0

            # Moyenne séances/traitement = épisodes terminés NON chroniques
            # + pour les chroniques on utilise une estimation haute (nb séances actuelles)
            # car on ne connaît pas leur fin → on les note séparément
            moy_seances = ep_termines['nb_seances'].mean() if not ep_termines.empty else df_ep['nb_seances'].mean()
            nb_termines = len(ep_termines)

            # p_stats conservé pour la compatibilité (filtre fantômes flux nouveaux)
            p_stats = df_physio.groupby("_pat").agg(
                date_min=("_date", 'min'),
                date_max=("_date", 'max')
            )

            # --- 4. FLUX NOUVEAUX PATIENTS ---
            # Logique par code :
            # - 7350 : bilan premier traitement → nouveau si le patient n'a PAS de séance
            #   7301/7311 dans les <delai_fin> jours AVANT la date du 7350
            #   (évite de compter un patient qui reprend un 2e traitement comme nouveau)
            # - 25.110 : première apparition du patient (traitements courts, pas de biais fantôme)
            seuil_fantomes = premiere_date + timedelta(days=28)

            # Index des séances physio par patient pour tester les antécédents
            physio_dates = df_physio.groupby("_pat")["_date"].apply(list).to_dict()

            def est_vraiment_nouveau(pat, date_bilan, delai):
                # Nouveau si aucune séance 7301/7311 dans les <delai> jours précédant le bilan
                if pat not in physio_dates:
                    return True
                seuil_avant = date_bilan - timedelta(days=delai)
                return not any(seuil_avant <= d < date_bilan for d in physio_dates[pat])

            # Garder un seul événement par patient pour 7350 (le premier bilan vraiment nouveau)
            # Pas de drop_duplicates : un patient peut avoir plusieurs traitements distincts.
            # est_vraiment_nouveau filtre déjà les 7350 d'un épisode en cours.
            nouveaux_7350 = (
                df_7350
                .sort_values("_date")
                .loc[lambda df: df.apply(lambda r: est_vraiment_nouveau(r["_pat"], r["_date"], delai_fin), axis=1)]
            )

            # 25.110 : première séance du patient par traitement distinct
            # On garde la première apparition uniquement (pas de code bilan disponible)
            nouveaux_25 = df_25.sort_values("_date").drop_duplicates(subset=["_pat"], keep="first")

            def stats_periode(jours):
                seuil = derniere_date - timedelta(days=jours)
                jo = jours_ouvres(seuil, derniere_date, jours_cabinet_flux)
                # 7350 dans la fenêtre, hors fantômes
                n_7350 = nouveaux_7350[
                    (nouveaux_7350["_date"] >= seuil) &
                    (nouveaux_7350["_date"] > seuil_fantomes)
                ]
                # 25.110 : première séance dans la fenêtre
                n_25 = nouveaux_25[nouveaux_25["_date"] >= seuil]
                count = len(n_7350) + len(n_25)
                return count, count / jo if jo > 0 else 0

            return {
                "moy_seances": moy_seances,
                "nb_termines": nb_termines,
                "rythme_reel": rythme,
                "flux_30":  stats_periode(30),
                "flux_60":  stats_periode(60),
                "flux_120": stats_periode(120),
                "flux_365": stats_periode(365),
                "derniere_date": derniere_date,
                "premiere_date": premiere_date,
                "nb_fichiers": nb_fichiers,
                "delai_fin": delai_fin,
                "nb_chroniques": nb_chroniques,
                "rdv_chron_sem": rdv_chron_sem,
            }

        data = get_full_analysis(uploaded_file, uploaded_file2, delai_fin_traitement, seuil_jour_flux)

        # --- INFOS EXPORT ---
        periode = f"{data['premiere_date'].strftime('%d.%m.%Y')} → {data['derniere_date'].strftime('%d.%m.%Y')}"
        nb_mois = round((data['derniere_date'] - data['premiere_date']).days / 30.5)
        if data['nb_fichiers'] == 2:
            st.success(f"✅ **2 fichiers fusionnés** — Historique de **{nb_mois} mois** ({periode})")
        else:
            st.info(f"📄 **1 fichier** — Historique de **{nb_mois} mois** ({periode})")

        st.caption(f"Moyenne séances/traitement : **{data['moy_seances']:.1f}** séances (sur {data['nb_termines']} épisodes terminés, pause > {data['delai_fin']}j) | {data['nb_chroniques']} patients chroniques actifs (≥52 séances sans interruption) — leurs places sont déduites de la capacité disponible")

        # --- AFFICHAGE FLUX ---
        st.subheader(f"📈 Recrutement Réel (Calculé au {data['derniere_date'].strftime('%d/%m/%Y')})")
        c_r1, c_r2, c_r3, c_r4 = st.columns(4)
        c_r1.metric("Derniers 30j", f"{data['flux_30'][0]} pat.", f"{data['flux_30'][1]:.2f} / j ouvré")
        c_r2.metric("Derniers 60j", f"{data['flux_60'][0]} pat.", f"{data['flux_60'][1]:.2f} / j ouvré")
        c_r3.metric("Derniers 120j", f"{data['flux_120'][0]} pat.", f"{data['flux_120'][1]:.2f} / j ouvré")
        c_r4.metric("Derniers 365j", f"{data['flux_365'][0]} pat.", f"{data['flux_365'][1]:.2f} / j ouvré")

        # --- FORMULAIRE CONFIGURATION ---
        with st.form("form_v11_1"):
            st.subheader("⚙️ Simulation des besoins (Cabinets A & B)")

            # Charger/sauvegarder config thérapeutes
            import io as _io2
            config_ther_file = st.file_uploader(
                "📥 Charger config_thérapeutes.xlsx",
                type="xlsx", key="config_capa_upload",
                help="Rechargez un fichier exporté précédemment pour pré-remplir le tableau."
            )
            if config_ther_file is not None:
                try:
                    df_loaded = pd.read_excel(config_ther_file)
                    cols_ok = {"Thérapeute", "Cabinet", "Places/Sem", "Semaines/an"}
                    if cols_ok.issubset(set(df_loaded.columns)):
                        st.session_state.capa_df = df_loaded[list(cols_ok)].copy()
                        st.success("✅ Configuration chargée.")
                    else:
                        st.warning("⚠️ Colonnes attendues : Thérapeute, Cabinet, Places/Sem, Semaines/an.")
                except Exception as e:
                    st.error(f"Erreur : {e}")

            if 'capa_df' not in st.session_state:
                st.session_state.capa_df = pd.DataFrame([
                    {"Thérapeute": f"Thérapeute {i}", "Cabinet": "A" if i <= 6 else "B",
                     "Places/Sem": 0, "Semaines/an": 43} for i in range(1, 13)
                ])

            config = {"Cabinet": st.column_config.SelectboxColumn("Cabinet", options=["A", "B"], required=True)}
            edited_df = st.data_editor(st.session_state.capa_df, column_config=config, use_container_width=True)

            col_p1, col_p2 = st.columns(2)
            with col_p1:
                in_seances = st.number_input("Séances / traitement", value=float(round(data['moy_seances'], 1)))
                in_rythme = st.slider("Rythme hebdomadaire", 0.5, 3.0, float(round(data['rythme_reel'], 1)))
            with col_p2:
                in_occup = st.slider("Taux d'occupation visé (%)", 50, 100, 85)
                in_jours = st.slider("Jours d'ouverture / semaine", 1, 6, 5)

            btn_go = st.form_submit_button("🚀 CALCULER ET COMPARER", use_container_width=True, type="primary")
            # Sauvegarder l'état du tableau à chaque rendu pour le download_button hors formulaire
            st.session_state.capa_df_current = edited_df

        # --- BOUTON SAUVEGARDER (hors formulaire, car st.download_button interdit dans st.form) ---
        if 'capa_df_current' in st.session_state:
            import io
            buf = io.BytesIO()
            st.session_state.capa_df_current.to_excel(buf, index=False, engine='openpyxl')
            buf.seek(0)
            st.download_button(
                label="💾 Sauvegarder la configuration (.xlsx)",
                data=buf,
                file_name="config_thérapeutes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Téléchargez ce fichier pour le recharger la prochaine fois sans tout ressaisir."
            )

        if btn_go:
            st.session_state.capa_df = edited_df

            def calc_capa(df_p):
                annuel = (df_p['Places/Sem'] * df_p['Semaines/an']).sum()
                return (annuel * (in_occup/100)) / 52.14

            df_act = edited_df[edited_df['Places/Sem'] > 0]
            c_tot = calc_capa(df_act)
            c_a   = calc_capa(df_act[df_act['Cabinet'] == "A"])
            c_b   = calc_capa(df_act[df_act['Cabinet'] == "B"])

            # Les chroniques sont soustraits une seule fois au niveau global
            # puis répartis proportionnellement entre cabinets
            chron = data['rdv_chron_sem']
            prop_a = c_a / c_tot if c_tot > 0 else 0.5
            prop_b = c_b / c_tot if c_tot > 0 else 0.5

            cd_tot = max(0, c_tot - chron)
            cd_a   = max(0, c_a - chron * prop_a)
            cd_b   = max(0, c_b - chron * prop_b)

            f_tot = (cd_tot * in_rythme) / in_seances
            f_a   = (cd_a   * in_rythme) / in_seances
            f_b   = (cd_b   * in_rythme) / in_seances

            st.markdown("---")

            # Info chroniques
            st.info(
                f"👴 **{data['nb_chroniques']} patients chroniques** actifs occupent en permanence "
                f"**{data['rdv_chron_sem']:.1f} RDV/semaine** — déduits de la capacité disponible."
            )

            t_all, t_a, t_b = st.tabs(["📊 TOTAL GLOBAL", "🏠 CABINET A", "🏠 CABINET B"])

            with t_all:
                besoin_j = f_tot / in_jours
                st.success(f"### Besoin Total : **{besoin_j:.1f}** nouveaux / jour")
                col1, col2, col3 = st.columns(3)
                col1.metric("Capacité totale", f"{c_tot:.1f} RDV/sem")
                col2.metric("Dont chroniques", f"{data['rdv_chron_sem']:.1f} RDV/sem")
                col3.metric("Capacité disponible", f"{cd_tot:.1f} RDV/sem")
                diff = data['flux_60'][1] - besoin_j
                st.metric("Équilibre (Réel 60j vs Théorique)", f"{data['flux_60'][1]:.1f} / jour", delta=round(diff, 1))

            with t_a:
                besoin_j_a = f_a / in_jours
                st.info(f"### Besoin A : **{besoin_j_a:.1f}** nouveaux / jour")
                col1, col2 = st.columns(2)
                col1.metric("Capacité totale A", f"{c_a:.1f} RDV/sem")
                col2.metric("Capacité disponible A", f"{cd_a:.1f} RDV/sem")

            with t_b:
                besoin_j_b = f_b / in_jours
                st.warning(f"### Besoin B : **{besoin_j_b:.1f}** nouveaux / jour")
                col1, col2 = st.columns(2)
                col1.metric("Capacité totale B", f"{c_b:.1f} RDV/sem")
                col2.metric("Capacité disponible B", f"{cd_b:.1f} RDV/sem")

    except Exception as e:
        st.error(f"❌ Erreur : {e}")


# --- INITIALISATION DE L'ÉTAT ---
if 'page' not in st.session_state:
    st.session_state.page = "accueil"
if 'analyse_lancee' not in st.session_state:
    st.session_state.analyse_lancee = False
if 'config_medecins' not in st.session_state:
    st.session_state.config_medecins = {}

# ==========================================
# 🏠 PAGE D'ACCUEIL (STRUCTURÉE PAR SOURCE DE DONNÉES)
# ==========================================
if st.session_state.page == "accueil":

    st.markdown("""
    <style>
    /* Page accueil */
    .accueil-hero {
        background: #F9EEF1;
        border-radius: 16px;
        padding: 20px 28px;
        margin-bottom: 32px;
        color: white;
    }
    .accueil-hero h1 { font-size: 1.8rem; font-weight: 700; margin: 0 0 4px 0; letter-spacing: -0.5px; color: #6D2B3D; }
    .accueil-hero p  { font-size: 0.85rem; color: #B5546A; margin: 0; }

    .section-label {
        font-size: 1.5rem;
        font-weight: 700;
        letter-spacing: 0px;
        text-transform: none;
        color: #B5546A;
        margin-bottom: 10px;
        padding-left: 2px;
    }
    .section-label-orange { color: #6D2B3D; }

    /* Boutons modules */
    div.stButton > button {
        background: #ffffff;
        border: 1.5px solid #e8edf2;
        border-radius: 10px;
        padding: 10px;
        font-size: 1.5rem;
        font-weight: 700;
        color: #6D2B3D;
        height: 120px !important;
        width: 100% !important;
        line-height: 1.3;
        transition: all 0.15s ease;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06);
        white-space: normal !important;
        word-wrap: break-word;
    }
    div.stButton > button:hover {
        border-color: #B5546A;
        background: #FDF3F5;
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(181,84,106,0.15);
    }
    .btn-factures div.stButton > button {
        border-left: 4px solid #6D2B3D;
        height: 120px !important;
    }
    .btn-factures div.stButton > button:hover { border-left-color: #B5546A; }


    </style>
    """, unsafe_allow_html=True)


    # Hero header
    _logo_b64 = "iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC"
    st.markdown(f"""
    <div class="accueil-hero" style="display:flex;align-items:flex-end;gap:24px;">
        <img src="data:image/png;base64,{_logo_b64}" style="height:70px;flex-shrink:0;" />
        <p style="font-size:0.85rem;opacity:0.7;margin:0;">Tableau de bord analytique pour cabinets paramédicaux</p>
    </div>
    """, unsafe_allow_html=True)

    # ── SECTION PRESTATIONS (5 modules, 2 rangées) ───────────────
    st.markdown('<div class="section-label section-label-orange">📑 Export Prestations</div>', unsafe_allow_html=True)

    # Rangée 1 : Médecins · Tarifs · Bilan
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("👨‍⚕️ Médecins", use_container_width=True):
            st.session_state.page = "medecins"
            st.rerun()
    with c2:
        if st.button("🏷️ Tarifs", use_container_width=True):
            st.session_state.page = "tarifs"
            st.rerun()
    with c3:
        if st.button("🏦 Bilan", use_container_width=True):
            st.session_state.page = "bilan"
            st.rerun()

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # Rangée 2 : Stats Patients · Rétrocession · 7350
    c4, c5, c6 = st.columns(3)
    with c4:
        if st.button("👥 Stats Patients", use_container_width=True):
            st.session_state.page = "stats_patients"
            st.rerun()
    with c5:
        if st.button("🤝 Rétrocession", use_container_width=True):
            st.session_state.page = "retrocession"
            st.rerun()
    with c6:
        if st.button("🔁 Position 7350", use_container_width=True):
            st.session_state.page = "pos7350"
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    # ── SECTION FACTURES (1 module) ──────────────────────────────
    st.markdown('<div class="section-label">📂 Export Factures</div>', unsafe_allow_html=True)
    col_f, _ = st.columns([1, 2])
    with col_f:
        st.markdown('<div class="btn-factures">', unsafe_allow_html=True)
        if st.button("📊 Facturation", use_container_width=True):
            st.session_state.page = "factures"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── SECTION GESTION INTERNE ───────────────────────────────────
    st.markdown('<div class="section-label">🏢 Gestion interne</div>', unsafe_allow_html=True)
    col_g, _ = st.columns([1, 2])
    with col_g:
        if st.button("🎓 Formations", use_container_width=True):
            st.session_state.page = "formations"
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

# ==========================================
# 📊 MODULE FACTURES (ORIGINAL RÉPARÉ)
# ==========================================
elif st.session_state.page == "factures":
    if st.sidebar.button("⬅️ Retour Accueil"):
        st.session_state.page = "accueil"
        st.rerun()

    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.title("📊 Analyse de la Facturation")
    uploaded_file = st.sidebar.file_uploader("Charger le fichier Excel (.xlsx)", type="xlsx", key="fact_file")

    if uploaded_file:
        try:
            @st.cache_data
            def lire_factures(f): return pd.read_excel(f, header=0)
            df_brut = lire_factures(uploaded_file)
            valider_colonnes(df_brut, 16, "Factures")
            st.sidebar.header("🔍 2. Filtres")
            _c_tmp = resoudre_colonnes(df_brut)
            if _c_tmp["fournisseur"] is None:
                df_brut["Fournisseur de prestation"] = "Cabinet"
                _c_tmp["fournisseur"] = "Fournisseur de prestation"
            fournisseurs = df_brut[_c_tmp["fournisseur"]].dropna().unique().tolist()
            sel_fournisseurs = st.sidebar.multiselect("Fournisseurs :", options=sorted(fournisseurs), default=fournisseurs)
            lois = df_brut[_c_tmp["loi"]].dropna().unique().tolist()
            sel_lois = st.sidebar.multiselect("Types de Loi :", options=sorted(lois), default=lois)
            regrouper_assureurs = st.sidebar.checkbox("Regrouper par groupe d'assureurs", value=False,
                help="Fusionne les assureurs appartenant au même groupe (ex. Le Groupe Mutuel + Philos → Groupe Mutuel)")
            st.sidebar.header("📅 3. Périodes & Simulation")
            options_p = {"Global": None, "6 mois": 6, "4 mois": 4, "3 mois": 3, "2 mois": 2, "1 mois": 1}
            periods_sel = st.sidebar.multiselect("Analyser les périodes :", list(options_p.keys()), default=["Global", "4 mois", "2 mois"])
            if st.sidebar.button("🚀 Analyser", type="primary", use_container_width=True):
                st.session_state.analyse_lancee = True
            st.sidebar.markdown("---")
            date_cible = st.sidebar.date_input("Date cible (simulation) :", value=datetime.today())
            btn_simuler = st.sidebar.button("🔮 Simuler", use_container_width=True)
            corriger_jours = st.sidebar.checkbox("🗓️ Correction par jour de versement", value=False,
                help="Exclut les assureurs dont le jour de paiement habituel ne tombe pas dans l'horizon simulé.")
            ass_disponibles_sim = sorted(df_brut[resoudre_colonnes(df_brut)["assureur"] or df_brut.columns[8]].dropna().unique().tolist()) if df_brut is not None else []

            _c = resoudre_colonnes(df_brut)
            # Fournisseur absent en mono-thérapeute → colonne virtuelle
            if _c["fournisseur"] is None:
                df_brut["Fournisseur de prestation"] = "Cabinet"
                _c["fournisseur"] = "Fournisseur de prestation"

            df = df_brut[
                (df_brut[_c["fournisseur"]].isin(sel_fournisseurs)) &
                (df_brut[_c["loi"]].isin(sel_lois))
            ].copy()
            df = df.rename(columns={
                _c["date_facture"]: "date_facture", _c["loi"]: "loi",
                _c["assureur"]: "assureur", _c["fournisseur"]: "fournisseur",
                _c["statut"]: "statut", _c["montant"]: "montant",
                _c["date_paiement"]: "date_paiement"
            })
            
            df["date_facture"] = df["date_facture"].apply(convertir_date)
            df["date_paiement"] = df["date_paiement"].apply(convertir_date)
            df = df[df["date_facture"].notna()].copy()
            df["montant"] = pd.to_numeric(df["montant"], errors="coerce").fillna(0)
            df["statut"] = df["statut"].astype(str).str.lower().str.strip()
            df["assureur"] = df["assureur"].fillna("Patient")
            # LCA : remboursement direct par le patient → assureur = "Patient"
            df.loc[df["loi"] == "LCA", "assureur"] = "Patient"

            # --- GROUPES D'ASSUREURS SUISSES ---
            # Mapping : nom exact dans Ephysio → nom du groupe affiché
            # LCA et LAI exclues du regroupement
            GROUPES_NOM = {
                # LAMal — Groupe Mutuel
                "Philos, caisse maladie":            "Groupe Mutuel, caisse maladie",
                "Caisse maladie Avenir":              "Groupe Mutuel, caisse maladie",
                "Easy Sana caisse maladie":           "Groupe Mutuel, caisse maladie",
                "SUPRA-1846 SA":                      "Groupe Mutuel, caisse maladie",
                # LAMal — CSS
                "Arcosana":                           "CSS Assurances",
                "Intras, caisse maladie":              "CSS Assurances",
                # LAMal — Helsana
                "Progrès (incl. Sansan)":              "Helsana Assurances",
                # LAMal — Visana
                "sana24":                              "Visana Services AG",
                "vivacare":                            "Visana Services AG",
                "GALENOS":                             "Visana Services AG",
                # LAA — Groupe Mutuel
                "Caisse maladie Avenir (accident)":    "Groupe Mutuel, caisse maladie (accident)",
            }
            LOI_EXCLUES_REGROUPEMENT = {"LCA", "LAI"}

            if regrouper_assureurs:
                def appliquer_groupe(row):
                    if row["loi"] in LOI_EXCLUES_REGROUPEMENT:
                        return row["assureur"]
                    return GROUPES_NOM.get(str(row["assureur"]).strip(), str(row["assureur"]).strip())
                df["assureur"] = df.apply(appliquer_groupe, axis=1)
                st.sidebar.caption("✅ Regroupement actif — LCA et LAI non fusionnées.")

            ajd = pd.Timestamp(datetime.today().date())
            f_att = df[df["statut"].str.startswith("en attente") & (df["statut"] != "en attente (annulé)")].copy()
            f_att["delai_actuel"] = (ajd - f_att["date_facture"]).dt.days
            st.metric("💰 TOTAL BRUT EN ATTENTE", f"{chf(f_att['montant'].sum())} CHF")

            # Facteur pessimiste : exclure les factures > 35 jours des projections de liquidités
            f_att_liq = f_att[f_att["delai_actuel"] <= 35].copy()
            f_att_old = f_att[f_att["delai_actuel"] > 35].copy()
            if not f_att_old.empty:
                st.caption(f"⚠️ {len(f_att_old)} facture(s) de plus de 35 jours exclues des projections ({chf_int(round(f_att_old['montant'].sum()))} CHF) — à traiter manuellement.")

            # Filtre assureurs (persistant)
            ass_dispo_sim = sorted(f_att_liq["assureur"].unique().tolist())
            sel_ass_sim = st.multiselect(
                "Filtrer par assureur (simulation) :", ass_dispo_sim,
                default=ass_dispo_sim, key="sel_ass_sim"
            )

            def _run_simulation(date_cible, periods_sel, options_p, df, f_att_liq,
                                sel_ass_sim, ass_dispo_sim, corriger_jours, ajd):
                """Calcule la simulation et retourne le dict résultat."""
                ts_cible = pd.Timestamp(date_cible)
                jour_semaine = ts_cible.weekday()
                if jour_semaine == 5:
                    ts_effective = ts_cible - pd.DateOffset(days=1)
                    note_weekend = f"⚠️ Samedi — les versements tombent le vendredi. Résultat calculé au **{ts_effective.strftime('%d.%m.%Y')}** (vendredi)."
                elif jour_semaine == 6:
                    ts_effective = ts_cible - pd.DateOffset(days=2)
                    note_weekend = f"⚠️ Dimanche — les versements tombent le vendredi. Résultat calculé au **{ts_effective.strftime('%d.%m.%Y')}** (vendredi)."
                else:
                    ts_effective = ts_cible
                    note_weekend = None
                jours_delta = (ts_effective - ajd).days
                if jours_delta < 0:
                    return None
                f_att_sim = f_att_liq[f_att_liq["assureur"].isin(sel_ass_sim)].copy()
                res_sim = []
                for p_nom in periods_sel:
                    val = options_p[p_nom]
                    limit = ajd - pd.DateOffset(months=val) if val else df["date_facture"].min()
                    p_hist_sim = df[(df["date_paiement"].notna()) & (df["date_facture"] >= limit)].copy()
                    p_hist_sim["delai"] = (p_hist_sim["date_paiement"] - p_hist_sim["date_facture"]).dt.days
                    p_hist_sim = p_hist_sim[(p_hist_sim["assureur"] != "Patient") & (p_hist_sim["delai"] >= 1)]
                    jv_sim = calculer_jours_versement(p_hist_sim) if corriger_jours else None
                    liq, _ = calculer_liquidites_fournisseur(f_att_sim, p_hist_sim, [jours_delta],
                                                              jours_versement=jv_sim, date_ref=ajd)
                    res_sim.append({"Période": p_nom, "Estimation (CHF)": f"{chf_int(round(liq[jours_delta]))}",
                                    "Assureurs": f"{len(sel_ass_sim)}/{len(ass_dispo_sim)}"})
                return {
                    "date": ts_cible.strftime("%d.%m.%Y"),
                    "jours": (ts_cible - ajd).days,
                    "note_weekend": note_weekend,
                    "res_sim": res_sim,
                    "date_cible_raw": str(date_cible),
                    "sel_ass_sim": sel_ass_sim,
                    "corriger_jours": corriger_jours,
                }

            # Lancer si bouton cliqué
            if btn_simuler:
                result = _run_simulation(date_cible, periods_sel, options_p, df, f_att_liq,
                                         sel_ass_sim, ass_dispo_sim, corriger_jours, ajd)
                if result:
                    st.session_state["sim_result"] = result

            # Recalculer si le filtre assureurs a changé (simulation déjà lancée)
            elif "sim_result" in st.session_state:
                sr = st.session_state["sim_result"]
                if sr.get("sel_ass_sim") != sel_ass_sim or sr.get("corriger_jours") != corriger_jours:
                    result = _run_simulation(date_cible, periods_sel, options_p, df, f_att_liq,
                                             sel_ass_sim, ass_dispo_sim, corriger_jours, ajd)
                    if result:
                        st.session_state["sim_result"] = result

            # Affichage persistant
            if "sim_result" in st.session_state:
                sr = st.session_state["sim_result"]
                st.markdown(f"**🔮 Simulation au {sr['date']}** — dans {sr['jours']} jour{'s' if sr['jours'] > 1 else ''}")
                if sr["note_weekend"]:
                    st.caption(sr["note_weekend"])
                df_sim = pd.DataFrame(sr["res_sim"])
                rows_html = "".join(f"<tr>{''.join(f'<td style=padding:6px 12px;border-bottom:1px solid #e0e0e0;>{v}</td>' for v in r)}</tr>" for r in df_sim.values)
                headers_html = "".join(f"<th style='padding:6px 12px;border-bottom:2px solid #b0c4d8;background:#AED6F1;text-align:left;font-weight:600;'>{c}</th>" for c in df_sim.columns)
                st.markdown(f"<table style='background:#D6EAF8;border-collapse:collapse;width:auto;font-size:0.9rem;'><thead><tr>{headers_html}</tr></thead><tbody>{rows_html}</tbody></table>", unsafe_allow_html=True)
                if st.button("✖ Effacer la simulation", key="clear_sim"):
                    del st.session_state["sim_result"]
                    st.rerun()

            if st.session_state.analyse_lancee:
                tab1, tab_open, tab_age, tab2, tab3, tab4 = st.tabs(["💰 Liquidités", "📋 Factures ouvertes", "📊 Âge des factures", "🕒 Délais", "⚠️ Retards", "📈 Évolution"])
                with tab2:
                    col_opt1, col_opt2, _ = st.columns([1, 1, 3])
                    show_med = col_opt1.checkbox("Médiane", value=True, key="show_med")
                    show_std = col_opt2.checkbox("Écart-type", value=True, key="show_std")
                for p_name in periods_sel:
                    val = options_p[p_name]
                    limit_p = ajd - pd.DateOffset(months=val) if val else df["date_facture"].min()
                    df_p = df[df["date_facture"] >= limit_p]
                    p_hist = df_p[df_p["date_paiement"].notna()].copy()
                    p_hist["delai"] = (p_hist["date_paiement"] - p_hist["date_facture"]).dt.days
                    # Exclure paiements directs patients (délai = 0) et délais négatifs (erreurs de saisie)
                    p_hist = p_hist[(p_hist["assureur"] != "Patient") & (p_hist["delai"] >= 1)]
                    jv = calculer_jours_versement(p_hist) if corriger_jours else None
                    with tab1:
                        st.subheader(f"Liquidités : {p_name}")
                        horizons = [10, 20, 30]
                        liq, t = calculer_liquidites_fournisseur(f_att_liq, p_hist, horizons,
                                                                  jours_versement=None, date_ref=ajd)
                        st.table(pd.DataFrame({"Horizon": [f"Sous {h}j" for h in horizons], "Estimation (CHF)": [f"{chf_int(round(liq[h]))}" for h in horizons]}))

                        if corriger_jours and jv:
                            jours_fr = {0:"Lun",1:"Mar",2:"Mer",3:"Jeu",4:"Ven",5:"Sam",6:"Dim"}
                            lignes = []
                            for ass_n, (wd, dec, detail) in sorted(jv.items()):
                                delta = jours_avant_prochain_versement(ajd, wd, dec)
                                lignes.append({
                                    "Assureur": ass_n,
                                    "Pattern (2 derniers mois)": detail,
                                    "Fiable": "✅" if not dec else "⚠️ <50%",
                                    "Prochain versement dans": f"{delta}j"
                                })
                            with st.expander("🗓️ Jours de versement détectés"):
                                st.dataframe(pd.DataFrame(lignes), use_container_width=True, hide_index=True)
                    with tab2:
                        st.subheader(f"Délais par assureur ({p_name})")
                        if not p_hist.empty:
                            stats = p_hist.groupby("assureur")["delai"].agg(
                                mean='mean', median='median', std='std', count='count'
                            ).reset_index()
                            stats.columns = ["Assureur", "Moyenne (j)", "Médiane (j)", "Écart-type (j)", "Nb factures"]
                            # Arrondir à 2 décimales
                            stats["Moyenne (j)"]    = stats["Moyenne (j)"].round(2)
                            stats["Médiane (j)"]    = stats["Médiane (j)"].round(2)
                            # Écart-type non significatif sous 5 factures → préfixer NS
                            stats["Écart-type (j)"] = stats.apply(
                                lambda r: f"NS {r['Écart-type (j)']:.2f}" if r["Nb factures"] < 5 and pd.notna(r["Écart-type (j)"])
                                else (round(r["Écart-type (j)"], 2) if pd.notna(r["Écart-type (j)"]) else r["Écart-type (j)"]),
                                axis=1
                            )
                            cols_to_show = ["Assureur", "Nb factures", "Moyenne (j)"]
                            if show_med: cols_to_show.append("Médiane (j)")
                            if show_std: cols_to_show.append("Écart-type (j)")
                            df_styled = stats[cols_to_show].sort_values("Moyenne (j)", ascending=False)
                            def colorier_ns(val):
                                if isinstance(val, str) and val.startswith("NS"):
                                    return "color: red; font-weight: bold"
                                return ""
                            st.dataframe(
                                df_styled.style.applymap(colorier_ns, subset=["Écart-type (j)"]) if show_std else df_styled,
                                use_container_width=True,
                                column_config={
                                    "Moyenne (j)":    st.column_config.NumberColumn(format="%.2f"),
                                    "Médiane (j)":    st.column_config.NumberColumn(format="%.2f"),
                                    "Écart-type (j)": st.column_config.TextColumn(help="NS = Non-significatif (< 5 factures)")
                                }
                            )
                            _pdf_buf = generer_pdf_tableau(f"Délais par assureur — {p_name}", df_styled, f"Période : {p_name}")
                            st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name=f"delais_{p_name}.pdf", mime="application/pdf", key=f"pdf_delais_{p_name}", use_container_width=True)
                    with tab3:
                        st.subheader(f"Analyse des retards > 30j ({p_name})")
                        df_att_30 = f_att[f_att["delai_actuel"] > 30].copy()
                        df_pay_30 = p_hist[p_hist["delai"] > 30].copy()
                        plus_30 = pd.concat([df_pay_30, df_att_30])
                        total_vol = df_p.groupby("assureur").size().reset_index(name="Volume Total")
                        ret_assur = plus_30.groupby("assureur").size().reset_index(name="Nb Retards")
                        merged = pd.merge(ret_assur, total_vol, on="assureur", how="right").fillna(0)
                        merged["Nb Retards"] = merged["Nb Retards"].astype(int)
                        merged["% Retard"] = (merged["Nb Retards"] / merged["Volume Total"] * 100).round(1)
                        st.metric(f"Total Retards ({p_name})", f"{int(merged['Nb Retards'].sum())} factures")
                        st.dataframe(merged[["assureur", "Nb Retards", "Volume Total", "% Retard"]].sort_values("% Retard", ascending=False), use_container_width=True)
                        _pdf_buf = generer_pdf_tableau(f"Retards > 30j — {p_name}", merged[["assureur", "Nb Retards", "Volume Total", "% Retard"]].sort_values("% Retard", ascending=False), f"Période : {p_name}")
                        st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name=f"retards_{p_name}.pdf", mime="application/pdf", key=f"pdf_retards_{p_name}", use_container_width=True)
                
                with tab4:
                    st.subheader("📈 Évolution du délai de remboursement")
                    ordre_chrono = ["Global", "6 mois", "4 mois", "3 mois", "2 mois"]
                    periodes_graph = {"Global": None, "6 mois": 6, "4 mois": 4, "3 mois": 3, "2 mois": 2}
                    evol_data = []
                    p_hist_global = df[df["date_paiement"].notna()].copy()
                    p_hist_global["delai"] = (p_hist_global["date_paiement"] - p_hist_global["date_facture"]).dt.days
                    # Classement global par CA total (base pour les tops)
                    ranking_assureurs = p_hist_global.groupby("assureur")["montant"].sum().sort_values(ascending=False)
                    tous_assureurs = ranking_assureurs.index.tolist()

                    for n, v in periodes_graph.items():
                        lim = ajd - pd.DateOffset(months=v) if v else df["date_facture"].min()
                        h_tmp = df[(df["date_paiement"].notna()) & (df["date_facture"] >= lim)].copy()
                        h_tmp["delai"] = (h_tmp["date_paiement"] - h_tmp["date_facture"]).dt.days
                        if not h_tmp.empty:
                            m = h_tmp.groupby("assureur")["delai"].mean().round(2).reset_index()
                            m["Période"] = n
                            evol_data.append(m)

                    if evol_data:
                        df_ev = pd.concat(evol_data)
                        df_pv = df_ev.pivot(index="assureur", columns="Période", values="delai")
                        cols_presentes = [c for c in ordre_chrono if c in df_pv.columns]
                        df_pv = df_pv[cols_presentes]

                        # --- Sélecteur de Top ---
                        options_top = {"Top 5": 5, "Top 10": 10, "Top 20": 20, "Global": None}
                        col_top, col_spacer = st.columns([1, 3])
                        with col_top:
                            top_choix = st.selectbox("Afficher :", list(options_top.keys()), index=0, key="evol_top")
                        nb_top = options_top[top_choix]
                        # Ranking basé sur le CA dans df_pv (assureurs ayant des données)
                        assureurs_disponibles = [a for a in tous_assureurs if a in df_pv.index]
                        defaut_sel = assureurs_disponibles[:nb_top] if nb_top else assureurs_disponibles

                        # Le key change selon le top choisi → force Streamlit à re-rendre
                        # le multiselect avec le bon default à chaque changement de top
                        # La clé intègre les lois et fournisseurs actifs → reset automatique si filtres changent
                        _filtre_key = "_".join(sorted(sel_lois)) + "_" + "_".join(sorted(sel_fournisseurs))
                        assur_sel = st.multiselect(
                            "Sélectionner les assureurs :",
                            options=df_pv.index.tolist(),
                            default=defaut_sel,
                            key=f"evol_assureurs_{top_choix}_{_filtre_key}"
                        )
                        if assur_sel:
                            df_plot = df_pv.loc[assur_sel].T
                            df_plot.index = pd.CategoricalIndex(df_plot.index, categories=ordre_chrono, ordered=True)
                            df_plot_sorted = df_plot.sort_index()

                            import plotly.graph_objects as go
                            fig_ev = go.Figure()
                            colors_ev = ['#636EFA','#EF553B','#00CC96','#AB63FA','#FFA15A',
                                         '#19D3F3','#FF6692','#B6E880','#FF97FF','#FECB52']
                            for i, assureur in enumerate(df_plot_sorted.columns):
                                vals = df_plot_sorted[assureur]
                                # Exclure les NaN — ne pas afficher comme 0
                                x_vals = [str(x) for x, v in zip(vals.index, vals) if pd.notna(v)]
                                y_vals = [v for v in vals if pd.notna(v)]
                                fig_ev.add_trace(go.Scatter(
                                    name=str(assureur), x=x_vals, y=y_vals,
                                    mode='lines+markers',
                                    line=dict(color=colors_ev[i % len(colors_ev)], width=2),
                                    marker=dict(size=7),
                                    connectgaps=False
                                ))
                            fig_ev.update_layout(
                                xaxis_title="Période",
                                yaxis_title="Délai moyen (jours)",
                                legend=dict(orientation="v", x=1.01, y=1),
                                height=450,
                                margin=dict(l=60, r=20, t=30, b=40),
                            )
                            st.plotly_chart(fig_ev, use_container_width=True)
                            st.dataframe(df_pv.loc[assur_sel].style.highlight_max(axis=1, color='#ff9999').highlight_min(axis=1, color='#99ff99'))
                            try:
                                # PDF combiné : graphique + tableau
                                import matplotlib
                                matplotlib.use('Agg')
                                import matplotlib.pyplot as plt
                                import matplotlib.ticker as mticker
                                from reportlab.lib.pagesizes import A4, landscape
                                from reportlab.lib import colors
                                from reportlab.lib.units import cm
                                from reportlab.platypus import SimpleDocTemplate, Image as RLImage, Paragraph, Spacer, Table, TableStyle
                                from reportlab.lib.styles import ParagraphStyle
                                from reportlab.lib.enums import TA_CENTER
                                import io as _io_pdf, tempfile, os

                                fig_mpl, ax = plt.subplots(figsize=(14, 5))
                                for col in df_plot_sorted.columns:
                                    ax.plot(df_plot_sorted.index.astype(str), df_plot_sorted[col], marker='o', linewidth=2, label=str(col))
                                ax.set_title("Évolution des délais de paiement", fontsize=13, fontweight='bold', pad=10)
                                ax.set_ylabel("Délai moyen (jours)", fontsize=10)
                                ax.legend(loc='upper left', fontsize=8, ncol=min(4, len(df_plot_sorted.columns)))
                                ax.grid(axis='y', linestyle='--', alpha=0.5)
                                plt.xticks(rotation=20, ha='right', fontsize=8)
                                plt.tight_layout()
                                tmpf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                                fig_mpl.savefig(tmpf.name, dpi=150, bbox_inches='tight')
                                plt.close(fig_mpl)

                                buf_pdf = _io_pdf.BytesIO()
                                page_w = landscape(A4)[0] - 3*cm
                                doc = SimpleDocTemplate(buf_pdf, pagesize=landscape(A4),
                                                        leftMargin=1.5*cm, rightMargin=1.5*cm,
                                                        topMargin=1.5*cm, bottomMargin=1.5*cm)
                                t_style = ParagraphStyle('t', fontSize=13, fontName='Helvetica-Bold', spaceAfter=4, alignment=TA_CENTER)
                                s_style = ParagraphStyle('s', fontSize=9, fontName='Helvetica', spaceAfter=8, textColor=colors.grey, alignment=TA_CENTER)
                                elems = [Paragraph("Évolution des délais de paiement", t_style),
                                         Paragraph(f"Généré le {datetime.today().strftime('%d.%m.%Y')}", s_style),
                                         Spacer(1, 0.2*cm),
                                         RLImage(tmpf.name, width=page_w, height=page_w * 5/14),
                                         Spacer(1, 0.4*cm),
                                         Paragraph("Tableau détaillé", t_style),
                                         Spacer(1, 0.2*cm)]
                                # Tableau
                                df_tab = df_pv.loc[assur_sel].copy()
                                df_tab = df_tab[[c for c in ordre_chrono if c in df_tab.columns]]
                                df_tab = df_tab.round(1)
                                cols_t = ["Assureur"] + list(df_tab.columns)
                                data_t = [cols_t] + [[str(idx)] + [str(v) for v in row] for idx, row in zip(df_tab.index, df_tab.values)]
                                col_w = page_w / len(cols_t)
                                tbl = Table(data_t, colWidths=[col_w]*len(cols_t), repeatRows=1)
                                tbl.setStyle(TableStyle([
                                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A6B9A')),
                                    ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                                    ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                                    ('FONTSIZE',   (0,0), (-1,-1), 8),
                                    ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
                                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4F9')]),
                                    ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
                                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                                    ('TOPPADDING',    (0,0), (-1,-1), 4),
                                ]))
                                elems.append(tbl)
                                doc.build(elems)
                                os.unlink(tmpf.name)
                                buf_pdf.seek(0)
                                st.download_button("📄 Télécharger graphique + tableau (PDF)", buf_pdf, file_name="evolution_delais.pdf", mime="application/pdf", key="pdf_evol_graph", use_container_width=True)
                            except Exception as _e:
                                st.caption(f"Export PDF indisponible : {_e}")
                # Onglet factures ouvertes — indépendant des périodes
                with tab_open:
                    st.subheader("Factures ouvertes par assureur")
                    if not f_att.empty:
                        # Délai moyen sur les 2 derniers mois (df a déjà le regroupement appliqué)
                        p_hist_2m = df[(df["date_paiement"].notna()) & (df["date_facture"] >= ajd - pd.DateOffset(months=2))].copy()
                        p_hist_2m["delai"] = (p_hist_2m["date_paiement"] - p_hist_2m["date_facture"]).dt.days
                        delai_moy = p_hist_2m.groupby("assureur")["delai"].mean().round(0).astype("Int64").to_dict()
                        ouv = f_att.groupby("assureur").agg(
                            nb=("montant", "count"),
                            total=("montant", "sum")
                        ).reset_index()
                        ouv["Délai moyen (j)"] = ouv["assureur"].map(delai_moy)
                        ouv.columns = ["Assureur", "Nb factures", "Montant (CHF)", "Délai moyen (j)"]
                        ouv = ouv.sort_values("Montant (CHF)", ascending=False)
                        ouv["Montant (CHF)"] = ouv["Montant (CHF)"].apply(lambda x: chf_int(round(x)))
                        total_nb = ouv["Nb factures"].sum()
                        total_chf = f_att["montant"].sum()
                        ouv.loc[len(ouv)] = ["TOTAL", total_nb, chf_int(round(total_chf)), ""]
                        st.table(ouv)
                    else:
                        st.info("Aucune facture ouverte.")

                with tab_age:
                    if not f_att.empty:
                        import plotly.graph_objects as go

                        tranches = [
                            (">45j",  46, 9999),
                            ("41–45j",41,  45),
                            ("36–40j",36,  40),
                            ("31–35j",31,  35),
                            ("26–30j",26,  30),
                            ("21–25j",21,  25),
                            ("16–20j",16,  20),
                            ("11–15j",11,  15),
                            ("6–10j",  6,  10),
                            ("0–5j",   0,   5),
                        ]
                        def tranche_age(age):
                            for label, lo, hi in tranches:
                                if lo <= age <= hi:
                                    return label
                            return ">60j"

                        f_att2 = f_att.copy()
                        f_att2["tranche"] = f_att2["delai_actuel"].apply(tranche_age)

                        ranking = f_att2.groupby("assureur")["montant"].sum().sort_values(ascending=False)
                        tous_assureurs_age = ranking.index.tolist()

                        col_type, col_top, col_spacer = st.columns([1, 1, 3])
                        chart_type = col_type.selectbox("Type", ["Colonnes", "Courbes"], key="age_chart_type")
                        top_opt = col_top.selectbox("Présélection", ["Top 5", "Top 10", "Top 20", "Tous"], key="age_top")
                        top_n = {"Top 5": 5, "Top 10": 10, "Top 20": 20, "Tous": None}[top_opt]
                        defaut_age = tous_assureurs_age[:top_n] if top_n else tous_assureurs_age

                        assur_sel_age = st.multiselect(
                            "Sélectionner les assureurs :",
                            options=tous_assureurs_age,
                            default=defaut_age,
                            key=f"age_assur_sel_{top_opt}"
                        )

                        if assur_sel_age:
                            pivot = f_att2.groupby(["tranche", "assureur"])["montant"].sum().unstack(fill_value=0)
                            ordre = [t[0] for t in tranches if t[0] in pivot.index]
                            pivot = pivot.loc[ordre]
                            cols_present = [a for a in assur_sel_age if a in pivot.columns]
                            pivot = pivot[cols_present]

                            fig = go.Figure()
                            colors = ['#636EFA','#EF553B','#00CC96','#AB63FA','#FFA15A',
                                      '#19D3F3','#FF6692','#B6E880','#FF97FF','#FECB52']
                            for i, assureur in enumerate(pivot.columns):
                                vals = pivot[assureur].values
                                color = colors[i % len(colors)]
                                if chart_type == "Colonnes":
                                    fig.add_trace(go.Bar(name=assureur, x=list(pivot.index), y=list(vals), marker_color=color))
                                else:
                                    fig.add_trace(go.Scatter(name=assureur, x=list(pivot.index), y=list(vals),
                                        mode='lines+markers', line=dict(color=color, width=2), marker=dict(size=7)))

                            fig.update_layout(
                                barmode='stack' if chart_type == "Colonnes" else None,
                                xaxis_title="Tranche d'âge",
                                yaxis_title="CA (CHF)",
                                yaxis=dict(tickformat=",.0f"),
                                legend=dict(orientation="v", x=1.01, y=1),
                                height=480,
                                margin=dict(l=60, r=20, t=30, b=40),
                            )
                            st.plotly_chart(fig, use_container_width=True)

                            # Descriptions textuelles
                            st.markdown("---")
                            st.subheader("Profil de paiement par assureur — 2 derniers mois")
                            p_hist_desc = df[
                                (df["date_paiement"].notna()) &
                                (df["date_facture"] >= ajd - pd.DateOffset(months=2))
                            ].copy()
                            p_hist_desc["delai"] = (p_hist_desc["date_paiement"] - p_hist_desc["date_facture"]).dt.days

                            # Historique global pour l'appréciation experte
                            p_hist_global = df[df["date_paiement"].notna()].copy()
                            p_hist_global["delai"] = (p_hist_global["date_paiement"] - p_hist_global["date_facture"]).dt.days
                            delai_moyen_global = p_hist_global["delai"].mean()

                            for assureur in assur_sel_age:
                                d = p_hist_desc[p_hist_desc["assureur"] == assureur]["delai"]
                                if len(d) < 3:
                                    st.markdown(f"**{assureur}** — historique insuffisant pour établir un profil.")
                                    continue
                                n = len(d)
                                p_10  = (d <= 10).mean()
                                p_20  = ((d > 10) & (d <= 20)).mean()
                                p_30  = ((d > 20) & (d <= 30)).mean()
                                p_ret = (d > 30).mean()

                                # Appréciation basée sur l'historique global
                                d_glob = p_hist_global[p_hist_global["assureur"] == assureur]["delai"]
                                delai_moy_ass = d_glob.mean() if len(d_glob) >= 3 else d.mean()
                                delai_med_ass = d_glob.median() if len(d_glob) >= 3 else d.median()
                                p_ret_glob = (d_glob > 30).mean() if len(d_glob) >= 3 else p_ret

                                # Ligne factuelle
                                parties = []
                                if p_10  > 0.01: parties.append(f"**{round(p_10*100)}%** en moins de 10 jours")
                                if p_20  > 0.01: parties.append(f"**{round(p_20*100)}%** entre 10 et 20 jours")
                                if p_30  > 0.01: parties.append(f"**{round(p_30*100)}%** entre 20 et 30 jours")
                                if p_ret > 0.01: parties.append(f"**{round(p_ret*100)}%** en retard (>30 jours)")
                                desc = ", ".join(parties) if parties else "profil non établi"

                                # Note experte
                                notes = []
                                # Classification basée sur le délai moyen des 2 derniers mois
                                delai_moy_recent = d.mean() if len(d) >= 3 else delai_moy_ass
                                if delai_moy_recent <= 15:
                                    notes.append("⭐⭐⭐ **Très bon payeur** — règle ses factures en moins de 15 jours en général.")
                                elif delai_moy_recent <= 25:
                                    notes.append("⭐⭐ **Bon payeur** — délai habituel entre 16 et 25 jours.")
                                elif delai_moy_recent <= 30:
                                    notes.append("⭐ **Payeur à la dernière minute** — paie entre 25 et 30 jours, dans les délais mais sans marge.")
                                else:
                                    notes.append("🔴 **Mauvais payeur** — délai moyen supérieur à 30 jours. Un suivi actif est recommandé.")

                                if p_ret_glob >= 0.20:
                                    notes.append("🔴 1 facture sur 5 dépasse 30 jours sur l'ensemble de l'historique — un suivi actif des impayés est recommandé.")
                                elif p_ret_glob >= 0.10:
                                    notes.append(f"🟡 Taux de retard de {round(p_ret_glob*100)}% sur l'historique global — quelques dossiers méritent un suivi ponctuel.")

                                if len(d_glob) >= 3 and delai_med_ass < delai_moy_ass * 0.7:
                                    notes.append("ℹ️ Quelques factures très tardives tirent la moyenne vers le haut — la majorité des paiements reste plus rapide.")

                                # Tendance : comparer 2 derniers mois vs période précédente
                                date_2m = ajd - pd.DateOffset(months=2)
                                date_4m = ajd - pd.DateOffset(months=4)
                                d_recent = p_hist_global[
                                    (p_hist_global["assureur"] == assureur) &
                                    (p_hist_global["date_facture"] >= date_2m)
                                ]["delai"]
                                d_ancien = p_hist_global[
                                    (p_hist_global["assureur"] == assureur) &
                                    (p_hist_global["date_facture"] >= date_4m) &
                                    (p_hist_global["date_facture"] < date_2m)
                                ]["delai"]
                                if len(d_recent) >= 5 and len(d_ancien) >= 5:
                                    delta = d_recent.mean() - d_ancien.mean()
                                    delta_ret = (d_recent > 30).mean() - (d_ancien > 30).mean()
                                    if delta <= -5 and delta_ret <= -0.05:
                                        notes.append(f"📈 Nette amélioration récente : délai moyen passé de {round(d_ancien.mean())} j à {round(d_recent.mean())} j sur les 2 derniers mois, avec moins de retards.")
                                    elif delta <= -3:
                                        notes.append(f"📈 Légère amélioration récente : délai moyen passé de {round(d_ancien.mean())} j à {round(d_recent.mean())} j.")
                                    elif delta >= 5 and delta_ret >= 0.05:
                                        notes.append(f"📉 Détérioration récente : délai moyen passé de {round(d_ancien.mean())} j à {round(d_recent.mean())} j sur les 2 derniers mois, avec davantage de retards. À surveiller.")
                                    elif delta >= 3:
                                        notes.append(f"📉 Légère détérioration récente : délai moyen passé de {round(d_ancien.mean())} j à {round(d_recent.mean())} j.")
                                    else:
                                        notes.append(f"➡️ Comportement stable sur les 4 derniers mois (délai moyen : {round(d_recent.mean())} j).")

                                st.markdown(f"🏦 **{assureur}** *(sur {n} factures)* : {desc}.")
                                for note in notes:
                                    st.markdown(note)
                                st.markdown("---")
                    else:
                        st.info("Aucune facture ouverte.")
        except Exception as e: st.error(f"Erreur d'analyse : {e}")

# ==========================================
# 🩺 MODULE MÉDECINS (ORIGINAL)
# ==========================================
elif st.session_state.page == "medecins":
    st.markdown("<style>.block-container { padding-left: 1rem; padding-right: 1rem; max-width: 100%; }</style>", unsafe_allow_html=True)
    if st.sidebar.button("⬅️ Retour Accueil"):
        st.session_state.page = "accueil"
        st.rerun()

    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.header("👨‍⚕️ Performance Médecins")
    st.caption("📌 Basé sur les dates de séance réelles — indépendant du comportement de facturation.")
    uploaded_file = st.sidebar.file_uploader("Export Prestations (.xlsx)", type="xlsx", key="med_up",
        help="Fichier exporté depuis Ephysio contenant les onglets 'Prestation' et 'Factures'.")

    # --- CONFIG MÉDECINS ---
    import io as _io
    st.sidebar.markdown("---")
    st.sidebar.markdown("**👨‍⚕️ Configuration des médecins**")
    config_med_file = st.sidebar.file_uploader("Charger config_medecins.xlsx", type="xlsx", key="med_config_up",
        help="Fichier avec col A = nom canonique, col B/C/D = variantes à fusionner.")
    if config_med_file is not None:
        try:
            df_cfg = pd.read_excel(config_med_file, dtype=str)
            mapping_cfg = {}
            for _, row in df_cfg.iterrows():
                canon = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if not canon or canon == 'nan': continue
                for val in row.iloc[1:]:
                    v = str(val).strip() if pd.notna(val) else None
                    if v and v != 'nan':
                        mapping_cfg[v] = canon
            st.session_state.config_medecins = mapping_cfg
            st.sidebar.success(f"✅ {len(df_cfg)} médecins, {len(mapping_cfg)} variantes")
        except Exception as e:
            st.sidebar.error(f"Erreur : {e}")

    st.sidebar.markdown("---")

    if uploaded_file:
        try:
            @st.cache_data
            def lire_medecins_prestations(f):
                xl = pd.ExcelFile(f)
                # Onglet Prestation (séances)
                ong_prest = next((s for s in xl.sheet_names if s.strip().lower() == "prestation"), None) or                             next((s for s in xl.sheet_names if "prestation" in s.lower()), xl.sheet_names[0])
                df_prest = pd.read_excel(f, sheet_name=ong_prest, header=0)
                df_prest.columns = [str(c).strip() for c in df_prest.columns]
                # Onglet Factures (médecins, fournisseurs, lois)
                ong_fact = next((s for s in xl.sheet_names if s.strip().lower() == "factures"), None) or                            next((s for s in xl.sheet_names if "facture" in s.lower()), None)
                if ong_fact is None:
                    raise ValueError("Onglet 'Factures' introuvable dans ce fichier.")
                df_fact = pd.read_excel(f, sheet_name=ong_fact, header=0)
                df_fact.columns = [str(c).strip() for c in df_fact.columns]
                return df_prest, df_fact

            df_prest, df_fact = lire_medecins_prestations(uploaded_file)

            # Résolution colonnes Factures
            _cm = resoudre_colonnes(df_fact)
            if _cm["fournisseur"] is None:
                df_fact["Fournisseur de prestation"] = "Cabinet"
                _cm["fournisseur"] = "Fournisseur de prestation"

            # Jointure Prestation × Factures sur numéro de facture
            col_num_prest = df_prest.columns[0]  # "Numéro de facture"
            col_num_fact  = df_fact.columns[0]   # "Numéro de facture"
            cols_fact_join = [col_num_fact, _cm["medecin"], _cm["fournisseur"], _cm["loi"], _cm["tp_tg"]]
            df_brut = df_prest.merge(
                df_fact[cols_fact_join].drop_duplicates(subset=[col_num_fact]),
                left_on=col_num_prest, right_on=col_num_fact, how="left"
            )

            # Bouton export config vierge basé sur les noms trouvés
            noms_bruts = sorted(df_brut[_cm["medecin"]].dropna().astype(str).str.strip().unique().tolist())
            df_export_cfg = pd.DataFrame({
                "Nom canonique": noms_bruts,
                "Variante 1": [""] * len(noms_bruts),
                "Variante 2": [""] * len(noms_bruts),
                "Variante 3": [""] * len(noms_bruts),
            })
            buf_cfg = _io.BytesIO()
            df_export_cfg.to_excel(buf_cfg, index=False, engine='openpyxl')
            buf_cfg.seek(0)
            st.sidebar.download_button(
                label="📥 Exporter liste médecins (.xlsx)",
                data=buf_cfg,
                file_name="config_medecins.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Téléchargez ce fichier, remplissez les colonnes B/C/D avec les variantes, puis rechargez-le ci-dessus."
            )

            st.sidebar.header("🔍 Filtres")
            fourn_med = sorted(df_brut[_cm["fournisseur"]].dropna().unique().tolist())
            sel_fourn_med = st.sidebar.multiselect("Fournisseurs :", fourn_med, default=fourn_med)
            seuil_jour_med = st.sidebar.number_input("Montant min. pour jour ouvert (CHF) :", min_value=0, max_value=500, value=50, step=10, key="seuil_med")
            exclure_mois_med = st.sidebar.toggle("Exclure le mois en cours", value=True, key="excl_mois_med")
            df_m_init = df_brut[df_brut[_cm["tp_tg"]].astype(str).str.upper() != "TG"].copy()
            df_m_init = df_m_init[df_m_init[_cm["fournisseur"]].isin(sel_fourn_med)]

            # Colonnes Prestation pour date et montant (onglet Prestation)
            col_date_prest   = df_prest.columns[1]   # "Date" (date séance)
            col_chiffre_prest = df_prest.columns[11]  # "Chiffre"

            def moteur_fusion_securise(df):
                noms_originaux = df[_cm["medecin"]].dropna().unique()
                mapping = {}
                def extraire_mots(texte):
                    mots = "".join(c if c.isalnum() else " " for c in str(texte)).upper().split()
                    return {m for m in mots if len(m) > 2}
                noms_tries = sorted(noms_originaux, key=len, reverse=True)
                for i, nom_long in enumerate(noms_tries):
                    mots_long = extraire_mots(nom_long)
                    for nom_court in noms_tries[i+1:]:
                        mots_court = extraire_mots(nom_court)
                        conflit = any(m in mots_long.symmetric_difference(mots_court) for m in MOTS_EXCLUSION)
                        if len(mots_long.intersection(mots_court)) >= 2 and not conflit:
                            mapping[nom_court] = nom_long
                return mapping

            regroupements = moteur_fusion_securise(df_m_init)
            df_m_init[_cm["medecin"]] = df_m_init[_cm["medecin"]].replace(regroupements)

            ajd = pd.Timestamp(datetime.today().date())
            # Fin du mois précédent — on exclut le mois en cours (incomplet)
            fin_mois_precedent = (ajd.replace(day=1) - pd.DateOffset(days=1))
            df_m_init["medecin"] = df_m_init[_cm["medecin"]].fillna("Sans médecin").astype(str).str.strip()
            # Utiliser date et montant de l'onglet Prestation (date de séance réelle)
            df_m_init["ca"] = pd.to_numeric(df_m_init[col_chiffre_prest], errors="coerce").fillna(0)
            df_m_init["date_f"] = df_m_init[col_date_prest].apply(convertir_date)
            # Pour les calculs de tendance (taux CHF/j) : jusqu'à aujourd'hui
            df_m = df_m_init[
                (df_m_init["ca"] > 0) &
                (df_m_init["date_f"].notna()) &
                (df_m_init["date_f"] <= ajd)
            ].copy()
            # Pour le graphique : optionnellement exclure le mois en cours
            if exclure_mois_med:
                df_m_graph = df_m[df_m["date_f"] <= fin_mois_precedent].copy()
                st.caption(f"📅 Graphique jusqu'au {fin_mois_precedent.strftime('%d.%m.%Y')} — tendances calculées jusqu'à aujourd'hui.")
            else:
                df_m_graph = df_m.copy()
                st.caption(f"📅 Graphique jusqu'au {ajd.strftime('%d.%m.%Y')} — mois en cours inclus.")
            # Appliquer le mapping de la config cabinet (variantes → nom canonique)
            if st.session_state.config_medecins:
                df_m["medecin"] = df_m["medecin"].replace(st.session_state.config_medecins)
                df_m_graph["medecin"] = df_m_graph["medecin"].replace(st.session_state.config_medecins)
                nb_mapped = df_m["medecin"].isin(st.session_state.config_medecins.values()).sum()
                if nb_mapped > 0:
                    st.caption(f"✅ Config cabinet active — {len(st.session_state.config_medecins)} variantes mappées")
            
            if not df_m.empty:
                ca_par_jour = df_m.groupby(df_m["date_f"].dt.date)["ca"].sum()
                jours_cabinet = set(ca_par_jour[ca_par_jour >= seuil_jour_med].index)

                # --- Sélecteur de méthode de tendance ---
                st.markdown("### 📊 Méthode de calcul de tendance")
                methode_tendance = st.radio(
                    "Comparer les 60 derniers jours avec :",
                    ["📅 Les 365 derniers jours (méthode actuelle)", "📆 Les mêmes 60 jours de l'année précédente (anti-saisonnalité)"],
                    horizontal=True, key="methode_tendance"
                )
                annee_sur_annee = "précédente" in methode_tendance

                t_60j = ajd - pd.DateOffset(days=60)
                jo_60 = jours_ouvres(t_60j, ajd, jours_cabinet)

                if annee_sur_annee:
                    t_ref_fin   = ajd   - pd.DateOffset(years=1)
                    t_ref_debut = t_60j - pd.DateOffset(years=1)
                if annee_sur_annee:
                    jo_ref = jours_ouvres(t_ref_debut, t_ref_fin, jours_cabinet)
                    label_ref = "CA même période N-1"
                    label_taux_ref = "Taux N-1 (CHF/j)"
                    ca_ref = df_m[(df_m["date_f"] >= t_ref_debut) & (df_m["date_f"] <= t_ref_fin)].groupby("medecin")["ca"].sum().reset_index(name=label_ref)
                else:
                    t_365j = ajd - pd.DateOffset(days=365)
                    jo_ref = jours_ouvres(t_365j, ajd, jours_cabinet)
                    label_ref = "CA 365j"
                    label_taux_ref = "Taux 365j (CHF/j)"
                    ca_ref = df_m[df_m["date_f"] >= t_365j].groupby("medecin")["ca"].sum().reset_index(name=label_ref)

                stats_ca = df_m.groupby("medecin")["ca"].sum().reset_index(name="CA Global")
                ca_60 = df_m[df_m["date_f"] >= t_60j].groupby("medecin")["ca"].sum().reset_index(name="CA 60j")
                tab_final = stats_ca.merge(ca_ref, on="medecin", how="left").merge(ca_60, on="medecin", how="left").fillna(0)
                tab_final["Taux 60j (CHF/j)"]  = (tab_final["CA 60j"] / jo_60).round(2)
                tab_final[label_taux_ref] = (tab_final[label_ref] / jo_ref).round(2)
                tab_final["Tendance"] = tab_final.apply(
                    lambda r: calculer_tendance(r["CA 60j"], r[label_ref], jo_60, jo_ref), axis=1
                )

                st.markdown("### 🏆 Sélection et Visualisation")
                c1, c2, c3 = st.columns([1, 1, 1.5]) 
                with c1: m_top = st.selectbox("Top :", [5, 10, 25, 50, "Tout"], index=1)
                with c2: t_graph = st.radio("Style :", ["📊 Barres", "📈 Courbes"], horizontal=True)
                with c3: visibility = st.radio("Option Tendance :", ["Données", "Ligne", "Les deux"], index=0, horizontal=True)

                tab_s = tab_final[tab_final["medecin"] != "Sans médecin"].sort_values("CA Global", ascending=False)
                tab_final["medecin"] = tab_final["medecin"].astype(str)
                tab_s["medecin"] = tab_s["medecin"].astype(str)
                def_sel = tab_s["medecin"].tolist() if m_top == "Tout" else tab_s.head(int(m_top))["medecin"].tolist()
                options_med = sorted(tab_final[tab_final["medecin"] != "Sans médecin"]["medecin"].astype(str).unique().tolist(), key=lambda x: x.lower())
                choix = st.multiselect("Sélection :", options=options_med, default=def_sel)

                if choix:
                    df_p = df_m_graph[(df_m_graph["medecin"].isin(choix)) & (df_m_graph["medecin"] != "Sans médecin")].copy()
                    df_p["M_Date"] = df_p["date_f"].dt.to_period("M").dt.to_timestamp()
                    df_p = df_p.groupby(["M_Date", "medecin"])["ca"].sum().reset_index()
                    base = alt.Chart(df_p).encode(
                        x=alt.X('M_Date:T', title="Mois", axis=alt.Axis(format='%m.%Y')),
                        y=alt.Y('ca:Q', title="CA (CHF)"),
                        color=alt.Color('medecin:N', legend=alt.Legend(orient='bottom', columns=2, labelLimit=0))
                    ).properties(height=600)
                    data_layer = base.mark_bar(opacity=0.6) if "Barres" in t_graph else base.mark_line(point=True)
                    trend_layer = base.transform_regression('M_Date', 'ca', groupby=['medecin']).mark_line(size=4, strokeDash=[6, 4])
                    chart = data_layer if visibility == "Données" else trend_layer if visibility == "Ligne" else data_layer + trend_layer
                    st.altair_chart(chart, use_container_width=True)
                    try:
                        _df_med_chart = df_p.groupby(["M_Date", "medecin"])["ca"].sum().unstack(fill_value=0)
                        _pdf_buf = generer_pdf_graphique_matplotlib("CA par médecin", _df_med_chart, sous_titre=f"Calculé au {datetime.today().strftime('%d.%m.%Y')}", ylabel="CA (CHF)")
                        st.download_button("📄 Télécharger le graphique en PDF", _pdf_buf, file_name="medecins_graphique.pdf", mime="application/pdf", key="pdf_med_graph", use_container_width=True)
                    except Exception as _e:
                        st.caption(f"Export PDF indisponible : {_e}")
                    _df_base = tab_final[tab_final["medecin"].isin(choix)].sort_values("CA Global", ascending=False).copy()
                    _df_base = _df_base.apply(lambda c: c.round(2) if c.dtype.kind == 'f' else c)

                    # Les deux modes utilisent les taux CHF/j — neutralise les jours fériés et fermetures
                    cols_affichage = ["medecin", label_taux_ref, "Taux 60j (CHF/j)", "Tendance", label_ref, "CA 60j", "CA Global"]
                    _df_disp_med = _df_base[cols_affichage].copy()
                    legende_med = (
                        f"🔵 <b>{label_taux_ref}</b> vs <b>Taux 60j</b> — comparaison en CHF/j (neutralise jours fériés et fermetures)"
                    )
                    st.markdown(f"<small>{legende_med}</small>", unsafe_allow_html=True)
                    st.dataframe(_df_disp_med, use_container_width=True, hide_index=True, column_config={
                        "medecin":            st.column_config.TextColumn("Médecin"),
                        label_taux_ref:       st.column_config.NumberColumn(f"⟵ {label_taux_ref}", format="%.2f", help="Période de référence"),
                        "Taux 60j (CHF/j)":  st.column_config.NumberColumn("⟵ Taux 60j (CHF/j)", format="%.2f", help="Période récente — à comparer avec la colonne précédente"),
                        "Tendance":           st.column_config.TextColumn("Tendance"),
                        label_ref:            st.column_config.NumberColumn(f"{label_ref} (CHF)", format="%.2f"),
                        "CA 60j":             st.column_config.NumberColumn("CA 60j (CHF)", format="%.2f"),
                        "CA Global":          st.column_config.NumberColumn("CA Global (CHF)", format="%.2f"),
                    })

                    _df_pdf_med = _df_disp_med.apply(lambda c: c.round(2) if c.dtype.kind == 'f' else c)
                    _pdf_buf = generer_pdf_tableau("Performance Médecins", _df_pdf_med, f"Calculé au {datetime.today().strftime('%d.%m.%Y')}")
                    st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name="medecins.pdf", mime="application/pdf", key="pdf_medecins", use_container_width=True)
        except Exception as e: st.error(f"Erreur technique : {e}")

# ==========================================
# 🏷️ MODULE TARIFS (PERFORMANCE & TENDANCES)
# ==========================================
elif st.session_state.page == "tarifs":
    if st.sidebar.button("⬅️ Retour Accueil"):
        st.session_state.page = "accueil"
        st.rerun()

    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.title("📊 Analyse des revenus mensuels et Tendances")
    uploaded_file = st.sidebar.file_uploader("📂 Déposer l'export Excel (onglet 'Prestation')", type="xlsx", key="tarif_up")

    if uploaded_file:
        try:
            ong_p = next((s for s in pd.ExcelFile(uploaded_file).sheet_names if 'Prestation' in s or 'prestation' in s.lower()), 'Prestation')
            df = pd.read_excel(uploaded_file, sheet_name=ong_p)
            nom_col_code = df.columns[2]   # C (Tarif)
            df[nom_col_code] = df[nom_col_code].apply(nettoyer_code_tarif)
            nom_col_nom  = df.columns[3]   # D (Nom de la prestation)
            nom_col_somme = df.columns[11] # L (Montant)
            date_cols = [c for c in df.columns if 'Date' in str(c)]
            nom_col_date = date_cols[0] if date_cols else df.columns[0]

            df[nom_col_somme] = pd.to_numeric(df[nom_col_somme], errors='coerce')
            df[nom_col_date] = pd.to_datetime(df[nom_col_date], errors='coerce')
            df = df[df[nom_col_somme] > 0].dropna(subset=[nom_col_date, nom_col_somme])
            
            # --- GESTION DE LA PÉRIODE ET AFFICHAGE ---
            st.sidebar.header("📅 Période & Graphique")
            exclure_actuel = st.sidebar.toggle("Exclure le mois en cours", value=True)
            y_axis_zero = st.sidebar.toggle("Forcer l'axe Y à zéro", value=False)
            
            maintenant = pd.Timestamp(datetime.today().date())
            
            if exclure_actuel:
                reference_date = maintenant.replace(day=1) - pd.Timedelta(days=1)
                df = df[df[nom_col_date] <= reference_date]
            else:
                reference_date = maintenant

            df['Profession'] = df[nom_col_code].apply(assigner_profession)

            # --- FILTRAGE ---
            st.sidebar.header("⚙️ Filtres")
            seuil_jour_tar = st.sidebar.number_input("Montant min. pour jour ouvert (CHF) :", min_value=0, max_value=500, value=50, step=10, key="seuil_tar")
            professions_dispo = sorted(df['Profession'].unique())
            metiers_actifs = [p for p in professions_dispo if st.sidebar.checkbox(p, value=True, key=f"t_check_{p}")]

            codes_possibles = df[df['Profession'].isin(metiers_actifs)]
            liste_codes = sorted(codes_possibles[nom_col_code].unique().astype(str))
            selection_codes = st.sidebar.multiselect("Codes à afficher :", options=liste_codes, default=liste_codes)

            view_mode = st.radio("Affichage :", ["Profession", "Code tarifaire"], horizontal=True)
            chart_type = st.radio("Style :", ["Barres", "Courbes"], horizontal=True)
            methode_tarif = st.radio(
                "Tendance — comparer les 60 derniers jours avec :",
                ["📅 Les 365 derniers jours (méthode actuelle)", "📆 Les mêmes 60 jours de l'année précédente (anti-saisonnalité)"],
                horizontal=True, key="methode_tarif"
            )

            df_filtered = df[df[nom_col_code].astype(str).isin(selection_codes)].copy()

            if not df_filtered.empty:
                # 1. GRAPHIQUE D'ÉVOLUTION
                df_filtered['Mois'] = df_filtered[nom_col_date].dt.to_period('M').dt.to_timestamp()
                target_col = "Profession" if view_mode == "Profession" else nom_col_code
                df_plot = df_filtered.groupby(['Mois', target_col])[nom_col_somme].sum().reset_index()
                
                color_map = COULEURS_PROF if view_mode == "Profession" else None
                if chart_type == "Barres":
                    fig = px.bar(df_plot, x='Mois', y=nom_col_somme, color=target_col, 
                                 barmode='group', color_discrete_map=color_map, text_auto='.2f')
                else:
                    fig = px.line(df_plot, x='Mois', y=nom_col_somme, color=target_col, 
                                  markers=True, color_discrete_map=color_map)
                
                # Application de la logique d'axe Y
                if y_axis_zero:
                    fig.update_yaxes(rangemode="tozero")
                else:
                    fig.update_yaxes(rangemode="normal")

                fig.update_xaxes(dtick="M1", tickformat="%b %Y")
                st.plotly_chart(fig, use_container_width=True)
                try:
                    _pdf_buf = generer_pdf_plotly(f"Évolution CA — {view_mode}", fig, sous_titre=f"Calculé au {reference_date.strftime('%d.%m.%Y')}")
                    st.download_button("📄 Télécharger le graphique en PDF", _pdf_buf, file_name="tarifs_graphique.pdf", mime="application/pdf", key="pdf_tarifs_graph", use_container_width=True)
                except Exception as _e:
                    try:
                        _df_tarif_chart = df_plot.pivot(index='Mois', columns=target_col, values=nom_col_somme).fillna(0)
                        _pdf_buf = generer_pdf_graphique_matplotlib(f"Évolution CA — {view_mode}", _df_tarif_chart, sous_titre=f"Calculé au {reference_date.strftime('%d.%m.%Y')}", ylabel="CA (CHF)")
                        st.download_button("📄 Télécharger le graphique en PDF", _pdf_buf, file_name="tarifs_graphique.pdf", mime="application/pdf", key="pdf_tarifs_graph", use_container_width=True)
                    except Exception as _e2:
                        st.caption(f"Export PDF indisponible : {_e2}")

                # 2. TABLEAU DES TENDANCES
                st.markdown(f"### 📈 Performance par Tarif (Base : {reference_date.strftime('%d.%m.%Y')})")

                ca_par_jour_t = df.groupby(df[nom_col_date].dt.date)[nom_col_somme].sum()
                jours_cabinet_t = set(ca_par_jour_t[ca_par_jour_t >= seuil_jour_tar].index)  # Jours réels du cabinet avec min de facturation
                annee_sur_annee_t = "précédente" in methode_tarif

                t_60j = reference_date - pd.DateOffset(days=60)
                jo_60 = jours_ouvres(t_60j, reference_date, jours_cabinet_t)

                if annee_sur_annee_t:
                    t_ref_fin   = reference_date - pd.DateOffset(years=1)
                    t_ref_debut = t_60j          - pd.DateOffset(years=1)
                if annee_sur_annee_t:
                    jo_ref      = jours_ouvres(t_ref_debut, t_ref_fin, jours_cabinet_t)
                    label_ref   = "CA même période N-1"
                    ca_ref = df_filtered[(df_filtered[nom_col_date] >= t_ref_debut) & (df_filtered[nom_col_date] <= t_ref_fin)].groupby(nom_col_code)[nom_col_somme].sum().reset_index(name=label_ref)
                else:
                    t_365j      = reference_date - pd.DateOffset(days=365)
                    jo_ref      = jours_ouvres(t_365j, reference_date, jours_cabinet_t)
                    label_ref   = "CA 365j"
                    ca_ref = df_filtered[df_filtered[nom_col_date] >= t_365j].groupby(nom_col_code)[nom_col_somme].sum().reset_index(name=label_ref)

                label_taux_ref = "Taux N-1 (CHF/j)" if annee_sur_annee_t else "Taux 365j (CHF/j)"

                # --- Groupement selon le mode d'affichage ---
                group_col = "Profession" if view_mode == "Profession" else nom_col_code

                if annee_sur_annee_t:
                    ca_ref_g = df_filtered[(df_filtered[nom_col_date] >= t_ref_debut) & (df_filtered[nom_col_date] <= t_ref_fin)].groupby(group_col)[nom_col_somme].sum().reset_index(name=label_ref)
                else:
                    ca_ref_g = df_filtered[df_filtered[nom_col_date] >= t_365j].groupby(group_col)[nom_col_somme].sum().reset_index(name=label_ref)

                stats_global = df_filtered.groupby(group_col)[nom_col_somme].sum().reset_index(name="CA Global")
                ca_60_g      = df_filtered[df_filtered[nom_col_date] >= t_60j].groupby(group_col)[nom_col_somme].sum().reset_index(name="CA 60j")

                tab_perf = stats_global.merge(ca_ref_g, on=group_col, how="left").merge(ca_60_g, on=group_col, how="left").fillna(0)
                tab_perf["Taux 60j (CHF/j)"] = (tab_perf["CA 60j"]  / jo_60).round(2)
                tab_perf[label_taux_ref]      = (tab_perf[label_ref] / jo_ref).round(2)
                tab_perf["Tendance"] = tab_perf.apply(
                    lambda r: calculer_tendance(r["CA 60j"], r[label_ref], jo_60, jo_ref), axis=1
                )

                # Pour le mode code : ajouter le nom de la prestation en tooltip
                if view_mode != "Profession":
                    noms_prestation = (
                        df_filtered.groupby(nom_col_code)[nom_col_nom]
                        .agg(lambda x: x.mode().iloc[0] if not x.mode().empty else "")
                        .reset_index().rename(columns={nom_col_nom: "Prestation"})
                    )
                    tab_perf = tab_perf.merge(noms_prestation, on=nom_col_code, how="left")

                tab_sorted = tab_perf.sort_values("CA Global", ascending=False)
                tab_sorted = tab_sorted.apply(lambda c: c.round(2) if c.dtype.kind == 'f' else c)

                def tendance_html(t):
                    if "Hausse" in t:    color = "#1a7f3c"
                    elif "Baisse" in t:  color = "#c0392b"
                    elif "Nouveau" in t: color = "#0066cc"
                    else:                color = "#666666"
                    return f'<span style="color:{color};font-weight:600">{t}</span>'

                COULEURS_PROF_HEX = {"Physiothérapie": "#00CCFF", "Ergothérapie": "#FF9900", "Massage": "#00CC96", "Autre": "#AB63FA"}

                rows = ""
                for _, r in tab_sorted.iterrows():
                    val = str(r[group_col])
                    if view_mode == "Profession":
                        hex_c = COULEURS_PROF_HEX.get(val, "#888")
                        first_cell = f'<td><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{hex_c};margin-right:6px"></span>{val}</td>'
                    else:
                        nom = str(r.get("Prestation", ""))
                        first_cell = f'<td><span title="{nom}" style="cursor:help;border-bottom:1px dotted #999">{val}</span></td>'
                    # Delta % pour mode N-1
                    ca_60_val = r["CA 60j"]
                    ca_ref_val = r[label_ref]
                    delta_pct = f"{((ca_60_val - ca_ref_val) / ca_ref_val * 100):+.1f}%" if ca_ref_val and ca_ref_val != 0 else "—"

                    # Les deux modes : taux CHF/j — neutralise jours fériés et fermetures
                    rows += (
                        f'<tr>'
                        f'{first_cell}'
                        f'<td style="text-align:right;background:#EBF5FB;font-weight:600">{chf(r[label_taux_ref])}</td>'
                        f'<td style="text-align:right;background:#EBF5FB;font-weight:600">{chf(r["Taux 60j (CHF/j)"])}</td>'
                        f'<td>{tendance_html(r["Tendance"])}</td>'
                        f'<td style="text-align:right">{chf(r[label_ref])}</td>'
                        f'<td style="text-align:right">{chf(r["CA 60j"])}</td>'
                        f'<td style="text-align:right">{chf(r["CA Global"])}</td>'
                        f'</tr>'
                    )

                first_col_header = "Profession" if view_mode == "Profession" else "Code"
                tooltip_note = "" if view_mode == "Profession" else "<p style='font-size:0.75rem;color:#999;margin-top:4px'>ℹ️ Survolez le code pour voir le nom de la prestation</p>"

                headers = (
                    f"<th>{first_col_header}</th>"
                    f"<th style='background:#D6EAF8'>⟵ {label_taux_ref}</th>"
                    f"<th style='background:#D6EAF8'>⟵ Taux 60j (CHF/j)</th>"
                    f"<th>Tendance</th>"
                    f"<th>{label_ref} (CHF)</th><th>CA 60j (CHF)</th><th>CA Global (CHF)</th>"
                )
                legende = f"<small>🔵 <b>{label_taux_ref}</b> vs <b>Taux 60j</b> — comparaison en CHF/j (neutralise jours fériés et fermetures)</small>"

                html_table = (
                    "<style>"
                    ".tarif-table{width:100%;border-collapse:collapse;font-size:0.9rem}"
                    ".tarif-table th{background:#f0f2f6;padding:8px 12px;text-align:left;border-bottom:2px solid #ddd;vertical-align:bottom;min-width:60px}"
                    ".tarif-table th:nth-child(n+3){text-align:right}"
                    ".tarif-table td{padding:6px 12px;border-bottom:1px solid #eee;white-space:nowrap}"
                    ".tarif-table td:nth-child(n+3){text-align:right}"
                    ".tarif-table tr:hover td{background:#f8f9fa}"
                    "</style>"
                    f"<p style='font-size:0.8rem;color:#555;margin-bottom:6px'>{legende}</p>"
                    "<table class='tarif-table'>"
                    f"<thead><tr>{headers}</tr></thead>"
                    f"<tbody>{rows}</tbody>"
                    "</table>"
                    f"{tooltip_note}"
                )
                st.markdown(html_table, unsafe_allow_html=True)
                # PDF du tableau tarifs
                if not tab_sorted.empty:
                    _cols_pdf_t = [group_col, "Tendance", "CA Global", label_ref, label_taux_ref, "CA 60j", "Taux 60j (CHF/j)"]
                    _cols_pdf_t = [c for c in _cols_pdf_t if c in tab_sorted.columns]
                    _df_pdf_t = tab_sorted[_cols_pdf_t].copy()
                    _df_pdf_t.columns = [str(c) for c in _df_pdf_t.columns]
                    _pdf_buf = generer_pdf_tableau(f"Performance Tarifs — {view_mode}", _df_pdf_t, f"Calculé au {reference_date.strftime('%d.%m.%Y')}")
                    st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name="tarifs.pdf", mime="application/pdf", key="pdf_tarifs", use_container_width=True)
            else:
                st.warning("Aucune donnée disponible pour cette sélection.")
                
        except Exception as e: st.error(f"Erreur Tarifs : {e}")
# ==========================================
# 🏦 MODULE BILAN COMPTABLE (V10 - AVEC LIGNE TOTAL)
# ==========================================
elif st.session_state.page == "bilan":
    if st.sidebar.button("⬅️ Retour Accueil"):
        st.session_state.page = "accueil"
        st.rerun()

    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.title("🏦 Bilan des Revenus par Fournisseur")
    up = st.sidebar.file_uploader("Fichier Excel (Export avec onglet Facture)", type="xlsx", key="bilan_up")
    
    if up:
        try:
            xl = pd.ExcelFile(up)
            ong_f = next((s for s in xl.sheet_names if 'Facture' in s or 'facture' in s.lower()), None)
            
            if not ong_f:
                st.error(f"L'onglet 'Facture' est introuvable. Onglets disponibles : {', '.join(xl.sheet_names)}")
                st.stop()
            
            df_f = pd.read_excel(up, sheet_name=ong_f)
            
           # --- CONFIGURATION DES COLONNES ---
            col_date_f = df_f.columns[2]   # C: Date de la facture
            col_fourn_f = df_f.columns[9]  # J: Fournisseur
            col_ca_f = df_f.columns[14]    # O: Montant (CA)
            col_paye_f = df_f.columns[15]  # P: Date de paiement
            
            df_f[col_date_f] = pd.to_datetime(df_f[col_date_f], errors='coerce')
            df_f[col_ca_f] = pd.to_numeric(df_f[col_ca_f], errors='coerce').fillna(0)
            df_f = df_f.dropna(subset=[col_date_f])

            # Extraction des années uniques
            annees = sorted(df_f[col_date_f].dt.year.unique().astype(int), reverse=True)
            
            # --- NOUVEAU : ALERTE MULTI-ANNÉES ---
            if len(annees) > 1:
                st.warning(
                    f"⚠️ **Attention :** L'export chargé contient des données sur {len(annees)} années différentes "
                    f"({min(annees)} à {max(annees)}). Le bilan est conçu pour analyser un exercice comptable unique. "
                    "Veuillez faire un export des prestations du 1er janvier au 31 décembre d'une seule année."
                )

            annee = st.sidebar.selectbox("Année d'analyse :", annees)
            df_sel = df_f[df_f[col_date_f].dt.year == annee].copy()

            # --- SECTION CHIFFRE D'AFFAIRES ---
            st.subheader(f"📊 Analyse du Chiffre d'Affaires ({annee})")
            vue_ca = st.radio("Affichage CA par Fournisseur :", ["Annuel (Cumulé)", "Mensuel (Détail)"], horizontal=True)

            if vue_ca == "Annuel (Cumulé)":
                ca_fourn = df_sel.groupby(col_fourn_f)[col_ca_f].sum().round(2).sort_values(ascending=False).reset_index()
                
                # Ajout de la ligne Total pour le cumul annuel
                total_val = ca_fourn[col_ca_f].sum()
                ligne_total = pd.DataFrame({col_fourn_f: ['TOTAL GÉNÉRAL'], col_ca_f: [total_val]})
                ca_fourn = pd.concat([ca_fourn, ligne_total], ignore_index=True)
                
                st.dataframe(
                    ca_fourn, 
                    use_container_width=True, 
                    hide_index=True,
                    column_config={
                        col_fourn_f: "Fournisseur", 
                        col_ca_f: st.column_config.NumberColumn("Total CA", format="%.2f CHF")
                    }
                )
                _pdf_buf = generer_pdf_tableau(f"Bilan CA {annee}", ca_fourn.rename(columns={col_fourn_f: "Fournisseur", col_ca_f: "Total CA (CHF)"}), f"Exercice {annee}")
                st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name=f"bilan_ca_{annee}.pdf", mime="application/pdf", key="pdf_bilan_ca", use_container_width=True)
            else:
                df_sel['Mois_Num'] = df_sel[col_date_f].dt.month
                nom_mois = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jun", "Jul", "Août", "Sep", "Oct", "Nov", "Déc"]
                
                pivot_fourn = df_sel.pivot_table(index=col_fourn_f, columns='Mois_Num', values=col_ca_f, aggfunc='sum', fill_value=0)
                pivot_fourn = pivot_fourn.reindex(columns=range(1, 13), fill_value=0)
                pivot_fourn.columns = nom_mois
                pivot_fourn = pivot_fourn.round(2)
                pivot_fourn['TOTAL'] = pivot_fourn.sum(axis=1).round(2)
                
                # Ajout de la ligne de Totalisation en bas du tableau mensuel
                pivot_total = pivot_fourn.sum(axis=0).to_frame().T
                pivot_total.index = ["TOTAL GÉNÉRAL"]
                pivot_final = pd.concat([pivot_fourn, pivot_total])
                
                pivot_final = pivot_final.round(2)
                st.dataframe(pivot_final.style.format(lambda x: chf(x) if isinstance(x, (int,float)) else str(x)).highlight_max(axis=0, color='#d4f1f9'), use_container_width=True)

            # --- SECTION IMPAYÉS AU 31.12 ---
            st.markdown("---")
            st.subheader(f"⏳ Factures Impayées au 31.12.{annee}")
            
            df_impayes = df_sel[df_sel[col_paye_f].isna()].copy()
            total_impayes = df_impayes[col_ca_f].sum()

            if total_impayes > 0:
                st.warning(f"Montant total restant à percevoir pour {annee} : **{chf(total_impayes)} CHF**")
                imp_par_fourn = df_impayes.groupby(col_fourn_f)[col_ca_f].sum().sort_values(ascending=False).reset_index()
                
                # Ajout de la ligne Total aussi pour les impayés
                ligne_total_imp = pd.DataFrame({col_fourn_f: ['TOTAL DES IMPAYÉS'], col_ca_f: [total_impayes]})
                imp_par_fourn = pd.concat([imp_par_fourn, ligne_total_imp], ignore_index=True)
                
                with st.expander("Voir le détail des impayés par fournisseur"):
                    st.dataframe(
                        imp_par_fourn, 
                        use_container_width=True, 
                        hide_index=True,
                        column_config={col_fourn_f: "Fournisseur", col_ca_f: st.column_config.NumberColumn("Montant dû", format="%.2f CHF")}
                    )
                    _pdf_buf = generer_pdf_tableau(f"Impayés au 31.12.{annee}", imp_par_fourn.rename(columns={col_fourn_f: "Fournisseur", col_ca_f: "Montant dû (CHF)"}), f"Exercice {annee}")
                    st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name=f"impayes_{annee}.pdf", mime="application/pdf", key="pdf_impayes", use_container_width=True)
            else:
                st.success(f"Toutes les factures de l'année {annee} sont marquées comme payées.")

        except Exception as e:
            st.error(f"Erreur d'analyse : {e}")

elif st.session_state.page == "stats_patients":
    render_stats_patients()

# ==========================================
# ==========================================
# 🤝 MODULE RÉTROCESSION
# ==========================================
elif st.session_state.page == "retrocession":
    import io as _io_retro

    if st.sidebar.button("⬅️ Retour Accueil"):
        st.session_state.page = "accueil"
        st.rerun()

    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.title("🤝 Calcul de Rétrocession")
    st.caption("Calculez la rétrocession due par un·e thérapeute indépendant·e à partir de son export Ephysio.")

    if not st.session_state.get("retro_warning_seen"):
        st.sidebar.warning("⚠️ Soyez attentif au fait que des factures rejetées sur cette période peuvent encore être non-traitées et ne figurent donc pas dans ce décompte.")
        if st.sidebar.button("OK, j'en suis conscient", key="retro_warning_ok", type="primary", use_container_width=True):
            st.session_state["retro_warning_seen"] = True
            st.rerun()
        st.stop()



    # --- SIDEBAR : FICHIERS ---
    st.sidebar.markdown("---")
    st.sidebar.markdown("**📂 Fichiers**")
    uploaded_retro = st.sidebar.file_uploader(
        "Export Prestations du/de la thérapeute (.xlsx)",
        type="xlsx", key="retro_up"
    )


    st.sidebar.markdown("---")
    st.sidebar.markdown("**⚙️ Grille de taux**")
    taux_file = st.sidebar.file_uploader(
        "Charger une grille de taux (.xlsx)",
        type="xlsx", key="retro_taux_up",
        help="Rechargez une grille sauvegardée pour pré-remplir les pourcentages."
    )

    # --- SIDEBAR : PÉRIODE ---
    st.sidebar.markdown("---")
    st.sidebar.markdown("**📅 Période**")
    periode_mode = st.sidebar.radio(
        "Filtrer par :",
        ["Tout l'export", "Trimestre", "Période personnalisée"],
        key="retro_periode_mode"
    )

    if uploaded_retro:
        try:
            @st.cache_data
            def lire_retro(f):
                df = pd.read_excel(f, sheet_name=None)
                # Chercher onglet Prestation (insensible casse)
                for k in df:
                    if k.strip().lower() == "prestation":
                        return df[k]
                return list(df.values())[0]

            df_r = lire_retro(uploaded_retro)
            df_r.columns = [str(c).strip() for c in df_r.columns]

            # Colonnes export Ephysio Prestations — détection par nom
            _cr = resoudre_colonnes(df_r)
            c_date = _cr["date_facture"] or df_r.columns[1]
            c_code = df_r.columns[2]   # Code tarifaire toujours col 2 dans Prestations
            c_mont = _cr["chiffre"] or _cr["montant"] or df_r.columns[11]

            df_r[c_date] = pd.to_datetime(df_r[c_date], errors="coerce")
            df_r[c_mont] = pd.to_numeric(df_r[c_mont], errors="coerce").fillna(0)
            df_r[c_code] = df_r[c_code].apply(nettoyer_code_tarif)

            # Garder uniquement les lignes avec montant positif
            df_r = df_r[(df_r[c_mont] > 0) & df_r[c_date].notna()].copy()

            date_min = df_r[c_date].min()
            date_max = df_r[c_date].max()

            # --- FILTRE PÉRIODE ---
            if periode_mode == "Trimestre":
                annees = sorted(df_r[c_date].dt.year.unique(), reverse=True)
                sel_annee = st.sidebar.selectbox("Année", annees, key="retro_annee")
                sel_trim = st.sidebar.selectbox("Trimestre", ["T1 (jan-mar)", "T2 (avr-jun)", "T3 (jul-sep)", "T4 (oct-déc)"], key="retro_trim")
                trim_map = {"T1 (jan-mar)": (1,3), "T2 (avr-jun)": (4,6), "T3 (jul-sep)": (7,9), "T4 (oct-déc)": (10,12)}
                m1, m2 = trim_map[sel_trim]
                df_f = df_r[(df_r[c_date].dt.year == sel_annee) & (df_r[c_date].dt.month.between(m1, m2))]
                label_periode = f"{sel_trim} {sel_annee}"
            elif periode_mode == "Période personnalisée":
                d1 = st.sidebar.date_input("Du", value=date_min.date(), key="retro_d1")
                d2 = st.sidebar.date_input("Au", value=date_max.date(), key="retro_d2")
                df_f = df_r[(df_r[c_date].dt.date >= d1) & (df_r[c_date].dt.date <= d2)]
                label_periode = f"{d1.strftime('%d.%m.%Y')} – {d2.strftime('%d.%m.%Y')}"
            else:
                df_f = df_r.copy()
                label_periode = f"{date_min.strftime('%d.%m.%Y')} – {date_max.strftime('%d.%m.%Y')}"

            if df_f.empty:
                st.warning("Aucune prestation sur la période sélectionnée.")
                st.stop()

            # Colonne patient (col 8, index 8)
            c_pat = _cr["num_patient"] or _cr["patient"] or df_r.columns[8]

            # --- DÉTECTION PAIRES 7311/7354 (séances à domicile) ---
            # Une paire domicile = même jour + même patient + présence de 7311 ET 7354
            df_7354 = df_f[df_f[c_code] == "7354"][[c_date, c_pat]].copy()
            df_7311 = df_f[df_f[c_code].isin(["7311", "7301"])][[c_date, c_pat, c_code, c_mont]].copy()

            # Jointure sur date + patient pour identifier les 7311 accompagnés d'un 7354
            paires = pd.merge(
                df_7311,
                df_7354.drop_duplicates().assign(_domicile=True),
                on=[c_date, c_pat], how="left"
            )
            paires["_domicile"] = paires["_domicile"].fillna(False)

            nb_domicile_7311 = paires[paires["_domicile"] & paires[c_code].isin(["7311"])].shape[0]
            nb_domicile_7301 = paires[paires["_domicile"] & paires[c_code].isin(["7301"])].shape[0]

            ca_domicile_7311 = paires.loc[paires["_domicile"] & (paires[c_code] == "7311"), c_mont].sum()
            ca_domicile_7301 = paires.loc[paires["_domicile"] & (paires[c_code] == "7301"), c_mont].sum()
            ca_cabinet_7311  = paires.loc[~paires["_domicile"] & (paires[c_code] == "7311"), c_mont].sum()
            ca_cabinet_7301  = paires.loc[~paires["_domicile"] & (paires[c_code] == "7301"), c_mont].sum()
            nb_cabinet_7311  = paires[~paires["_domicile"] & (paires[c_code] == "7311")].shape[0]
            nb_cabinet_7301  = paires[~paires["_domicile"] & (paires[c_code] == "7301")].shape[0]

            if nb_domicile_7311 + nb_domicile_7301 > 0:
                st.info(f"🏠 **{nb_domicile_7311 + nb_domicile_7301} séances à domicile détectées** — séparées dans la grille ci-dessous.")

            # --- AGRÉGAT PAR CODE ---
            agg = df_f.groupby(c_code).agg(
                CA=(c_mont, "sum"),
                Nb_lignes=(c_mont, "count")
            ).reset_index().rename(columns={c_code: "Code"})

            # Remplacer les lignes 7311 et 7301 par des versions split cabinet/domicile
            rows_extra = []
            for code_base, ca_cab, nb_cab, ca_dom, nb_dom in [
                ("7311", ca_cabinet_7311, nb_cabinet_7311, ca_domicile_7311, nb_domicile_7311),
                ("7301", ca_cabinet_7301, nb_cabinet_7301, ca_domicile_7301, nb_domicile_7301),
            ]:
                if code_base in agg["Code"].values:
                    agg = agg[agg["Code"] != code_base]  # retirer la ligne globale
                    if nb_cab > 0:
                        rows_extra.append({"Code": f"{code_base} (cabinet)", "CA": round(ca_cab, 2), "Nb_lignes": nb_cab})
                    if nb_dom > 0:
                        rows_extra.append({"Code": f"{code_base} (domicile)", "CA": round(ca_dom, 2), "Nb_lignes": nb_dom})

            if rows_extra:
                agg = pd.concat([agg, pd.DataFrame(rows_extra)], ignore_index=True)

            agg = agg.sort_values("CA", ascending=False).reset_index(drop=True)

            # --- CHARGEMENT GRILLE DE TAUX SAUVEGARDÉE ---
            taux_precharges = {}
            if taux_file is not None:
                try:
                    df_taux = pd.read_excel(taux_file, dtype=str)
                    if "Code" in df_taux.columns and "Taux (%)" in df_taux.columns:
                        for _, row in df_taux.iterrows():
                            code = str(row["Code"]).strip()
                            try:
                                taux_precharges[code] = float(str(row["Taux (%)"]).replace(",", "."))
                            except:
                                taux_precharges[code] = 0.0
                        st.sidebar.success(f"✅ {len(taux_precharges)} taux chargés")
                except Exception as e:
                    st.sidebar.error(f"Erreur grille : {e}")

            # --- INTERFACE PRINCIPALE ---
            st.subheader(f"📋 Grille de rétrocession — {label_periode}")
            st.caption(f"**{len(agg)} codes tarifaires** trouvés | CA total : **{chf(df_f[c_mont].sum())} CHF**")
            st.markdown("Saisissez le taux de rétrocession pour chaque code. Mettez **0%** pour ne pas prélever sur une position.")
            st.markdown("---")

            # Construire le dataframe éditable
            agg["Taux (%)"] = agg["Code"].map(taux_precharges).fillna(0.0)
            agg["CA (CHF)"] = agg["CA"].round(2)
            agg["Nb prestations"] = agg["Nb_lignes"]

            df_edit = agg[["Code", "Nb prestations", "CA (CHF)", "Taux (%)"]].copy()

            edited = st.data_editor(
                df_edit,
                column_config={
                    "Code": st.column_config.TextColumn("Code tarifaire", disabled=True),
                    "Nb prestations": st.column_config.NumberColumn("Nb prestations", disabled=True, format="%d"),
                    "CA (CHF)": st.column_config.NumberColumn("CA (CHF)", disabled=True, format="%.2f"),
                    "Taux (%)": st.column_config.NumberColumn(
                        "Rétrocession (%)",
                        min_value=0.0, max_value=100.0, step=0.5, format="%.2f",
                        help="Entrez le % de rétrocession pour ce code. 0 = aucune retenue."
                    ),
                },
                use_container_width=True,
                hide_index=True,
                key="retro_grid"
            )

            # --- SAUVEGARDE GRILLE ---
            buf_taux = _io_retro.BytesIO()
            edited[["Code", "Taux (%)"]].to_excel(buf_taux, index=False, engine='openpyxl')
            buf_taux.seek(0)
            st.sidebar.download_button(
                label="💾 Sauvegarder la grille de taux",
                data=buf_taux,
                file_name="grille_retrocession.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            # --- CALCUL ---
            st.markdown("---")
            if st.button("🧮 Calculer la rétrocession", type="primary", use_container_width=True):

                edited["Rétrocession (CHF)"] = (edited["CA (CHF)"] * edited["Taux (%)"] / 100).round(2)
                detail = edited[edited["Taux (%)"] > 0].copy()
                total_retro = detail["Rétrocession (CHF)"].sum()
                total_ca    = edited["CA (CHF)"].sum()
                ca_couvert  = detail["CA (CHF)"].sum()

                st.subheader("📊 Résultat")

                col1, col2, col3 = st.columns(3)
                col1.metric("CA total période", f"{chf(total_ca)} CHF")
                col2.metric("CA soumis à rétrocession", f"{chf(ca_couvert)} CHF")
                col3.metric("💰 Rétrocession due", f"{chf(total_retro)} CHF",
                    delta=f"{(total_retro/total_ca*100):.2f}% du CA total" if total_ca > 0 else None)

                st.markdown("#### Détail par code")
                st.dataframe(
                    detail[["Code", "Nb prestations", "CA (CHF)", "Taux (%)", "Rétrocession (CHF)"]].sort_values("Rétrocession (CHF)", ascending=False),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "CA (CHF)":           st.column_config.NumberColumn(format="%.2f"),
                        "Taux (%)":           st.column_config.NumberColumn(format="%.2f"),
                        "Rétrocession (CHF)": st.column_config.NumberColumn(format="%.2f"),
                    }
                )
                _df_retro_pdf = detail[["Code", "Nb prestations", "CA (CHF)", "Taux (%)", "Rétrocession (CHF)"]].sort_values("Rétrocession (CHF)", ascending=False)
                _pdf_buf = generer_pdf_tableau(f"Décompte Rétrocession — {label_periode}", _df_retro_pdf, f"Total dû : {chf(total_retro)} CHF")
                st.download_button("📄 Télécharger en PDF", _pdf_buf, file_name=f"retrocession_{label_periode.replace(' ','_')}.pdf", mime="application/pdf", key="pdf_retro", use_container_width=True)

                # Codes à 0% pour info
                codes_exclus = edited[edited["Taux (%)"] == 0]["Code"].tolist()
                if codes_exclus:
                    st.caption(f"Codes sans rétrocession (0%) : {', '.join(codes_exclus)}")

                # --- EXPORT DÉCOMPTE EXCEL ---
                buf_out = _io_retro.BytesIO()
                with pd.ExcelWriter(buf_out, engine='openpyxl') as writer:
                    # Onglet décompte
                    rows_decompte = []
                    rows_decompte.append({"": "DÉCOMPTE DE RÉTROCESSION", " ": ""})
                    rows_decompte.append({"": "Période", " ": label_periode})
                    rows_decompte.append({"": "Date de calcul", " ": datetime.today().strftime("%d.%m.%Y")})
                    rows_decompte.append({"": "", " ": ""})
                    pd.DataFrame(rows_decompte).to_excel(writer, sheet_name="Décompte", index=False)

                    ws = writer.sheets["Décompte"]
                    # En-tête tableau
                    headers = ["Code tarifaire", "Nb prestations", "CA (CHF)", "Taux (%)", "Rétrocession (CHF)"]
                    for col_idx, h in enumerate(headers, 1):
                        ws.cell(row=6, column=col_idx).value = h

                    for row_idx, (_, row) in enumerate(detail.iterrows(), 7):
                        ws.cell(row=row_idx, column=1).value = row["Code"]
                        ws.cell(row=row_idx, column=2).value = int(row["Nb prestations"])
                        ws.cell(row=row_idx, column=3).value = float(row["CA (CHF)"])
                        ws.cell(row=row_idx, column=4).value = float(row["Taux (%)"])
                        ws.cell(row=row_idx, column=5).value = float(row["Rétrocession (CHF)"])

                    total_row = 7 + len(detail)
                    ws.cell(row=total_row, column=1).value = "TOTAL"
                    ws.cell(row=total_row, column=3).value = float(round(ca_couvert, 2))
                    ws.cell(row=total_row, column=5).value = float(round(total_retro, 2))

                    # Onglet données brutes
                    df_f[[c_date, c_code, c_mont]].rename(columns={
                        c_date: "Date", c_code: "Code tarifaire", c_mont: "Montant (CHF)"
                    }).to_excel(writer, sheet_name="Données brutes", index=False)

                buf_out.seek(0)
                st.download_button(
                    label="📥 Télécharger le décompte (.xlsx)",
                    data=buf_out,
                    file_name=f"retrocession_{label_periode.replace(' ', '_').replace('–','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )

        except Exception as e:
            st.error(f"❌ Erreur : {e}")
    else:
        st.info("👈 Chargez l'export Prestations du/de la thérapeute dans la sidebar pour commencer.")


# ==========================================
# 🔁 MODULE POSITION 7350
# ==========================================
elif st.session_state.page == "pos7350":
    st.markdown('''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IiiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NuctcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo837nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++fill81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQtalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr73xyx9tunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.markdown("<style>.block-container { padding-left: 1rem; padding-right: 1rem; max-width: 100%; }</style>", unsafe_allow_html=True)

    if st.sidebar.button("⬅️ Retour Accueil", key="btn_back_7350"):
        st.session_state.page = "accueil"
        st.rerun()

    st.title("🔁 Suivi de la position 7350")
    st.caption("Identifie les patients actifs pour qui la position 7350 peut être refacturée — 36 séances 7301/7311 ou 6 mois écoulés depuis le dernier 7350.")

    st.sidebar.markdown("---")
    st.sidebar.markdown("**📂 Export Prestations**")
    f1 = st.sidebar.file_uploader("Export récent (obligatoire)", type="xlsx", key="up_7350_1")
    f2 = st.sidebar.file_uploader("Export plus ancien (optionnel)", type="xlsx", key="up_7350_2")
    st.sidebar.markdown("---")
    jours_inactif = st.sidebar.number_input("Jours sans séance = inactif", min_value=14, max_value=180, value=30, key="jours_inactif_7350")

    if f1 is not None:
        try:
            import numpy as np

            @st.cache_data(show_spinner=False)
            def charger_7350(file_bytes, nom_fichier):
                df = pd.read_excel(file_bytes, sheet_name="Prestation")
                _cm = resoudre_colonnes(df)
                c_date  = _cm["date_facture"] or df.columns[1]
                c_tarif = df.columns[2]
                c_pat   = _cm["patient"] or df.columns[8]
                c_mont  = _cm["chiffre"] or df.columns[11]
                c_num   = df.columns[0]  # Numéro de facture — toujours col 0
                df[c_date]  = pd.to_datetime(df[c_date], errors="coerce")
                df[c_tarif] = df[c_tarif].astype(str).str.strip()
                df[c_pat]   = df[c_pat].astype(str).str.strip()
                df[c_mont]  = pd.to_numeric(df[c_mont], errors="coerce").fillna(0)
                df[c_num]   = df[c_num].astype(str).str.strip()
                # Uniquement les lignes réellement facturées (Chiffre > 0)
                df = df[df[c_mont] > 0]
                df = df.dropna(subset=[c_date, c_pat])
                df = df[df[c_pat].str.lower() != "nan"]
                return df[[c_date, c_tarif, c_pat, c_num]].rename(columns={c_date: "date", c_tarif: "tarif", c_pat: "patient", c_num: "num_facture"})

            df_raw = charger_7350(f1, f1.name)
            if f2 is not None:
                df_raw2 = charger_7350(f2, f2.name)
                df_raw = pd.concat([df_raw, df_raw2]).drop_duplicates().reset_index(drop=True)

            df_raw = df_raw.sort_values("date")
            ajd = pd.Timestamp(datetime.today().date())

            df_7350   = df_raw[df_raw["tarif"] == "7350"].copy()
            df_physio = df_raw[df_raw["tarif"].isin(["7301", "7311"])].copy()

            # Associer chaque 7350 à son type (7301 ou 7311) par co-occurrence le même jour
            # Un 7350 facturé le même jour qu'un 7311 appartient au cas 7311, idem pour 7301
            def type_du_bilan(pat, date_bilan):
                """Retourne '7301' ou '7311' selon le code facturé le même jour que le 7350."""
                meme_jour = df_physio[
                    (df_physio["patient"] == pat) &
                    (df_physio["date"] == date_bilan)
                ]["tarif"].tolist()
                if "7311" in meme_jour: return "7311"
                if "7301" in meme_jour: return "7301"
                return None  # 7350 isolé — type inconnu

            tous_patients = sorted(df_raw["patient"].unique())

            cas_36     = []
            cas_6mois  = []
            cas_manuel = []
            cas_jamais = []  # patients actifs sans aucun 7350 dans l'export

            for pat in tous_patients:
                bilans_bruts = sorted(df_7350[df_7350["patient"] == pat]["date"].tolist())

                # Séparer les bilans par type via co-occurrence
                bilans_par_type = {"7301": [], "7311": []}
                for d in bilans_bruts:
                    t = type_du_bilan(pat, d)
                    if t in bilans_par_type:
                        bilans_par_type[t].append(d)
                    else:
                        # Type inconnu → attribuer aux deux (conservateur)
                        bilans_par_type["7301"].append(d)
                        bilans_par_type["7311"].append(d)

                for position in ["7301", "7311"]:
                    sub_pos = df_physio[(df_physio["patient"] == pat) & (df_physio["tarif"] == position)].sort_values("date")
                    seances_pos = sub_pos["date"].tolist()
                    num_factures_pos = sub_pos["num_facture"].tolist()
                    if not seances_pos:
                        continue

                    derniere_seance = seances_pos[-1]
                    jours_depuis_seance = (ajd - derniere_seance).days
                    actif = jours_depuis_seance <= jours_inactif
                    if not actif:
                        continue

                    bilans_type = sorted(bilans_par_type[position])
                    dernier_bilan = bilans_type[-1] if bilans_type else None

                    if dernier_bilan is not None:
                        # >= pour inclure la séance du jour du 7350 comme séance n°1
                        seances_depuis = [s for s in seances_pos if s >= dernier_bilan]
                        jours_depuis   = (ajd - dernier_bilan).days
                        date_facturable = dernier_bilan + pd.DateOffset(days=1)
                    else:
                        seances_depuis  = seances_pos
                        jours_depuis    = 0  # pas de 7350 → géré dans cas_jamais, pas cas_6mois
                        date_facturable = seances_pos[0]

                    nb_seances_depuis = len(seances_depuis)

                    # Détection reprise après pause
                    # Une pause n'est réelle que si les deux séances sont sur des factures différentes
                    if dernier_bilan is not None and len(seances_depuis) >= 1:
                        idx_depuis = len(seances_pos) - len(seances_depuis)
                        nums_depuis = num_factures_pos[idx_depuis:]
                        seances_apres = sorted(zip(seances_depuis, nums_depuis), key=lambda x: x[0])
                        pause_detectee = any(
                            (seances_apres[i][0] - seances_apres[i-1][0]).days > jours_inactif
                            and seances_apres[i][1] != seances_apres[i-1][1]  # factures différentes
                            for i in range(1, len(seances_apres))
                        )
                        if pause_detectee:
                            cas_manuel.append({
                                "Patient":                pat,
                                "Position":               position,
                                "Dernière séance":        derniere_seance.strftime("%d.%m.%Y"),
                                "Dernier 7350":           dernier_bilan.strftime("%d.%m.%Y"),
                                "7350 facturable depuis": date_facturable.strftime("%d.%m.%Y"),
                                "Remarque":               "Reprise après pause — vérifier si nouveau cas",
                            })
                            continue

                    if nb_seances_depuis >= 36:
                        cas_36.append({
                            "Patient":                pat,
                            "Position":               position,
                            "Séances depuis 7350":    nb_seances_depuis,
                            "Dernière séance":        derniere_seance.strftime("%d.%m.%Y"),
                            "Dernier 7350":           dernier_bilan.strftime("%d.%m.%Y") if dernier_bilan else "Jamais",
                            "7350 facturable depuis": date_facturable.strftime("%d.%m.%Y"),
                        })

                    if dernier_bilan is not None and jours_depuis >= 183:
                        cas_6mois.append({
                            "Patient":                   pat,
                            "Position":                  position,
                            "Jours depuis dernier 7350": jours_depuis,
                            "Dernière séance":           derniere_seance.strftime("%d.%m.%Y"),
                            "Dernier 7350":              dernier_bilan.strftime("%d.%m.%Y"),
                            "7350 facturable depuis":    (dernier_bilan + pd.DateOffset(days=183)).strftime("%d.%m.%Y"),
                        })

                    # Jamais de 7350 dans l'export pour ce cas
                    if dernier_bilan is None:
                        premiere_seance = seances_pos[0]
                        anciennete_jours = (ajd - premiere_seance).days
                        if anciennete_jours >= 14:  # au moins 2 semaines de recul
                            if anciennete_jours >= 90:
                                fiabilite = "🔴 Très probable"
                            elif anciennete_jours >= 30:
                                fiabilite = "🟠 Probable"
                            else:
                                fiabilite = "🟡 Incertain (export court)"
                            cas_jamais.append({
                                "Patient":           pat,
                                "Position":          position,
                                "1ère séance export": premiere_seance.strftime("%d.%m.%Y"),
                                "Dernière séance":   derniere_seance.strftime("%d.%m.%Y"),
                                "Nb séances":        len(seances_pos),
                                "Ancienneté (jours)": anciennete_jours,
                                "Fiabilité":         fiabilite,
                            })

            # Métriques
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("🟢 Critère 36 séances", len(cas_36))
            c2.metric("🟢 Critère 6 mois", len(cas_6mois))
            c3.metric("🔍 À analyser manuellement", len(cas_manuel))
            c4.metric("⚠️ Jamais de 7350", len(cas_jamais))

            st.caption("⚠️ Si un patient a plusieurs cas actifs (ex: épaule + pied), les séances sont comptées ensemble — le module peut sous-estimer les opportunités.")

            # --- Calcul erreurs facturation (pour le compteur d'onglet) ---
            df_mix = df_raw.copy()
            has_7301_num = set(df_mix[df_mix["tarif"] == "7301"]["num_facture"].unique())
            has_7311_num = set(df_mix[df_mix["tarif"] == "7311"]["num_facture"].unique())
            nums_erreur  = has_7301_num & has_7311_num
            erreurs_fact = []
            for num in sorted(nums_erreur):
                rows = df_mix[df_mix["num_facture"] == num]
                pat  = rows["patient"].iloc[0]
                date = rows["date"].iloc[0]
                codes = sorted(rows["tarif"].unique())
                seances_pat = df_physio[df_physio["patient"] == pat]["date"]
                if seances_pat.empty: continue
                if (ajd - seances_pat.max()).days > jours_inactif: continue
                if (ajd - date).days > jours_inactif: continue
                erreurs_fact.append({
                    "N° facture":     num,
                    "Patient":        pat,
                    "Date":           date.strftime("%d.%m.%Y"),
                    "Codes présents": ", ".join(codes),
                })

            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                f"🟢 36 séances ({len(cas_36)})",
                f"🟢 6 mois ({len(cas_6mois)})",
                f"🔍 À analyser ({len(cas_manuel)})",
                f"⚠️ Jamais de 7350 ({len(cas_jamais)})",
                f"🚨 Erreurs facturation ({len(erreurs_fact)})",
            ])

            with tab1:
                st.subheader("🟢 Critère : 36 séances 7301/7311 atteintes")
                if cas_36:
                    st.dataframe(pd.DataFrame(cas_36).sort_values("Séances depuis 7350", ascending=False),
                        use_container_width=True, hide_index=True)
                else:
                    st.info("Aucun patient n'a atteint 36 séances depuis son dernier 7350.")

            with tab2:
                st.subheader("🟢 Critère : 6 mois écoulés depuis le dernier 7350")
                if cas_6mois:
                    df_6mois = pd.DataFrame(cas_6mois)
                    df_6mois["_sort"] = pd.to_numeric(df_6mois["Jours depuis dernier 7350"], errors="coerce")
                    df_6mois = df_6mois.sort_values("_sort", ascending=False).drop(columns=["_sort"])
                    st.dataframe(df_6mois, use_container_width=True, hide_index=True)
                else:
                    st.info("Aucun patient n'a atteint 6 mois depuis son dernier 7350.")

            with tab3:
                st.subheader("🔍 À analyser manuellement — reprise après pause")
                st.caption("Ces patients ont eu une interruption de traitement et sont revenus. Vérifier s'il s'agit d'un nouveau cas (nouveau 7350 possible dès la 1ère séance).")
                if cas_manuel:
                    st.dataframe(pd.DataFrame(cas_manuel), use_container_width=True, hide_index=True)
                else:
                    st.info("Aucun cas de reprise détecté.")

            with tab4:
                st.subheader("⚠️ Patients actifs sans aucun 7350 dans l'export")
                st.caption(
                    "La fiabilité dépend de la période couverte par l'export : "
                    "🔴 **Très probable** = 1ère séance > 90j · "
                    "🟠 **Probable** = 30–90j · "
                    "🟡 **Incertain** = < 30j (le 7350 a peut-être été facturé avant la période de l'export)."
                )
                if cas_jamais:
                    df_jamais = pd.DataFrame(cas_jamais).sort_values("Ancienneté (jours)", ascending=False)
                    st.dataframe(df_jamais, use_container_width=True, hide_index=True)
                else:
                    st.info("Aucun patient actif sans 7350 détecté dans l'export.")

            with tab5:
                st.subheader("🚨 Erreurs de facturation — 7301 et 7311 sur la même facture")
                st.caption("Un patient ne peut pas avoir les deux codes dans la même facture. Chaque code correspond à un cas distinct et doit être sur une facture séparée.")
                if erreurs_fact:
                    st.dataframe(pd.DataFrame(erreurs_fact), use_container_width=True, hide_index=True)
                    st.warning(f"⚠️ {len(erreurs_fact)} facture(s) à corriger.")
                else:
                    st.success("✅ Aucune facture mixte 7301+7311 pour les patients actifs.")

        except Exception as e:
            st.error(f"❌ Erreur : {e}")
            import traceback; st.code(traceback.format_exc())
    else:
        st.info("👈 Chargez l'export Prestations dans la sidebar pour commencer.")

# ==========================================
# 🎓 MODULE FORMATIONS
# ==========================================
elif st.session_state.page == "formations":

    _logo_b64_f = "iVBORw0KGgoAAAANSUhEUgAAALYAAABQCAYAAAC07Y+bAAA5gUlEQVR4nO29d3xU1fY+vNbeZ3pJgQRCBwsK6lXBhmUSG0WKijNeew9YroqAKCgnI6LYRRQlNvRr4c4IgiCgoskgylVBvAiIBUFagJA2feacvdf7x8wgIpAJJMB9fz5+hvjJzJmsfc7ae6/9rIbwN/7GYQqf2809fr/4kMhqf/TFm6lq+1V6LHGcHo1a0GQkZjCuV/JyFiinHje9eOiAbwkAkVL/4qEW/m/8jT0ho9Tzny4/2/DLphd4MHS8nkxCUhcADAGIQEEGRoMRdIsxIQpaP9HvsdHjCVPv/a3Yf+Owg8/t4x6/R3z88HP9Db9tniuDYRaXUiBDBAQGlPocARAQSJSC59mdEMp3zL7w2Qcv8Xs87G/F/ht7BBFhWVkZlgGAv2dPdGfeWLUqpVZlZYSI1Nx/V1VVVub10uJPv+ganfHRMqhryNUY6gig7FVWAEIptTybw1jXJu/hi54c++BeP/w3/t8CEaHf72cFL6zCyoBXIqIEAPLu7QJv6p0Kl0uB4mKoBJBer1ceqBw9V69GRJDzK5ZMtEcTufUMdLYPpQYAQAAkxgwN4ZDgHO779IN5s/5esf8fB6kqq6ysZCWBgL7zlwiwTpJ507JVhYlvl9sNSdGBHHaTgRMlfvy9Ktn76Ejrk3pW9T6mewNof1xGAAhuHwOfW+7Pak5EiIi0cMmKNomX3lyLobCVGEtLlM31IGwGI491KZqeueB/ScGbfftrKlQA5t0POQgA8DCQHyC15YPXC14ACQCwgciy9rWZZ8oNm4qTDaFTIRbvTolEIerCbEIOTOEABJBIxEE3cELAGmazbOAW63LutC9hR3VdVHztkF9ApBbtCpdLKa6sFE1R8ApVVUq8Xv0TdbLH+NuGf0ciUQEMebbXEwApBCBb5WxUAAFAUVLHyMPilu8FGfl0AUCHTlACQASQYDQ0TQ4EQCJQk7riRdQbv6BlQETo93iYx+sVgAAL3559Aq365YY1pWOH8KTe1aAJYLoOmtBBJwIJBFEAiZQeLUOGCULGWGsllmhtUMInw7aam+Jrf08svHX8f1jbwnei4271lyDWAeLOg2BTZExK/WizBCCEJtF2iIhCCCBN76h8MGlKO+O2hveYlBZBRIiH3+pNBMQQUZLUyECDBz7p3aqqKmsOm64JQJ/bzXDWTLHg3kkvG+Na77jUJSAwAEive3sGMQBJoFsZtyanvvkxEY0qKyuDgyw/+NxujogCAMSCV2f8Q/nld69YsOgisy6VeCIBMSlkjDOJiAgICIgIqX/+bA8gkg4AOkmKa0kJkhABTEoi6eL1IZdx+AMPLhz96It87PCpJXl59QTA0nZGViuBLgTfLyOCCIAhYCKJimXpf026rpxuVAwoDlP+jwCAIUJSaCAgaToUMkwrLVU85eXagvsee8yxqfrmaDQCRmR/+szu9y5teoAgAivjEHFaN3FpeBIQqYwIvN69Hs2aHRleeB6RUxn3pEqLlt6hxJPGhJaEEEMdERlyzgD+oNP2AUyPFQGAQYpcozgJSdEoKNFoB0s4PjFx7xM3fTrh+fGo/uttQARSVYZZTGaz3RIB3A/7ARFASECrOalIxgkUHo6DtEkEAjoMnTaIhEAoGUsSKgfdDknbftr8x14ablm19t6GSFgjzhiRbPxeIUiDBIzm2Ori7uJzhvTtu0VVVeZNsQ4HBRUul1Li9+vzX3r7DOV29XVzfbh7MB4FDZlAxvi+qLQmAIGAI0PQAaghHhXGeKybKam/9cmdEwbhFRfcjqedVkNuH8e9mCbVPXsSAACSsiSJhEDEdk6hLEBE0qQoKBTDSpb+DcPUdsEw5dc5rF475SJijYyt2ZE50Mx9etr5xjXrno+FwkIyVICINyo3AHIhGXNYIdm5/dAhffuu87l9/GCaIBUul1ISCOgfTXrpBsO3PwRg247u9YmYDowRImR9MGsKUvYLKhqCDIWCunlbzeU0/cMv58yYeRz6PaJCVfc4kTwejyAAPHrs8GUJo7LOhBxon0bebpBEismE2KndGwddUf6X4PP5eInXq899Y9axph/X/1tvCDKdA2KWBiBKKaxOJxMndL9l4IN3LKpQVaWpB6kDQYVLVUoCAX3h+KdV2+q1ryVr65UkA8EQFTg4TBhDxpSGZFxnOxq6WxZ+++Unz79+XonXq+9NustcKu+KGOcFrSdYbFaGUmZ10CYA3cwYj9lMP7e6/9ZX/lbsvUBVVebxeGTF0qWtDf9ZPpvVhfI1RImA2d0zSZrT7lAiXdtP6DfyltczK38Li70TFaqqlAS8+ifjnxxn3rCtLBIKCckZALXMKr0vMEQlLqUQtUEn//7nOZ9MftVV4vXqPp/vL7J4A17d53bzfk/c93owzzYz32o3kpTa3lZuAiBJpBsJFHDaheGEY6/ujRj9W7H3ACLCnt7VuJRIib/x0UxTbfCoGAmR7dYtSepOs8UQbp3z1oCHRoyvcB1cpfa50zvNpOc9ht+3PRwMNujEGct2p2kJIALXEISoa7DgyrX+ee/OOsLj8QhVVf+ig26fT6pCsNxn1WtChbn+HLvDYAJkJIQkIn3nS0rBJGGO0axgnnMbHdttyAW3X/Otz+3m/1OK3SR7a/+BlcVl3MPeE9WjJr5mr6k/J6QndcTsHAUEpDsUoxLJdwY6PjXuJp+QvLiy7KCZHyn2wyPmvzO7p+mXTa8nQiEpUwfEQ04KIAJPAgmlIVyAi76bQUSGnqtXI+1GWGCKNaI+iLELn33QEzvuqFIqyF9ldjqZ02xVnJbUy2F3cOawNWgd204PX1R8ygWjh32UYX/+Z2JFiEgqjJnjVmOLTsYKVeUlXq8+b9wTDzk2bLu6IRHXGKIhWxlNyJRknmO9vHGw+zjEpKqqzNMCwUJ7+fvo93iggkiJ3T7+TR6MWBPIRIaHbtJ3AUggkkiARJJlqMsUS4ESGBAiYlNNG0TkEaFpOaFY77llTz/m8fvv8Xk8HADEbp8jIkJABLy39OUKotfhzfdPF79tODFuMuZxIGk0mX629D5xscvVuwrgD0oToHlonhYHEQmHwaRECnLmt7mi7za1Sy7zer3NriwZO/ijR164wbZm/YMNkbAOjGWn1EDSQACQ6wiJc04cOODkk6t9Ph/3eA7eYdHv8TCP3y8WPPj0GEcwenK91HSGrGnPmEAQSWbmCjOajExnCLrCgWOKWJZCgFECZ1KCltQgJjUAZFmbaQAAiKiEIhFh+n3r3Z+84n/nwpvdy/bkocw4dHxuHy9JeWsXp19/gs/t5m6fT6adTwAAoBg5p7gOOgHpQEDUNC9mNqM4oG2QCISNKUo877nC8lyZpw9ijIiwuRU7o9TzX/y/YsO3K1+JhkKCOOPZCE4AxCUQd9h57Lij3YOudK862Eqtqipze71y4eyFbbTZnzwQiUYlMtakOAuQkmwGI9dNRhB26/JYft7HmGtdgh2KqvKsFtC4QvVbd3CtpvYIrbbhRFbXUGyIRE8xCeLhRJyAMcjyWaNkDAyxJEa/XT4JEM8H8O/1wx6/RwAR+vx+VrBq1Z++v7isTCCi2J3vVqTJyI1EOVbFAM3teUREiMTjoJHcr+8lIGkE5HqeY6sywDW4BDFMqsqwmZ0bPp+Pl3g8+qw33j2aVyz3i1AYBcdsHxKAlMJqdyihIzoMGzTixo8rVFUp8XgOajxIcSUwBNDnf/blCFtMszdkEe6ZAQFJJoFZLFbU27b6RDmh5xPnlnoWUiK5t0u+BoB3wKDA58+8dlrytw13m7bV/FPGE6ABScTGmSNE4FEtKUxhdt7Hj7x4Tt/7hy/aZ1wJInl2M1cAYGf47O5QlMGuWlxbPU7XhJFSOHDdZgBEknOjOQLbd1xj/G1TDw1IQhNsvdQqSIC5zrh24rGXXDjwvN/J5+PYzKtghtZbsGFDPk56+QMejbWOAQiELA+LkrRcm8MQ6tLuyYHj7yo/2LQewM5wT33Rzz8XhB9/bXg0ESfMMiqOgKSRkLE8R1w/svNdFzxwR/kfEXqqUn17zz/vjH6Agh6rECoBSgJece4d134NAFd8+tjUt9kvG1811YcKEySyUm5iSIogSGzZMhIAFu1r1W4qWvSkPPeFN/uYv/lhvhYO2wVi1o4NSGVECLPDoUS7FV08UL1ndksoDBGhHz3MTT5cMOrRBdbN284Lakk97cBoFJJIzzGalUhR/swBU7yXfd5nnFJcWdakUM3mQObeLCh75lbr2s1Tg9GwjlnY1imlZkzmOrYmTzn20sHDr1viAzcHnxuyNaNIVVn53Co+bFm59pl/dk9a+M0nYkddOx1JQuOcP6EkQLs1Yb1qcA9XX9e65gpuUwAAK1S12Uj7WG0+7z/lruS8N/y9jIFlC2UkahGp8MPsJ5Ek3WF3GIJdi+5uKaUGAEjTevq80Y+85qiqOa9BSzaFARF2blDirZxfFzz1wLXjn36QVRaDLDnISg0AUOwFCYyB3Frj0RNJQsYaDSHKnAuY0xqnY7pePHj4dV8vLZ1m6F0+TANP9itnOqhJLi0tNfR2D1n1wSvvDLQuXr6EBUMGyXAnmbK3yyWSnkNojn713SAAeK44FeveLIpNzaU0qqoy75S7xMIVKwqVae+9h8GQJUEksuWAAdJbu8VqCLZtNfki7z2TW8q5sZMBGfv4g45N1TcE4zENWXZKLYmkGRjX8py/s4v7Du2NGCWVGHoPXmBTBmo6Yi5QESiKvDbn1LiuIbAsvKNCSpvTycNHdR4xYPSwr32qauztHbZXo7ox9C4v15aWTjP0vvnK5R89+NQkx3pSg5GIaMwkQsZQTyZBr62/EBCfK/Y2j6+i2TjhNFMhiQjE9A/fVbbXdY6BTKTjuyVlkcYgSeo5ZrMh3K7V7P6PjRlBl7lbxLmRUep5E1+42rZ5x0OhcDhJHJEA9D95toh0AtCBQGTGQADSIAkwzxkVp3W/9Pzz+2z2+Xz8UCg1AEBx+hlGP/76VIsurZQ6YO17d0QQVoOBR0x88UVld02pcKmKx+vVDlSWOUVbBKkqa/PQPc9EzMo2AwJv7LkTAUvqOlAkdsrHUtoQvLJR+bNAcyk2QlkZVqxbZ15w76Oz29RFSwABnIrJZEbOjICMS0ISUhKRDghi9wFLImHnRiWSn/N1h8fvv7oMESHFTTbr1p4JbFrwwpt9jL+sfyXW0AAKY0YbKopDMSpOs2WnZ8tptioOg1GxKgZuBGScCLkuyJCbw7QTjrqm/43XfFehqsrBpPX+gsrUDxZLns4JgLKIpiYhEYxGwDatHiJNh+rCngTNkD/l9XplZSWw3ogN0mZ502KyAADt894gAGokSUlqBcYXph8FAEDqgRMYzeKgISJARPnxJVcUQWHrnzZuqPrO3K6gQ4xzJ0Vi7UHTWpOQ7S0ENoMuma5pkJAaCAIBDBAQyQKMJ/Ocv7PB5196HGLY5/bx5qb10gyImDf51R5s+epFJl3yRK4jAibjRmE2bUDOt2mx+O+sYxFwxkDbUQemeLKQTIa2Iho/AiKx9jaDITfapd3I/nfeOOtQMCC7o7pwNQEiCF07XojGF2sAkEbGWcxqWo8PjVxEE0bh3uKj90ue23sSBQA/6tp+frI2NBrilElE2DsQhAW5Eq9tOAYAvq+EMgYHaGc3i2JnVtW+Jx6zDgBG/+lNzoB0wZcs+b6ttuaX7on1m0+juvoSDMdOsWsiV08mQeg6UJ4zJE7pccmAC8/a0hK0XprGpPOHDs0LfxgYay4sfDbeJm9hwYk9V51yzimbkaHc15qFZhN8VVHZJvTrlq79rxn6n3Q2yCFVagAAt98vkTOgRKJIkoSUm3vvAyEAaTIYmWYzzx+AmFBdLgV2zVA/UHncbokAVHHC6T9Gl6+JckBr4xqKgESAjB/RXHIckGITAZaVuXYeDoqhGNavBwVgPdiiUdktL496lZeLtKtzc/r1OTB49OslK9uGKr64SPy+5XpF03tpx3S9csDNVy6vUFUFW8C5kZl8X23YEO837o7rcbeE2gqXS6kuLGQRq5UBdAGA9RB0OunMREL+VlcnPX6/OOOMM7YBwDYVIKsUp5ZGplzBUl1Yq0vvb6MLAYSNrI8kARUFlPzc/wIAFBcXgzcQaDaZMve5uLhnzfzptFnh/KgEiX2H+yIAEEFy49ZU2l/lgcux34pNKcciAfwx270QAADYk1Jiaa9SpWcfEyvKz6dVXq9+2mnHbQWAV4GzV+e/91GX/pf0X6+qKmvprb1Pp04xAIB5//qX6ZdfABYv2Kr7wS9KGlm1XOBSriztjgAAw8rLD/lKvStWA2Cewg1GImgslQoJUZc68Fh0AwBA9erVzU5PEgCiwrW5w8YFWTQJjZtHCFIIMNusnQHS5tUBYr8UO7NSrFv+TO7KOb+dH9EkMSKUKImZbRiqSkSYrK065uwTEqf3a1OltLm5vnxZuQbL/hjJtNJSw9FFRVTi9eoZpW7plClVVRlUVjJvIKAPmDIlAQAACgJpHzgCr6zo8NNX65w5HVp1YGbChs2xaE4ubDlj8Ik7Opx2yxZkqAfKd65suEvgzWFRtIIkZJGACQAIKDQNtB9+rWtRgRgDZjI2qUQF2i225vrzTVZsIkAPIiMinHjRxR8kauIuXco/FgpE0DUJhBLWLK+muVNxx7gzL9xgNBpWtjuiYGn3M7stPPuGB9cMKy/XAABUl0spKy6WLbm1pydNptyBJPrK8tbI6X12bA5dVL+jofe4s6YeJYRsIxISDSs3AzIAXZMAnOD7Rb9GjebZGx8bcsmy/DY5i47s3fmzc2+b8KvH7xeAmJK/MiBwf7KqmwkdATDCkGcjABFJo9nC2OkndYI5r/ynoEePFvE+U1Jj826815SqftropwGRgb6jdgsAQMH2A5epyYpdVuzifr5If/bKy19p2BRyRRKJJDLguwqfGgsyiBAyxAKGrEBhrFeoOnLdmmW/6+PPGbCksHPOm7e+MuRtRE/MGwhAS6zYRIB+jztVHIYjzJ04+tgfvvr1xgdd3qHJiN5VahJ0IUFICZIkAEOKxRLyT2MIkZUj627gvHvt78Erf1m+KTHhgoGLC9rmvVH66q2zEU8PevHPscAHG8UA2oKkFkbGCtInx70rBiKglKBtr2m3z8/tJzK7ecU3q9oAUTdN6ICNOYwIADkDysvZAQAAxQBwgGZ/k3hs1eVSvIGA/tzVV47e9lPtTVEtoXMDMzLG+K7/IWOcMUTGGQFDEkgySboeicX1SENcqd8SPPv35VtfHtvn5e8nX335rRnnjs/tbjbXvqqqDBHI4/eL+c+M7v6s+58zvnj/u+9r1tWNClZHu0aiCRnXNV2AFMCAGGfEEPEvY1AYEYOU/ImEHg3FTTXr689bu3zTmxMufHjFGyNuGElUYff4/cINwJsliCxLYKbMOecJaTJu44wBNGYaIYLUdKB4zAUI1Nw2dmVZGScATHxacaZNgFXuKSJvDyBEYAbDxuaSI2vFdrvd3BsI6G+OLnVv+6n68VA0pjPOeCPbDEJq8WMAqCBjCuNIAklE4gkRaYgfXbW6ZuqECwZV+ifd3dPj9wvV5TpgCtLndvO0F9T8wnVXTQjMWL58ww9bLg8H48a40HRiIBlDhqlgJ56Rc49fRumiMIAKIirIkHQkEU0kxPaN9Z1/XLThSe+5T3z72h2ll/sZCEQkdzNO0MZQ5nJxIAI0GzdwxoCoUaOWx/UkyGDknIradbluv18252SsXr2aEIHElu1XgqYBsCzOIEQ8gQDYqeNKgD/qixwIslJsVVWZ3+8Xb48dduqPi9e/WV8fkUxhHPYvfBsBgCMiJwQZisX02i2hc5bPXr3k5VuvudwbCOhuAE77uU2qLpfi8fuFb9KI4x8dcMmXG1ZsfSDUELNoIAVjSGllPhCP6075UUEKReN67dbQMT99tXbGpIGXvrN2xTtt/H6/aM7dZ18oLi5O2VyK4XvGlUYPawiAOoCwJ/T8+IS3L0MAqiwraxZZU/HUfvnZlHd6GkLRQZFkgqARc5cAyIAMdZOyreCGS1cDAHg8ngM2SbN7wJWVDABg2cc/DdfDmpkxJvZTqf/y9xljikAhgnVRx9qlW2Y8d/WVD/g5Ex4A1tS8h9JevQzeQEB/7e7rB/13zurF29bXnByOx3WmMII/VubmAwEyxhRiIMPRuNixrvaKt0fN+Nr/yF2nNdfuky14x6LvdI6AWay+iIDJZJKwLnjfz0Smau/qZonDL+ixCgGR4j+uedqQ1AyArFEFRQRhUgyENuviExEj6QXh4KzYGSRjWgKwBdLHIL36heKi6sfqCZM9lz7t5yA86GbZrtyqS1XKly3Tnr7qmut+XbJ5VkNNxCkZCcaY0kyTcF9gjDMeF7pevaG28/I5ayrfGHXzEG8goLe0cheXpYLE7CW9vk1wrOOIjQYeASBLCCEtCf2IX0Y/8rgH/GLZsGEHJOfOilnjnrje3hC5MKppArLIgyQhEUxGZO3avg0AsKqZWJomKXYqV6CFlIQAmcJYMBLTqtbUjZjsufhpP/eLYmj85qQOtV79zTGlg7b9tPnVhoYYB44SssyCaS4gogIKimBdxPzT4vXvvTHq8hZXbkQkH7h5n+OPr4VWuYtN3ECIjcdZIEMejkWFrbrhznlPTruxd3m5Nq1XqWF/Vu5MYNnHM+YcZ9q8fUo8GhOUTehsOm4lzGi9Y3S/BQSAZV5vszBLh1ddEQJkChrCsYRW9VNoxLNXXP1UgOE+FcOXPtTOmzzmpF+WbHw3Ek4ypiDtT8mBZgEBBwVlQ12c//Rl3XvvPnhHf28goLekzV2g9kjto3k5b6LJiCSyM1GJMxYLhoTxh19f/eiJl28YtqxcQ0SqUFUlGwUnIpxWOs2wyuOhT9euzaFF387CcNSuIWWVLSVJSovFispRXab1wU6xSlXl2bDe2SDLbT5F8/2r51kvmblhWEzXdNx7+tSuReT3b3VHAKmTluOwGLr07nhn6YuvTXG73dy/G0+c4b6Xzn+46KOnlny9Y1OwI3AQkMUqv0/Z4QDlBwBCkKAR5rdzNPS//YxTz7hs/C8t5V1Nc8dQQWSLDh/7s7Kjoa2echg1nncIQExIMjsdTO/abqpBvXtcCWI9QGrRKOjRA6t79qRMc6XKVal8x8qAV+7MdDEZYcFtD3xq3rzj/LDQsk0skYokkK3zqu3TJh57FmI9pKJEm0Wxm2OLJAASJAEJiCMhpsNYgZAIEdPpjk1QNgJgCioNwai+7vtNk18dccuKm555ObCrYmQeJtFXlon9Hnmvfku4I3HKOgl3d/mlJERIlfcEIsBMdhWC3mT5AQAJGCggQtsiuZ+/tvQtIjrTg56M86RZuePMKluCGJ4/4fnnrZHkxIZYVGSVLQ6AkjOIBoPS8Yu4LX7bg/0+fXjKY8q4O3wZBd8jGMLS+obW9a/6TxGbtt1kqtpxfljXGs2YyYBISqvDqYTbFqhnI9ZVuFSlpBk7Pez/is2QkyQBEhSTQQHFwIEUBG6AoGI2Si2WVGQS7KBLSCZ1SEpBvKl2L4LOdFBy2ztWlS386IQyRPACSCLAYnTxxYbF+sT+g2bUrG+4PKprWSfhpkFSSsmAcZPCwWBWIElCmCzGEDcaQIvGUU/KHC4RkgkdElIHzlE01W6XJHWbwaQU9SwcO+KdGY/uaedpDmSqJlWuq8uJP/rEGr6jrkBLxfllXxlAkjAicrPNBhEj3wJGw+c8J2cptGu9zmmzYRQAYr9tsBmITqRw9GQZDp9oSspWkNQgqiUkMpbV3yICYWWcxwpyv7K+8JCr2uMhz3t+0ZzTff9WbARBOilWs0kxOI21OfmW2Y7W1s86HNF+Za++7bYUndRD/rjgB+PywNouOzbWntFQHb44FoyfHQ8leVIKyVgWh1AEkppER74NjuzdYQxjTI5PNwQqK3bxAAb0p4deNmnbmrrLY5qmsSzzFVMgARK4w2LhzMq3OPKsH3Y8uuDTnI6FPwy8tHMNdOwLP1X62Nfzfu4S3Bo5o357cEg0GD8vEUzyuNAlVxhmy7Qwxng0kRTVG2ofWPLuo++eccX9v7eESYKI5HP7uKdrXv2CCVPuN8e015LRkI6YnbIBpA6UGoBMhsNkQGxnNhivxobo1WLLNkgwBgwAbJoOTBLougZJoUOUSCIyyrY4T6oBEoFutyb4P465qQRRJ1XNpotCk9BkxZaSBCdmchZaqrsc32nqoNvPmVbY85qqnYLd/6ePVwHAEm7kT384adS533/64731VaG+sYQGuK8MZgQgnXSnw2rocHzhsGseK//IDW7u9XpFZvd4/qarhm3+77YxaZ7akNWNQQASJIyMc0u+aetRJ3d69roJl72Muf1qd37m7j9dsQMAlqKBTamYMu6Mb+avuq92c3BwOBwH5I1mYKdAqV6yyQbN+tlbX98HCMNXr17dIgdbj98jfG4371d21+sfDRt7uUPT+gb1hM6gSTsZQ4agAZCmJwVoBCjpj8EiAqXr9hEiy8bc+RNICqvNroS6FN190bCr1vjcPo7e5k+ta5JQmibIZjbzdkcXzLz4ntNPvOmFl8sKe1xT5SbgqsulqKrKiAgzL5/bzVWXSxFJwS6657HPH/h0Tr9O/yga5nBaEiQI91rzWJBms5gMhUfmTbrt5bfKVZdL8YN/p1K/PuKmfptXbn8pGI6JtAc0K0ghdavJyFt1zn2//+1n9Lp+yiuPYW6/Whe4FJ/bzXeXX1VVprpcCmmSFQ+fsGTMvPeH9CzpelNuK1tkX/LvDmTIY8kkRRui134z68mOfr9ftFRMyaoePYiEROMVfUsTTusOE6FCqWJFTQKmQiEURFSAM46cceCMA0OOgAoQNLl0nSTSnUaLEm7l/PdFZXc/X+FquUL4TbKx7z31vP9r1yV/+8gPZo0USZEO2cyul5/P7eZ+P4Af/OKdcSP6rapY80FDbciECkqgPyaYJKlbmFEpOCJnxth5c644SztLCUCKLvP4/WLmEyOP++6DlUsaaiI2VBCyMgkQQOpSd1gtStGxeY+P8M0aIzXRpJDTlPx+8AOI95+8t893H66aV1/dkEMMs6qZQkS6xWBUCo9sdd/omf7HMve0Udn3A5lSYR89XX6ucfnPC0QkwnT2R/evQwEiEBbGuN4qb5Vtwp1nbCm4Leqmlotnz2qgZYGAAADoM/S4h+9+zz9SJAWqqsq8gYCerWAev1/4wS9Ke/UyXDnxmQU9Szpe7Mw1x0mTlOkQRUTCCIqS38X5zb0f3Huj0ASrpEqhqirz+P2i6tephSs//mluqCZqB46UrZ0rhdRtZpNSeFTOU3fN8I+RmuB/yJ/dep+SH8S0Xr0Ml4x6/KsTLug+wJ5jjqAkmU2HK0RETRMQrA1eRkToDQRaLP7ck+7zctE9pZ/Lbu1LrTk5nAlxsOqL/wUEIBUiJvKcQRpw5mVnFRSE3GqPFunFnkFWip0hzS8eM/knSNWXpP09/JQvW6aV9io1XDmxfEHbYwqvcubYudQlAYJghNxRaN3Qf9g5lyL2iamqCoAIXq8XiMj4yh0ff9hQFeosOWVd85mIdItiVPI65nwwetYHo1wCFCLa777fw5Yt06aVlhrcDz79Vbuebe62WkychMzmu5hGghIRceLHT43tDgByT9X8mwuZPi99vSOmJ7t3Hm7MyyVFEqM9p+61GNI1GKXidEDy5OMu7ze435oKVVVaOme0afHYqQdxwLOsfFm5Vtqrl+HO6e/M6tq7aLjdbmIyIcGRbw2f3PfYi08aNHJzJvS0GFycG7l80u15q2FTw2lJEnrWBSOBhAG44iyyLbvvwzuvEZpgxap6wNvfsPJyTXW5lLvfnPGKo61tiYL8L4XL9wAEIIE6KevWbDkbAHYGl7UUSrxevcKlKuffd+u05Mk9BhsK8mttyBUiqWdTwKg5gJJ0m92uxDu3uXvg8CsXHKySFU26sc1JUZUvW6apLpdyy9Tp0/I75N7bpl0uHnlK+39efP/Ty1U1FXpa2quXIQAB/cnLhk6s+a3OHRP79Hj+CQQgUQBzFFirSq7qeSliSVhV1WYbQ8/CQhKagGPP6PyQ2WYEIRtPOUREEJqEcF301OaQIRuUBFLKPeCO6+aJc04/XS9q/bnTalcUSUhALargRKQ7LVZDqG3+1IHekc8dzDosBy3bY29QQWUT+MNy5mP39blk1MSvMhxvaa9ehvJly7SppdfevOG7LS+HwjGdKUzJktYj0qTMLbDLkwcc47p07OQlLZS6hUTEx5/bf1XD1sjRkoFsxESSnJC17pr3pfrJ3LOEJprdC7k37Kw9TYSfeZ+7Q27Z/qApFC2IJuKgAwhMRWkwaCadkETCoRh5tDDv0/7PqX39Hg87mMnPhzwIygteKYWAS0ZN/EqFlFKrLpdSvmyZ9tq9wy7Y+N+qF8PhePa0XooBEXanhXc6sd11l46dvCSTfNDcsrvAxRFRt+fZA0ZFAWyMVkMAAoJIfaxAT+rNEnecLTz+dIcuRDiv7K4pyk2DT4i3yXuYt87b7rTauAU5RykRJAkC0uEADppEICzIeaJ1zlrHqBsuB0RY1aNlD4u745ArdgY+t5t7IZX36A0E9M/KH+j525INM8INMQ4KZu3pkyJF67U5upVa+vzr72aSD1pC5mJX6qfZwr7hCkIWiREohARu4AUAS3IAdlaoOihIm2Hkc7t5yWmnbb3w2fEPWr3/Ol72OOJ2vSC/guXlxGw2G88xWhQjIKP9mHcEIA1ASDn2ejj3lIvP7ty5zu92t3hpjd1x2DRX8vj9KVrP65Urv3oqf2ZZxfvhmkg+cBCYZWcqSaRbFaPiKLJPv+ff/ofS5kzL2XTFABAAKOrWtmbbL/VAsQRgo11rEBB1BFjSYmI1Bk/GQeTxMCwq2g4AU4Hj1C9XrOwcmfflaayu4RQtFj/fsHnbiZoQlA4CaxQEQEyS5A67Eu9xxBUDhw5aeajqGx42ik0A6PF6kYjwkYFD/PWbQ0fpKLNvmQEkTMgVZ3v7F2Pnvl/6PSKftnSpXt6i218xAASAG1rzppmmBADB9P8fmmNO2iwQRISVZWW82OsV2LPn7wDwO1jNvk9uum8ZpFKlsgsdAACQUtjtDiXYpd1dA0fectAYkD0hi3YOkFXJkwMEDuvVS5m5Yrn2pOey1+s3hs7VSGTNgACA4BK5o43990Ej+l+GiFqq/EJL23SVAACQiGwW1ASvNSJnAN3SypK93rQE0vdIJyKsKC5Tfq4sw063j680V9efHBR61hF7kqSea7Eq4bb5Uwd6Rzw3rbTUUNIMNbf3F40JfTCUGkp79VLKly3TJl99xbja34LXx5KahixrpSYShPZ8S+jUQUcPOrHvbdszHLgKKnNDC2aLVwIAAFZvrG0ndJEJ7NonGANIJjAC0CvRYnI1EenGp6xk8QS908iJb1trQ32CWkLPXqlJdxrMSrRt60/7PXbfvypcLqV02rRDWt9wr4KrqfeIiAywr7obB4gMA1J+27VXbP5h28PBSFRHBbOn9YQUjhwL69ar05UXjXzmhwwDoqoq84JX+sEvVGgZD9/qQCEBAIWDiTOFvkuZt72BgBgyUAyySjGeEAaAZssYORBUFpdxj98v5o2Z9KRje507pCWy9xekYquVeJ5jlfGBWz2ASJXFxYe8puEeH7jqcileADnluiv++hell81hCpLP7W525cgwIG+Pu+2s9d9VvR4Ox7OvV4IAJKSwWcxKq655d9703CtzMwyI6nIpXq9Xvnjz1ddPHX79zV5IUYjNKTsRoA/8kmixI1IfPzepC4BGQziJGDKwOqwbhC7Bvf8pbM2GClVVSgJefa73mTsdm3aMrI9GmuQEMxAxkZ9TR0POvbgkL6/e7/MddAZkT/jLg8go28ynRvxjy881r9Str+87+dqr/tncdTIoHdj0xatqtx+/WDcz1BA3Ac8yWg/SIahGk9LqyPznR/3bNyWz8mfk9z10Z6/1K7a9/Pt3m172e68/u7mzxcuKXRwB6NmrpgyVUdGGQIpdoxT3BCIgRWFgdlhXAQH0cLkOqYNsZ4Opx6cOta+rmhwMB3VoSsKAJGJ2q9S7tbukX7+SXw92N+J94U8PIhNFt7Li8bb/nfvT3Eh9zBaOJcS2X6qnLp//aBdvIKA3R+BOutMV7djxf87P3/1udnhHpJBY0wKbzNygtO6W9/Fon+8uF7mUskBgZxRgxcxJHVZ+/vOsaCjOI8GY/O+nW+f5J97fbMqdic4jqrLVbg6WRaLpFnRZXMcMDPLa5C4GSLnlD1SW/QWlSybMmfTiqcbV69+OBYOSGMs6xhqlFBa7nWtHdiodMOb2wCHvxbMbdj7kVP6tF4jINKHf4DmhbeEOMqVsEKmN5y2Y+p93iOi8MsREuk3Ffm03RITFWMyISEwaNOTt4NbQcTpIHbPO8iChEFOcbawrR796+eWISESpBj5eRKDqDxwPX/XyrND2WCdiJAAY1u8I2/87f8U8/8T7B7jHPfpFxl2/P/IDAA7r3VsBDtqTnjueiNfFO6cmZSO0JIJkwDgYcUfxtaf8B54BcPv8e+zFTUTo9/t3TnK3292sNqvPnWqF8ul7c7vBx1/OkaGwSTCU2So1EWk5Vpsh1K714wPG3/XaoWZA9oSdN8+DbsaNXD55mefdhs2h3lpK2TgA8qTU9doNoTO85w+a7kWUuJ+VUdMUHAZ4QH/Kc/Fz9RsjA+NNCWxCkCSA5ba1155380kXY76nwed2M0CEYizmBhOXj9zw+jsNVdFTkqDrmHKXMOAo63c02L+fv3zeO2NuPrd82TLNBa6samfsLr8KLl6+bJn22KVXjqj5reHWqJbUsyk3QETSyDnZc0yfdj35hno3uPnuseA+t5sTAENE8ng8IvNK5TOmMnyaIu/exuD2e+R/duxwys+++ZDXhwq1VIvtrHdLh9FkiBbkfjDgybFjVJdLOdw6PACkZ2gmm+Ppf3ombf+pZkw4nvhLM08i0g3AlFadWs0dN3/c1Yi9G1wAyu1uN3n8fgn7oAUJAMtcLu4NBHQi4k8Ovey16t9qro0kdJ3xJjAgOkl7joWOO7/reVdPLF+UCWxSXS7F+0VAf/ySy1/b8ev2GyLJpM7Yn1suE4AEQcyeY9W7Htf21mGvv/kKSMgqC2jXTgjAASZfcfWYDSs2T0pqmkg3C228OIyU0mGzsK69Ow4Y9tLr83cPyiJQWbrHIQTmB4q0qu1HSoNitACLGK4evPp0xCDAgdXhztB6bp9Pfnz3Qwss22ovDOnJrHdLIhBWzrlWkLfM8bz37I8RE2WpOhiHnNnZHVha2stQXr5Me+XOG4ev/WrDi8F9RNERkTBxhee1d6459sxutw8te/bzTBSyC1xKJnYCimFng5zKAEAg3afmP7Of6Pr5tC9fq9vUUBzTsq9BkUntctotSqcT299w6ytvTM9MxszPZ6/+p7dq1fbx4XhCZ3yvUYAkBYHdasKcds7pV4w87/6uJXdsTb/HVJfrz6tWAMALAQHpSfvN+xM6fvH2iqe2ra1xRxNJmW3JNwKQXCLaWptXT1z08Unpxk70x/sppf7khTdcsGrtKIrGzuGSnMgYSCKQCq/iBflzQq5THr2kf8n6/VXuCleKAZk3+pFyx5Ydt9THozrLoud6SkaSRkIGBflb9Vvdpwz4xz82HYhJ2tJAAIA3R9587povf/8kWB8FVBjbFzNBRIIRcovDBAUdnL7OPds9f9lDU5bs3oXrz9csdjx3xYvXV2+uGR+pjbVuolcRSJJmN5sM7f/RZtK/3njn/tKTUzZypkbHtOHX3LhuadWrocjOmt37UjYiKaVZMXKj07CtdYe8Z0uuP33GSYPuWQ/6X2eDYlLgg0n3Hfvjoh+vqt0WvDUZ1PITQhPZlhsAAJBSCrvFzAu65l45aubMd3fNd1RdquJd5NXnqc+qlg1byyAchbjQQEiSaa8kKgzRajRD0mmtpWOPLL3g7htmNlW5p5WWGoaVl2sfjX/qfue6qkeCkYgGWZasSGXBSOI5Ti1++vFnDx52zbeHEwOyJ+D8Z0Z3XzxzxeLgjkhr2C2xdh+QUhIzMQ4GqwJGu3FNTmvrEm6x/GRi2m8FR+ZDvC6CNVsShULox9dujfRL1Cc7JbQkAIMmFZ2RRLqFG5RW3fJmPDBv9hUPaGcpXggIt9vN/H6/eOOeawb+uHjrnHAwKpCzrOOJiUggAbeYTMCtPGJ1mhY78m1LQ3XRtds31keO7t25QzwU7xQNJs4I10dPhoRQ4poGxKCp1aYEJ+TOtrbl3s/mnVpWVibLvF7CdJSdx+8Xc8Y9fm/e5prHGurrBXEOkGppt3MclIp21Q1EBiXHCcmTup/T/84bv8hWuTO03vxHp7ota9b5oqGQTqmJmW0MiG51OJVo987/vGjs7f8+HBq3Ngbl63mrPw3tiLYmBSVmp9QAAIwxBA2ESIQF4+HkMbEdsWMYZ8A4wqbVtUBEIDQJUhAkdB0ISaR7kWStFAQkjMiUnCLHt/d98NQN9+MsVkaVwotIbgDwAwBDkzXddkUCQdalGBCRAwJFkwkBcbDFGxJ9Q1WRvgQEJsZh4/dbgKSEpC5AkARAEMiQYVOcKghAugST3QLte7a5AxF1n9vNEUD6fG7u8fjFvHdmn6B8snhiKBjSSeF7pNsQAAHBoDEULBTm8vs1r3+1YcPxZ3TqFIdGSqalm8Hqs1+fcbph0fLpiVBEyCbQeiSlnmNzKA0dCu8fOPb2fx+ODMiewKwOyzsmswJZJqTuBuSMIRIDmZC6Hk0m9XA0IRqCUREMx0Q0mdTjQtOBgUwzB1mzEIQgmUDuKLBt7D/i1CGIXeOqqu50QXv8fuF2u/k1T73iKzgq/0G7xWQgQU294YiICnIkHaSIakk9pml6QtdENJnQY7qmy9SEpKbKD5Cqj2I1mnh+x5zHh02Z/tWfVlh/6tto+ep7zAldEQyhMWVDAh4nqTsS4ojwS+9ehQBUoap7nWikqgw9HjH//fldzN+smknBsDXbSqgAKcIgx2JVQh0LXh/4yKhJFSkG5LBXagAAdt+cWffld3L67EazQtRkxdj5Pbv0aEm1J2KMI6Z+B01NaECQpElwtrbFzxh63MW9+z9QlQls2vVjfr9fqi6XMtr33sO5nZ3vWA0mgyTany0SIdV+IyUv4q6y71cnBALSTVwx5HRyfnXvrPfGucHN3Sn2KBWi6/eLCklmqKk/L5ZMULbmGSKCFDrpNcEhAHtvQEpECF6A74lsfOGSWcaa+nYJkFkVqgRI1fGzK0Yl3Mr5uWXi6FIfuXlxZeVha1PvDiY1ye6bPeZ6a6FhoREUAxAc2hmZqtlHOTk21qVXuyv6/+uJ7/aR2kVllZVC6pKNef/pm/I6Wz8xIVNo/5S72UBAOpdMcRSYVl8y+pzLEFH0UHtkYpsB1BR/rs9b2IkJvUgnmfUqigAsqeuImnY0ERn2cl+wsriYI3tIbr57wv+Za4MnhYWeFd+eHoAwM84TeY61WHqZuwRRX9XCdUCaGyxVuqNPTCZjg5xFpk8tBoNBStIOUYiwlLqknBwr77xyx9sunvzaB41VTEJESo2ha/y+Dy8d4ijKX25CRaFDNEEzfL811/ybvXNuvx6ue6pUVcU9BQahyWRGReFN6V5LAEhSAhgNOVsBjOm/+aenNa20VCkJBPSP7nn4cWd1/SUNyXjWlWgJgDgRk3nOkOxfMqTfccfV+ty+v+yWhzuY15sq3OINBOLqwpcG53XOnWs3mQxClxKyaPnQfCBBgpjDaWUdT+py+7CXXn0x2zJgmTEg3hA/67ZLLsjrYP3EoigGKUjLpkpTswCBSJJmQkXJb+9cPWhEz3NGTf/3xj2ZUFCWksks2Q6paVFkDLItg4AIkisKUDyxsS1AlNL1ITPvV6iqMqy8XJv70ORSR3XD6GCTuGogLkgqTjskundx9x9y3qoKteXq67UkGEAqyVNVgSF2jT+44MNBXU4umpSTY2WoAyMivUWVA4GklDqTjOe2ttUd3af94FunvTy1qbXtUmNQ2QWXXlczbsELQ9oe3foDh9VkELqkg1DaS0hNos1kMhR1bz3/puevLe7jeWazey90XKbx6JkXnFlFdttPJuT79NzuCiIiI1cIc53LEZEqXX+0ssvQcB++MP1C82+bpkZCIUFN4NtRSmF12Hm8W4fhg0aVfvy/QOvtDTsPEl5vqqC6ntTZra+9df9RZ3a7sFX7nJU2o1GRmsQWUPBUJwGd0GE2K2275X97+tB/nHnjs9Pn7G/BxoxyI3aNj5kz65L2JxRNsNstjEtgUsoDKimwFwghpeQSeW5rW6JLr6Kx93/0wYB2R3uqM70x93ZhpapyRCTWofAts9XKQFJWqyJKgqSBIzu2y6sAANWFqWafpKqsxOvVZ/tmH2le8cvbIhiWAkEAkSAifZ8vIJ2kTDitdiVSlP/0wPF3lf8vKzXAXg4sbnBzP/gF0QbL5CtH31WzJXhvsiGZF40ngBBE6lyNCPtXvkESkSQJiolzMDkMkY7HtH3i1tf/71FETDZHYZt0pxAEAPnqiFtcG1dUPRerjZ0QSyRBkBTpjKf9LQ6Tkp+AG5Ch2WaE3HY5C489s/3IIfc+twJSRXQazYwhIiwrK8NLysosW4c/8LWpNtgzIoWGCHvzBhJJqeVb7Ma6ToVvXDTp3ut9l6VjZdLBUWd1O+ZIuWTV4rxQvCBCAhTMrp66JAk2gwmq822zBkzxDq0480ylpLJSNNq++jDGXh/srgpW88O7HWc8Mf+WqvXV1+lRrZMW10ETAgQQIIJIZ+cjYGpN32n0EaUqqVKqEzIRcAUQTEYDcCtPFHZpPfPEMzurJXdM+hXgj2ZJzTW4zMpPtMHy9GUjb6neUnc3JKhrPK6BLiUQA8nSEgMCIqT65wDskuxJQCnPHxEQcAYIJkUBZuZgybUEio4seOL219/8SI/rOxeEbOXLxFrMn+7vaVjy3wpjfaggqCUIkUlIVycmIEACAil5vs0BNYW582xPjRta7fFomcpKme/5xDdnsHntlpJgMhZjqXqCjcuARAowIIMSS44Z9sRgxGi6v8//rFIDNLJi7RqVBwBAtNL+1j3PXrD5t7qh4bpwn0RM7wqaBKFLEFKClASAAEIIQETgnAMQAOcMFIUDKCDsuZZVuQX29/9xQc+3i28Z/wuI1CRy+/2yJRKHd+35sm2bzz7Hu+CKqt92XBOqi/aipLQKTYAuZCrYiCQwzgEZgNAJkAg4Y6mXwoA4gtlq/DWn0PFJ1+Pbvv3PR6d+JZICAAD3xnw0hoxSzps17whj5dLJoqbuIqMgEFKAJAKODBhnoFlNcSjIm1L/6JgHPIjJ/z8oX0siu8Lvu4Ztpq8iudgx//HZJ29eu/2Muh3h4xIx/SgSeptwg5Zjdthy9aSmkx6vsuXYg4qBrc5tbfm207FtK/uPevI7RMysalxV1f0uSZwtdp+g3KRA4IXx3X9etvG8hrpgn3BdtEciLjojh7xIXQQTMQG5BRZg3JhARhvNVuNPua2dy3IKbJ9f+fgD3yK2j6a/Gn1uNztg0ykTJYcAFS/NOB82VbmTW7YehSTzwWqrMrTK+5qf3HPWOUPOXwEAsDelJlVl5VVV/OiiIvq5qqpJZtbRRUVUXFaWVRH//wX8fz6qsrNXur56AAAAAElFTkSuQmCC"
    st.markdown(f'''<div style="position:fixed;top:60px;right:24px;z-index:999;opacity:0.85;">
        <img src="data:image/png;base64,{_logo_b64_f}" style="height:32px;" />
    </div>''', unsafe_allow_html=True)
    st.markdown("<style>.block-container { padding-left: 1rem; padding-right: 1rem; max-width: 100%; }</style>", unsafe_allow_html=True)

    if st.sidebar.button("⬅️ Retour Accueil", key="btn_back_formations"):
        st.session_state.page = "accueil"
        st.rerun()

    st.title("🎓 Suivi des formations")

    mode = st.sidebar.radio("Mode", ["📋 Créer fichier de configuration", "📊 Suivi des formations"], key="formations_mode")

    # ── MODE 1 : CRÉER CONFIGURATION ─────────────────────────────
    if mode == "📋 Créer fichier de configuration":
        st.subheader("📋 Configuration de l'équipe")
        st.caption("Renseignez les informations de chaque employé. Téléchargez ensuite le fichier Excel généré — il servira de base pour le suivi des formations.")

        # Nombre d'employés
        n = st.number_input("Nombre d'employés", min_value=1, max_value=30, value=3, step=1, key="nb_emp")

        # Tableau de saisie
        st.markdown("**Données des employés**")

        cols_header = st.columns([2, 2, 1.2, 1.5, 1.8, 2, 2, 1.8, 1.5, 1.2])
        headers = ["Nom", "Prénom", "Taux (%)", "Jours droit\n(à 100%)", "Budget CHF\n(à 100%)", "Date engagement", "Date avenant", "Taux avenant (%)", "Décompte", "Report solde"]
        for col, h in zip(cols_header, headers):
            col.markdown(f"<small><b>{h}</b></small>", unsafe_allow_html=True)

        employees = []
        for i in range(int(n)):
            c = st.columns([2, 2, 1.2, 1.5, 1.8, 2, 2, 1.8, 1.5, 1.2])
            nom         = c[0].text_input("", key=f"nom_{i}", placeholder="Dupont", label_visibility="collapsed")
            prenom      = c[1].text_input("", key=f"prenom_{i}", placeholder="Marie", label_visibility="collapsed")
            taux        = c[2].number_input("", min_value=10, max_value=100, value=100, step=5, key=f"taux_{i}", label_visibility="collapsed")
            jours       = c[3].number_input("", min_value=0.0, max_value=30.0, value=3.0, step=0.5, key=f"jours_{i}", label_visibility="collapsed")
            budget      = c[4].number_input("", min_value=0, max_value=20000, value=1500, step=100, key=f"budget_{i}", label_visibility="collapsed")
            date_eng    = c[5].date_input("", key=f"date_eng_{i}", label_visibility="collapsed")
            date_av     = c[6].date_input("", key=f"date_av_{i}", value=None, label_visibility="collapsed")
            taux_av     = c[7].number_input("", min_value=0, max_value=100, value=0, step=5, key=f"taux_av_{i}", label_visibility="collapsed", help="0 = pas d'avenant")
            decompte    = c[8].selectbox("", ["Civil", "Engagement"], key=f"decompte_{i}", label_visibility="collapsed")
            report      = c[9].checkbox("", key=f"report_{i}", label_visibility="collapsed")
            employees.append({
                "Nom": nom, "Prénom": prenom, "Taux (%)": taux,
                "Jours droit (100%)": jours, "Budget CHF (100%)": budget,
                "Date engagement": str(date_eng),
                "Date avenant": str(date_av) if date_av else "",
                "Taux avenant (%)": taux_av if taux_av > 0 else "",
                "Type décompte": decompte,
                "Report solde": "Oui" if report else "Non",
            })

        if st.button("💾 Générer le fichier Excel de configuration", type="primary"):
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # ── Feuille Employés ──
            ws_emp = wb.active
            ws_emp.title = "Employés"

            header_fill = PatternFill("solid", start_color="6D2B3D")
            header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            data_font   = Font(name="Arial", size=10)
            border_thin = Border(
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD"),
            )
            alt_fill    = PatternFill("solid", start_color="F9EEF1")

            emp_headers = ["Nom", "Prénom", "Taux (%)", "Jours droit (100%)", "Budget CHF (100%)",
                           "Date engagement", "Date avenant", "Taux avenant (%)", "Type décompte", "Report solde",
                           "Jours effectifs", "Budget effectif CHF"]
            col_widths  = [18, 16, 12, 18, 18, 16, 14, 18, 14, 14, 14, 20]

            for ci, (h, w) in enumerate(zip(emp_headers, col_widths), 1):
                cell = ws_emp.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws_emp.column_dimensions[get_column_letter(ci)].width = w
            ws_emp.row_dimensions[1].height = 30

            for ri, emp in enumerate(employees, 2):
                row_fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
                vals = [
                    emp["Nom"], emp["Prénom"], emp["Taux (%)"],
                    emp["Jours droit (100%)"], emp["Budget CHF (100%)"],
                    emp["Date engagement"], emp["Date avenant"],
                    emp["Taux avenant (%)"], emp["Type décompte"], emp["Report solde"],
                    f"=D{ri}*C{ri}/100",
                    f"=E{ri}*C{ri}/100",
                ]
                for ci, val in enumerate(vals, 1):
                    cell = ws_emp.cell(row=ri, column=ci, value=val)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal="center" if ci > 2 else "left")

            ws_emp.freeze_panes = "A2"

            # ── Feuille Formations ──
            ws_form = wb.create_sheet("Formations")
            form_headers = [
                "Employé (Nom Prénom)", "Date", "Nom de la formation", "Organisme",
                "Coût formation CHF", "Coût déplacement CHF", "Coût hébergement CHF", "Coût nourriture CHF",
                "Total CHF", "Payé par", "Remboursé", "Date remboursement", "Notes"
            ]
            form_widths = [22, 12, 30, 20, 18, 20, 20, 18, 12, 14, 12, 18, 25]
            for ci, (h, w) in enumerate(zip(form_headers, form_widths), 1):
                cell = ws_form.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws_form.column_dimensions[get_column_letter(ci)].width = w
            ws_form.row_dimensions[1].height = 30

            # 3 lignes exemple
            example_rows = [
                ["Dupont Marie", "2026-03-01", "Formation IASTM niveau 1", "Institut XY",
                 450, 80, 0, 30, "=E2+F2+G2+H2", "Employé", "Non", "", "Remboursement à planifier"],
                ["Dupont Marie", "2026-04-15", "Congrès physiothérapie", "SSRPM",
                 300, 150, 200, 60, "=E3+F3+G3+H3", "Cabinet", "—", "", ""],
                ["Martin Paul", "2026-05-10", "Formation manuelle avancée", "HES-SO",
                 600, 0, 0, 0, "=E4+F4+G4+H4", "Partagé 50/50", "Partiel", "", "Cabinet a payé 300 CHF"],
            ]
            ex_fill = PatternFill("solid", start_color="FFF3CD")
            for ri, row in enumerate(example_rows, 2):
                for ci, val in enumerate(row, 1):
                    cell = ws_form.cell(row=ri, column=ci, value=val)
                    cell.font = Font(name="Arial", size=10, color="888888", italic=True)
                    cell.fill = ex_fill
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal="center" if ci > 2 else "left")

            ws_form.freeze_panes = "A2"

            # ── Feuille Tableau de bord ──
            ws_db = wb.create_sheet("Tableau de bord")
            ws_db["A1"] = "Ce tableau est généré automatiquement par l'application 36.9° Analytique."
            ws_db["A1"].font = Font(bold=True, italic=True, color="B5546A", name="Arial", size=10)
            ws_db["A2"] = "Importez ce fichier dans le module Formations pour voir le suivi en temps réel."
            ws_db["A2"].font = Font(italic=True, color="888888", name="Arial", size=10)
            ws_db.column_dimensions["A"].width = 65

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                "⬇️ Télécharger le fichier de configuration",
                data=buf,
                file_name="formations_config.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ── MODE 2 : SUIVI DES FORMATIONS ────────────────────────────
    else:
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import math

        st.subheader("📊 Suivi des formations")

        fichier = st.sidebar.file_uploader("📂 Fichier de configuration (.xlsx)", type=["xlsx"], key="formations_file")

        if fichier is None:
            st.info("👈 Chargez votre fichier formations_config.xlsx dans la sidebar pour commencer.")
            st.stop()

        # ── Lecture du fichier ──
        try:
            df_emp   = pd.read_excel(fichier, sheet_name="Employés")
            fichier.seek(0)
            df_form_raw = pd.read_excel(fichier, sheet_name="Formations")
        except Exception as e:
            st.error(f"Erreur lecture fichier : {e}")
            st.stop()

        # Nettoyer colonnes
        df_emp.columns    = [str(c).strip() for c in df_emp.columns]
        df_form_raw.columns = [str(c).strip() for c in df_form_raw.columns]

        # Calculer taux effectif et droits
        ajd = pd.Timestamp(datetime.today().date())

        def taux_effectif(row):
            if pd.notna(row.get("Date avenant")) and str(row.get("Date avenant","")).strip() not in ("","None","NaT"):
                taux_av = row.get("Taux avenant (%)", 0)
                if pd.notna(taux_av) and float(taux_av) > 0:
                    return float(taux_av)
            return float(row.get("Taux (%)", 100))

        df_emp["_taux_eff"] = df_emp.apply(taux_effectif, axis=1)
        df_emp["_jours_eff"] = df_emp["Jours droit (100%)"].astype(float) * df_emp["_taux_eff"] / 100
        df_emp["_budget_eff"] = df_emp["Budget CHF (100%)"].astype(float) * df_emp["_taux_eff"] / 100
        df_emp["_nom_complet"] = df_emp["Nom"].astype(str).str.strip() + " " + df_emp["Prénom"].astype(str).str.strip()

        # Calculer l'année de référence par employé (civil ou engagement)
        def annee_debut(row):
            decompte = str(row.get("Type décompte","Civil")).strip()
            if decompte == "Civil":
                return pd.Timestamp(f"{ajd.year}-01-01")
            else:
                try:
                    eng = pd.Timestamp(str(row["Date engagement"]))
                    debut = pd.Timestamp(f"{ajd.year}-{eng.month:02d}-{eng.day:02d}")
                    if debut > ajd:
                        debut = pd.Timestamp(f"{ajd.year-1}-{eng.month:02d}-{eng.day:02d}")
                    return debut
                except:
                    return pd.Timestamp(f"{ajd.year}-01-01")

        df_emp["_debut_periode"] = df_emp.apply(annee_debut, axis=1)

        # Filtrer les formations (exclure lignes exemple en italique / sans employé valide)
        noms_valides = set(df_emp["_nom_complet"].str.lower().tolist())
        col_emp_form = df_form_raw.columns[0]
        col_date_form = df_form_raw.columns[1]

        df_form = df_form_raw.copy()
        df_form[col_date_form] = pd.to_datetime(df_form[col_date_form], errors="coerce")
        df_form = df_form[df_form[col_emp_form].astype(str).str.strip().str.lower().isin(noms_valides)]
        df_form = df_form.dropna(subset=[col_date_form])

        # ── TABS ──
        tab1, tab2, tab3, tab4 = st.tabs(["➕ Saisir formation", "📊 Tableau de bord", "💰 Paiements", "📅 Soldes & reports"])

        # ── TAB 1 : SAISIE ──
        with tab1:
            st.subheader("➕ Ajouter une formation")
            noms_emp = sorted(df_emp["_nom_complet"].tolist())

            c1, c2 = st.columns(2)
            f_emp   = c1.selectbox("Employé", noms_emp, key="f_emp")
            f_date  = c2.date_input("Date de la formation", key="f_date")
            f_nom   = st.text_input("Nom de la formation", key="f_nom")
            f_org   = st.text_input("Organisme / prestataire", key="f_org")

            st.markdown("**Coûts**")
            cc = st.columns(4)
            f_cout_form  = cc[0].number_input("Formation (CHF)", min_value=0, value=0, step=10, key="f_cf")
            f_cout_dep   = cc[1].number_input("Déplacement (CHF)", min_value=0, value=0, step=10, key="f_cd")
            f_cout_heb   = cc[2].number_input("Hébergement (CHF)", min_value=0, value=0, step=10, key="f_ch")
            f_cout_nour  = cc[3].number_input("Nourriture (CHF)", min_value=0, value=0, step=10, key="f_cn")
            f_total = f_cout_form + f_cout_dep + f_cout_heb + f_cout_nour

            cp = st.columns(3)
            f_paye_par  = cp[0].selectbox("Payé par", ["Employé", "Cabinet", "Partagé 50/50", "Partagé autre"], key="f_pp")
            f_rembourse = cp[1].selectbox("Statut remboursement", ["Non", "Partiel", "Oui", "—"], key="f_rb")
            f_notes     = cp[2].text_input("Notes", key="f_notes")

            st.info(f"💰 **Total : CHF {f_total:,.2f}**")

            if st.button("➕ Ajouter cette formation", type="primary", key="btn_add_form"):
                if not f_nom.strip():
                    st.warning("Veuillez saisir le nom de la formation.")
                else:
                    # Recharger le workbook et ajouter la ligne
                    fichier.seek(0)
                    wb_edit = load_workbook(fichier)
                    ws_f = wb_edit["Formations"]
                    last_row = ws_f.max_row + 1
                    data_font_e = Font(name="Arial", size=10)
                    border_thin_e = Border(
                        left=Side(style="thin", color="DDDDDD"),
                        right=Side(style="thin", color="DDDDDD"),
                        bottom=Side(style="thin", color="DDDDDD"),
                    )
                    new_row = [
                        f_emp, str(f_date), f_nom, f_org,
                        f_cout_form, f_cout_dep, f_cout_heb, f_cout_nour,
                        f_total, f_paye_par, f_rembourse, "", f_notes
                    ]
                    for ci, val in enumerate(new_row, 1):
                        cell = ws_f.cell(row=last_row, column=ci, value=val)
                        cell.font = data_font_e
                        cell.border = border_thin_e

                    buf = io.BytesIO()
                    wb_edit.save(buf)
                    buf.seek(0)
                    st.success(f"✅ Formation '{f_nom}' ajoutée. Téléchargez le fichier mis à jour ci-dessous.")
                    st.download_button(
                        "⬇️ Télécharger le fichier mis à jour",
                        data=buf,
                        file_name="formations_config.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_updated"
                    )

        # ── TAB 2 : TABLEAU DE BORD ──
        with tab2:
            st.subheader("📊 Tableau de bord — Droits et consommation")

            col_total = df_form.columns[8] if len(df_form.columns) > 8 else None
            col_paye  = df_form.columns[9] if len(df_form.columns) > 9 else None

            dashboard_rows = []
            for _, emp in df_emp.iterrows():
                nom = emp["_nom_complet"]
                debut = emp["_debut_periode"]
                jours_droit = emp["_jours_eff"]
                budget_droit = emp["_budget_eff"]
                report_solde = str(emp.get("Report solde","Non")).strip() == "Oui"

                # Formations de l'employé sur la période courante
                df_e = df_form[df_form[col_emp_form].astype(str).str.strip().str.lower() == nom.lower()]
                df_e_periode = df_e[df_e[col_date_form] >= debut] if not df_e.empty else df_e

                nb_formations = len(df_e_periode)
                # Chaque ligne = 1 jour de formation (simplification) — nb_jours = nombre de dates distinctes
                jours_pris = len(df_e_periode[col_date_form].dt.date.unique()) if not df_e_periode.empty else 0

                if col_total and not df_e_periode.empty:
                    budget_consomme = pd.to_numeric(df_e_periode[col_total], errors="coerce").sum()
                else:
                    budget_consomme = 0.0

                jours_restants  = jours_droit - jours_pris
                budget_restant  = budget_droit - budget_consomme
                pct_jours  = int(jours_pris / jours_droit * 100) if jours_droit > 0 else 0
                pct_budget = int(budget_consomme / budget_droit * 100) if budget_droit > 0 else 0

                alert = ""
                if pct_jours >= 100 or pct_budget >= 100:
                    alert = "🔴"
                elif pct_jours >= 80 or pct_budget >= 80:
                    alert = "🟠"
                else:
                    alert = "🟢"

                dashboard_rows.append({
                    "": alert,
                    "Employé": nom,
                    "Taux (%)": f"{emp['_taux_eff']:.0f}%",
                    "Jours droit": f"{jours_droit:.1f}j",
                    "Jours pris": f"{jours_pris}j",
                    "Jours restants": f"{jours_restants:.1f}j",
                    "Budget droit": f"CHF {budget_droit:,.0f}",
                    "Consommé": f"CHF {budget_consomme:,.0f}",
                    "Solde budget": f"CHF {budget_restant:,.0f}",
                    "Nb formations": nb_formations,
                    "Report": "✅" if report_solde else "—",
                })

            if dashboard_rows:
                st.dataframe(pd.DataFrame(dashboard_rows), use_container_width=True, hide_index=True)
                st.caption("🟢 < 80% utilisé · 🟠 ≥ 80% · 🔴 dépassé")
            else:
                st.info("Aucune donnée à afficher. Ajoutez des employés dans la configuration.")

        # ── TAB 3 : PAIEMENTS ──
        with tab3:
            st.subheader("💰 Suivi des paiements")

            if df_form.empty:
                st.info("Aucune formation enregistrée.")
            else:
                col_rembourse = df_form.columns[10] if len(df_form.columns) > 10 else None
                col_paye_par  = df_form.columns[9] if len(df_form.columns) > 9 else None
                col_total_p   = df_form.columns[8] if len(df_form.columns) > 8 else None
                col_nom_form  = df_form.columns[2] if len(df_form.columns) > 2 else None

                # Filtre employé
                choix_emp = st.selectbox("Filtrer par employé", ["Tous"] + sorted(df_emp["_nom_complet"].tolist()), key="filtre_paiement")
                df_show = df_form.copy()
                if choix_emp != "Tous":
                    df_show = df_show[df_show[col_emp_form].astype(str).str.strip().str.lower() == choix_emp.lower()]

                # A rembourser par le cabinet (payé par employé + non remboursé)
                df_a_rembourser = df_show[
                    df_show[col_paye_par].astype(str).str.strip().str.lower().isin(["employé","employe"]) &
                    df_show[col_rembourse].astype(str).str.strip().str.lower().isin(["non", "partiel"])
                ] if col_paye_par and col_rembourse else pd.DataFrame()

                if not df_a_rembourser.empty:
                    total_a_rembourser = pd.to_numeric(df_a_rembourser[col_total_p], errors="coerce").sum()
                    st.error(f"⚠️ **À rembourser à l'employé : CHF {total_a_rembourser:,.2f}**")
                    cols_display = [col_emp_form, col_date_form, col_nom_form, col_total_p, col_paye_par, col_rembourse]
                    cols_display = [c for c in cols_display if c]
                    st.dataframe(df_a_rembourser[cols_display], use_container_width=True, hide_index=True)
                else:
                    st.success("✅ Aucun remboursement en attente.")

                st.markdown("---")
                st.subheader("Toutes les formations")
                cols_all = [col_emp_form, col_date_form, col_nom_form, col_total_p, col_paye_par, col_rembourse]
                cols_all = [c for c in cols_all if c and c in df_show.columns]
                st.dataframe(df_show[cols_all], use_container_width=True, hide_index=True)

        # ── TAB 4 : SOLDES & REPORTS ──
        with tab4:
            st.subheader("📅 Soldes de droits & reports")
            st.caption(f"Période de référence : du début de l'année (civil) ou de l'anniversaire d'engagement jusqu'au {ajd.strftime('%d.%m.%Y')}")

            solde_rows = []
            for _, emp in df_emp.iterrows():
                nom = emp["_nom_complet"]
                debut = emp["_debut_periode"]
                jours_droit = emp["_jours_eff"]
                budget_droit = emp["_budget_eff"]
                report_actif = str(emp.get("Report solde","Non")).strip() == "Oui"
                decompte = str(emp.get("Type décompte","Civil")).strip()

                df_e = df_form[df_form[col_emp_form].astype(str).str.strip().str.lower() == nom.lower()]
                df_e_periode = df_e[df_e[col_date_form] >= debut] if not df_e.empty else df_e
                jours_pris = len(df_e_periode[col_date_form].dt.date.unique()) if not df_e_periode.empty else 0
                budget_consomme = pd.to_numeric(df_e_periode[df_form.columns[8]], errors="coerce").sum() if not df_e_periode.empty and len(df_form.columns) > 8 else 0.0

                solde_jours   = jours_droit - jours_pris
                solde_budget  = budget_droit - budget_consomme

                # Fin de période
                if decompte == "Civil":
                    fin_periode = pd.Timestamp(f"{ajd.year}-12-31")
                else:
                    try:
                        eng = pd.Timestamp(str(emp["Date engagement"]))
                        fin_periode = debut + pd.DateOffset(years=1) - pd.DateOffset(days=1)
                    except:
                        fin_periode = pd.Timestamp(f"{ajd.year}-12-31")
                jours_restants_periode = (fin_periode - ajd).days

                solde_rows.append({
                    "Employé": nom,
                    "Type décompte": decompte,
                    "Fin de période": fin_periode.strftime("%d.%m.%Y"),
                    "Jours restants dans la période": f"{jours_restants_periode}j",
                    "Solde jours": f"{solde_jours:.1f}j",
                    "Solde budget": f"CHF {solde_budget:,.0f}",
                    "Report activé": "✅ Oui" if report_actif else "❌ Non",
                    "Solde reporté (N+1)": f"{solde_jours:.1f}j / CHF {solde_budget:,.0f}" if report_actif and solde_jours > 0 else "—",
                })

            if solde_rows:
                st.dataframe(pd.DataFrame(solde_rows), use_container_width=True, hide_index=True)
                st.caption("Le report de solde est une option indicative — activez-le par employé dans le fichier de configuration.")
            else:
                st.info("Aucun employé configuré.")
